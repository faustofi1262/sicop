import os
import tempfile
import uuid
from flask import Blueprint, render_template, request, session
from app.decorators import login_required
from openpyxl import load_workbook
from io import BytesIO
from decimal import Decimal, InvalidOperation
from app.database import get_connection
import unicodedata
from flask import (
    Blueprint,
    render_template,
    request,
    session,
    redirect,
    url_for,
    flash
)
def normalizar_texto(valor):

    if valor is None:
        return ""

    texto = str(valor).strip().upper()

    texto = unicodedata.normalize(
        "NFKD",
        texto
    )

    texto = "".join(
        caracter
        for caracter in texto
        if not unicodedata.combining(caracter)
    )

    return texto


def decimal_seguro(valor):

    if valor in (None, ""):
        return Decimal("0.00")

    try:

        if isinstance(valor, Decimal):
            return valor

        # Si Excel ya entregó un número real
        if isinstance(valor, (int, float)):
            return Decimal(str(valor))

        texto = str(valor).strip()

        texto = texto.replace("$", "")
        texto = texto.replace(" ", "")

        # ==========================================
        # FORMATO DEL PAC:
        # 1.234.567,89  -> 1234567.89
        # ==========================================
        if "," in texto:

            texto = texto.replace(".", "")
            texto = texto.replace(",", ".")

        return Decimal(texto)

    except (InvalidOperation, ValueError, TypeError):

        return Decimal("0.00")

def valor_texto(fila, indice):

    if indice is None:
        return None

    if indice >= len(fila):
        return None

    valor = fila[indice]

    if valor is None:
        return None

    texto = str(valor).strip()

    return texto if texto else None
from openpyxl import load_workbook

planificacion = Blueprint(
    "planificacion",
    __name__,
    url_prefix="/planificacion"
)


@planificacion.route("/")
@login_required()
def inicio():
    return "Módulo de Planificación funcionando"
@planificacion.route("/pac")
@login_required()
def pac_gestion():
    return render_template(
        "planificacion/pac_gestion.html"
    )
@planificacion.route("/pac/vista-previa", methods=["POST"])
@login_required()
def pac_vista_previa():

    archivo = request.files.get("archivo_pac")

    if not archivo or archivo.filename == "":
        return "No se seleccionó archivo", 400

    # ==========================================
    # DATOS INGRESADOS EN EL FORMULARIO
    # ==========================================
    anio = request.form.get("anio")
    numero_reforma = request.form.get("numero_reforma")
    fecha_reforma = request.form.get("fecha_reforma")
    descripcion = request.form.get("descripcion")
    estado = request.form.get("estado")

    # ==========================================
    # LEER EXCEL EN MEMORIA
    # NO SE GUARDA NINGÚN ARCHIVO
    # ==========================================
    contenido = archivo.read()

    # ==========================================
    # GUARDAR COPIA TEMPORAL DEL EXCEL
    # PARA CONFIRMAR LA IMPORTACIÓN DESPUÉS
    # ==========================================
    nombre_temporal = f"pac_{uuid.uuid4().hex}.xlsx"

    ruta_temporal = os.path.join(
        tempfile.gettempdir(),
        nombre_temporal
    )

    with open(ruta_temporal, "wb") as archivo_temporal:
        archivo_temporal.write(contenido)

    session["pac_temp_path"] = ruta_temporal
    session["pac_temp_nombre"] = archivo.filename
        
    libro = load_workbook(
        BytesIO(contenido),
        data_only=True,
        read_only=True
    )

    hoja = libro.active

    # ==========================================
    # BUSCAR AUTOMÁTICAMENTE FILA DE ENCABEZADOS
    # ==========================================
    hoja.reset_dimensions()
    fila_encabezados = None
    for numero_fila, fila in enumerate(
        hoja.iter_rows(values_only=True),
        start=1
    ):

        valores = [
            str(valor).strip().upper()
            if valor is not None else ""
            for valor in fila
        ]

        # Buscamos varios campos característicos del PAC
        tiene_programa = any(
            "PROGRAMA" in valor
            for valor in valores
        )

        tiene_fuente = any(
            "FUENTE" in valor
            for valor in valores
        )

        tiene_renglon = any(
            "RENGLON" in valor or "RENGLÓN" in valor
            for valor in valores
        )

        if (
            tiene_programa
            and tiene_fuente
            and tiene_renglon
        ):
            fila_encabezados = numero_fila
            break

    if not fila_encabezados:
        libro.close()
        return (
            "No se pudo identificar la fila "
            "de encabezados del PAC.",
            400
        )

    # ==========================================
    # OBTENER ENCABEZADOS
    # ==========================================
    encabezados = []

    for celda in hoja[fila_encabezados]:

        valor = celda.value

        encabezados.append(
            str(valor).strip()
            if valor is not None else ""
        )

    # ==========================================
    # MAPEAR COLUMNAS DEL PAC
    # ==========================================
    mapa_columnas = {
        nombre.strip().upper(): indice
        for indice, nombre in enumerate(encabezados)
        if nombre
    }

    idx_ejercicio = mapa_columnas.get("EJERCICIO")
    idx_programa = mapa_columnas.get("PROGRAMA")
    idx_partida = mapa_columnas.get("RENGLON")
    idx_fuente = mapa_columnas.get("FUENTE")
    idx_total = mapa_columnas.get("TOTAL")
    idx_unidad = mapa_columnas.get("DEPARTAMENTO REQUIRENTE")
  

    # ==========================================
    # ANALIZAR TODO EL PAC
    # SIN GUARDAR EN BASE DE DATOS
    # ==========================================
    registros = []

    total_registros = 0
    total_pac = Decimal("0.00")

    unidades = set()
    programas = set()
    partidas = set()
    fuentes = set()

    for fila in hoja.iter_rows(
        min_row=fila_encabezados + 1,
        values_only=True
    ):

        # Ignorar filas totalmente vacías
        if not any(
            valor is not None and str(valor).strip() != ""
            for valor in fila
        ):
            continue

        # Ignorar filas que no correspondan al ejercicio cargado
        if idx_ejercicio is not None:

            ejercicio = str(
                fila[idx_ejercicio] or ""
            ).strip()

            if ejercicio != str(anio):
                continue

        total_registros += 1

        if idx_total is not None:
            total_pac += decimal_seguro(
                fila[idx_total]
            )

        if idx_unidad is not None:
            valor = str(
                fila[idx_unidad] or ""
            ).strip()

            if valor:
                unidades.add(valor)

        if idx_partida is not None:
            valor = str(
                fila[idx_partida] or ""
            ).strip()

            if valor:
                partidas.add(valor)

        if idx_fuente is not None:
            valor = str(
                fila[idx_fuente] or ""
            ).strip()

            if valor:
                fuentes.add(valor)
        if idx_programa is not None:
            valor = str(
                fila[idx_programa] or ""
            ).strip()

            if valor:
                programas.add(valor)
        # Solo mostramos los primeros 10 en la vista previa
        if len(registros) < 10:
            registros.append(list(fila))


    total_pac = total_pac.quantize(
        Decimal("0.01")
    )

    total_unidades = len(unidades)
    total_programas = len(programas)
    total_partidas = len(partidas)
    total_fuentes = len(fuentes)

    return render_template(
        "planificacion/pac_vista_previa.html",
        nombre_archivo=archivo.filename,
        anio=anio,
        numero_reforma=numero_reforma,
        fecha_reforma=fecha_reforma,
        descripcion=descripcion,
        estado=estado,
        fila_encabezados=fila_encabezados,
        encabezados=encabezados,
        registros=registros,

        total_registros=total_registros,
        total_pac=total_pac,
        total_unidades=total_unidades,
        total_programas=total_programas,
        total_partidas=total_partidas,
        total_fuentes=total_fuentes
    )
@planificacion.route("/pac/importar", methods=["POST"])
@login_required()
def pac_importar():

    ruta_temporal = session.get("pac_temp_path")

    if not ruta_temporal:

        flash(
            "No existe un archivo PAC pendiente de importar.",
            "danger"
        )

        return redirect(
            url_for("planificacion.pac_gestion")
        )

    if not os.path.exists(ruta_temporal):

        flash(
            "El archivo temporal del PAC ya no está disponible. "
            "Realice nuevamente la vista previa.",
            "danger"
        )

        return redirect(
            url_for("planificacion.pac_gestion")
        )


    # ==========================================
    # DATOS DE LA VERSIÓN
    # ==========================================
    anio = request.form.get("anio")
    numero_reforma = (
        request.form.get("numero_reforma") or ""
    ).strip()

    fecha_reforma = (
        request.form.get("fecha_reforma") or None
    )

    descripcion = (
        request.form.get("descripcion") or None
    )

    estado = (
        request.form.get("estado") or "HISTORICO"
    ).strip().upper()

    nombre_archivo = session.get(
        "pac_temp_nombre",
        "PAC.xlsx"
    )


    if not anio or not numero_reforma:

        flash(
            "Faltan los datos de año o número de reforma.",
            "danger"
        )

        return redirect(
            url_for("planificacion.pac_gestion")
        )


    # ==========================================
    # ABRIR EXCEL TEMPORAL
    # ==========================================
    libro = load_workbook(
        ruta_temporal,
        data_only=True,
        read_only=True
    )

    hoja = libro.active

    hoja.reset_dimensions()

    fila_encabezados = None


    # ==========================================
    # BUSCAR FILA DE ENCABEZADOS
    # ==========================================
    for numero_fila, fila in enumerate(
        hoja.iter_rows(values_only=True),
        start=1
    ):

        valores = [
            normalizar_texto(valor)
            for valor in fila
        ]

        tiene_programa = any(
            "PROGRAMA" == valor
            for valor in valores
        )

        tiene_fuente = any(
            "FUENTE" == valor
            for valor in valores
        )

        tiene_renglon = any(
            "RENGLON" == valor
            for valor in valores
        )

        if (
            tiene_programa
            and tiene_fuente
            and tiene_renglon
        ):
            fila_encabezados = numero_fila
            break


    if not fila_encabezados:

        libro.close()

        flash(
            "No fue posible identificar los encabezados del PAC.",
            "danger"
        )

        return redirect(
            url_for("planificacion.pac_gestion")
        )


    # ==========================================
    # OBTENER ENCABEZADOS
    # ==========================================
    encabezados = []

    for celda in hoja[fila_encabezados]:

        encabezados.append(
            normalizar_texto(celda.value)
        )


    # ==========================================
    # BUSCAR POSICIÓN DE LAS COLUMNAS
    # ==========================================
    def indice_columna(nombre):

        nombre = normalizar_texto(nombre)

        try:
            return encabezados.index(nombre)
        except ValueError:
            return None


    idx_ejercicio = indice_columna(
        "EJERCICIO"
    )

    idx_programa = indice_columna(
        "PROGRAMA"
    )

    idx_partida = indice_columna(
        "RENGLON"
    )

    idx_fuente = indice_columna(
        "FUENTE"
    )

    idx_cpc = indice_columna(
        "CODIGO CATEGORIA CPC A NIVEL 9"
    )

    idx_tipo_compra = indice_columna(
        "TIPO COMPRA (Bien, obras, servicio o consultoría)"
    )

    idx_descripcion = indice_columna(
        "DETALLE DEL PRODUCTO (Descripción de la contratación)"
    )

    idx_cantidad = indice_columna(
        "CANTIDAD ANUAL"
    )

    idx_valor_unitario = indice_columna(
        "COSTO UNITARIO (Dólares)"
    )

    idx_procedimiento = indice_columna(
        "PROCEDIMIENTO SUGERIDO (son los procedimientos de contratación)"
    )

    idx_tipo_regimen = indice_columna(
        "TIPO DE RÉGIMEN (común, especial)"
    )

    idx_tipo_presupuesto = indice_columna(
        "TIPO DE PRESUPUESTO (proyecto de inversión, gasto corriente)"
    )

    idx_subtotal = indice_columna(
        "SUBTOTAL"
    )

    idx_iva = indice_columna(
        "IVA"
    )

    idx_total = indice_columna(
        "TOTAL"
    )

    idx_codigo_necesidad = indice_columna(
        "CÓDIGO NECESIDAD"
    )

    idx_unidad = indice_columna(
        "DEPARTAMENTO REQUIRENTE"
    )

    idx_departamento_principal = indice_columna(
        "DEPARTAMENTO PRINCIPAL"
    )

    idx_bloque = indice_columna(
        "BLOQUE"
    )


    # ==========================================
    # VALIDAR COLUMNAS CRÍTICAS
    # ==========================================
    columnas_criticas = {
        "PROGRAMA": idx_programa,
        "RENGLON": idx_partida,
        "FUENTE": idx_fuente,
        "TOTAL": idx_total,
        "DEPARTAMENTO REQUIRENTE": idx_unidad
    }

    faltantes = [
        nombre
        for nombre, indice
        in columnas_criticas.items()
        if indice is None
    ]

    if faltantes:

        libro.close()

        flash(
            "El PAC no contiene estas columnas necesarias: "
            + ", ".join(faltantes),
            "danger"
        )

        return redirect(
            url_for("planificacion.pac_gestion")
        )


    # ==========================================
    # CONEXIÓN / TRANSACCIÓN
    # ==========================================
    conn = get_connection()
    cur = conn.cursor()

    total_importados = 0
    print("=== INICIO IMPORTACION PAC ===")
    print("AÑO:", anio)
    print("REFORMA:", numero_reforma)
    print("ESTADO:", estado)
    print("ARCHIVO:", ruta_temporal)

    try:

        # ======================================
        # EVITAR IMPORTAR DOS VECES
        # LA MISMA REFORMA
        # ======================================
        cur.execute("""
            SELECT id
            FROM pac_versiones
            WHERE anio = %s
              AND numero_reforma = %s
        """, (
            int(anio),
            numero_reforma
        ))

        existe = cur.fetchone()

        if existe:

            raise ValueError(
                f"La reforma {numero_reforma} del año "
                f"{anio} ya fue importada."
            )


        # ======================================
        # SI LA NUEVA ES VIGENTE,
        # LA ANTERIOR PASA A HISTÓRICO
        # ======================================
        if estado == "VIGENTE":

            cur.execute("""
                UPDATE pac_versiones
                SET estado = 'HISTORICO'
                WHERE anio = %s
                  AND estado = 'VIGENTE'
            """, (
                int(anio),
            ))


        # ======================================
        # CREAR CABECERA DE LA VERSIÓN
        # ======================================
        cur.execute("""
            INSERT INTO pac_versiones (
                anio,
                numero_reforma,
                descripcion,
                fecha_reforma,
                estado,
                nombre_archivo,
                usuario_id
            )
            VALUES (
                %s, %s, %s, %s, %s, %s, %s
            )
            RETURNING id
        """, (
            int(anio),
            numero_reforma,
            descripcion,
            fecha_reforma,
            estado,
            nombre_archivo,
            session.get("user_id")
        ))

        pac_version_id = cur.fetchone()[0]

        print("VERSION PAC CREADA:", pac_version_id)
        # ======================================
        # IMPORTAR DETALLE DEL PAC
        # ======================================
        for fila in hoja.iter_rows(
            min_row=fila_encabezados + 1,
            values_only=True
        ):

            # Ignorar fila completamente vacía
            if not any(
                valor is not None
                and str(valor).strip() != ""
                for valor in fila
            ):
                continue


            # Solo ejercicio seleccionado
            if idx_ejercicio is not None:

                ejercicio = valor_texto(
                    fila,
                    idx_ejercicio
                )

                if ejercicio != str(anio):
                    continue


            cur.execute("""
                INSERT INTO pac_detalle (
                    pac_version_id,
                    unidad,
                    departamento_principal,
                    programa,
                    partida,
                    fuente,
                    cpc,
                    tipo_compra,
                    descripcion,
                    procedimiento_sugerido,
                    cantidad,
                    valor_unitario,
                    subtotal,
                    iva,
                    total,
                    codigo_necesidad,
                    tipo_regimen,
                    tipo_presupuesto,
                    bloque
                )
                VALUES (
                    %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s,
                    %s, %s, %s, %s
                )
            """, (
                pac_version_id,

                valor_texto(
                    fila,
                    idx_unidad
                ),

                valor_texto(
                    fila,
                    idx_departamento_principal
                ),

                valor_texto(
                    fila,
                    idx_programa
                ),

                valor_texto(
                    fila,
                    idx_partida
                ),

                valor_texto(
                    fila,
                    idx_fuente
                ),

                valor_texto(
                    fila,
                    idx_cpc
                ),

                valor_texto(
                    fila,
                    idx_tipo_compra
                ),

                valor_texto(
                    fila,
                    idx_descripcion
                ),

                valor_texto(
                    fila,
                    idx_procedimiento
                ),

                decimal_seguro(
                    fila[idx_cantidad]
                    if idx_cantidad is not None
                    else 0
                ),

                decimal_seguro(
                    fila[idx_valor_unitario]
                    if idx_valor_unitario is not None
                    else 0
                ),

                decimal_seguro(
                    fila[idx_subtotal]
                    if idx_subtotal is not None
                    else 0
                ),

                decimal_seguro(
                    fila[idx_iva]
                    if idx_iva is not None
                    else 0
                ),

                decimal_seguro(
                    fila[idx_total]
                ),

                valor_texto(
                    fila,
                    idx_codigo_necesidad
                ),

                valor_texto(
                    fila,
                    idx_tipo_regimen
                ),

                valor_texto(
                    fila,
                    idx_tipo_presupuesto
                ),

                valor_texto(
                    fila,
                    idx_bloque
                )
            ))

            total_importados += 1


        # ======================================
        # TODO SALIÓ BIEN
        # ======================================
        conn.commit()

        libro.close()


        # ======================================
        # BORRAR ARCHIVO TEMPORAL
        # SOLO DESPUÉS DEL COMMIT
        # ======================================
        try:

            os.remove(ruta_temporal)

        except OSError:

            pass


        session.pop(
            "pac_temp_path",
            None
        )

        session.pop(
            "pac_temp_nombre",
            None
        )


        flash(
            f"✅ PAC importado correctamente. "
            f"{total_importados:,} registros cargados.",
            "success"
        )

        return redirect(
            url_for(
                "planificacion.pac_gestion"
            )
        )


    except Exception as e:

        conn.rollback()

        try:
            libro.close()
        except Exception:
            pass
        print("ERROR IMPORTANDO PAC:", repr(e))
        flash(
            f"❌ No fue posible importar el PAC: {e}",
            "danger"
        )

        return redirect(
            url_for(
                "planificacion.pac_gestion"
            )
        )


    finally:

        cur.close()
        conn.close()
@planificacion.route("/pac/historial")
@login_required()
def pac_historial():

    conn = get_connection()
    cur = conn.cursor()

    try:

        cur.execute("""
            SELECT
                pv.id,
                pv.anio,
                pv.numero_reforma,
                pv.fecha_reforma,
                pv.estado,
                pv.nombre_archivo,
                pv.fecha_carga,
                COUNT(pd.id) AS total_registros,
                COALESCE(SUM(pd.subtotal), 0) AS subtotal,
                COALESCE(SUM(pd.iva), 0) AS iva,
                COALESCE(SUM(pd.total), 0) AS total_pac
            FROM pac_versiones pv
            LEFT JOIN pac_detalle pd
                ON pd.pac_version_id = pv.id
            GROUP BY
                pv.id,
                pv.anio,
                pv.numero_reforma,
                pv.fecha_reforma,
                pv.estado,
                pv.nombre_archivo,
                pv.fecha_carga
            ORDER BY
                pv.anio DESC,
                pv.id DESC
        """)

        versiones = cur.fetchall()

        return render_template(
            "planificacion/pac_historial.html",
            versiones=versiones
        )

    finally:
        cur.close()
        conn.close()
@planificacion.route("/pac/version/<int:version_id>")
@login_required()
def pac_version_detalle(version_id):

    unidad = request.args.get("unidad", "").strip()
    programa = request.args.get("programa", "").strip()
    tipo_presupuesto = request.args.get("tipo_presupuesto", "").strip()
    partida = request.args.get("partida", "").strip()
    fuente = request.args.get("fuente", "").strip()

    conn = get_connection()
    cur = conn.cursor()

    try:

        # ==============================
        # CABECERA DE LA VERSIÓN
        # ==============================
        cur.execute("""
            SELECT
                id,
                anio,
                numero_reforma,
                fecha_reforma,
                estado,
                nombre_archivo,
                fecha_carga
            FROM pac_versiones
            WHERE id = %s
        """, (version_id,))

        version = cur.fetchone()

        if not version:
            flash("La versión del PAC no existe.", "danger")
            return redirect(
                url_for("planificacion.pac_historial")
            )


        # ==============================
        # FILTROS DISPONIBLES
        # ==============================
        cur.execute("""
            SELECT DISTINCT unidad
            FROM pac_detalle
            WHERE pac_version_id = %s
              AND unidad IS NOT NULL
              AND unidad <> ''
            ORDER BY unidad
        """, (version_id,))
        unidades = [r[0] for r in cur.fetchall()]

        cur.execute("""
            SELECT DISTINCT programa
            FROM pac_detalle
            WHERE pac_version_id = %s
              AND programa IS NOT NULL
              AND programa <> ''
            ORDER BY programa
        """, (version_id,))
        programas = [r[0] for r in cur.fetchall()]

        cur.execute("""
            SELECT DISTINCT tipo_presupuesto
            FROM pac_detalle
            WHERE pac_version_id = %s
            AND tipo_presupuesto IS NOT NULL
            AND tipo_presupuesto <> ''
            ORDER BY tipo_presupuesto
        """, (version_id,))

        tipos_presupuesto = [r[0] for r in cur.fetchall()]
        cur.execute("""
            SELECT DISTINCT partida
            FROM pac_detalle
            WHERE pac_version_id = %s
              AND partida IS NOT NULL
              AND partida <> ''
            ORDER BY partida
        """, (version_id,))
        partidas = [r[0] for r in cur.fetchall()]

        cur.execute("""
            SELECT DISTINCT fuente
            FROM pac_detalle
            WHERE pac_version_id = %s
              AND fuente IS NOT NULL
              AND fuente <> ''
            ORDER BY fuente
        """, (version_id,))
        fuentes = [r[0] for r in cur.fetchall()]


        # ==============================
        # DETALLE
        # ==============================
        sql = """
            SELECT
                id,
                unidad,
                programa,
                partida,
                fuente,
                cpc,
                tipo_compra,
                descripcion,
                procedimiento_sugerido,
                subtotal,
                iva,
                total
            FROM pac_detalle
            WHERE pac_version_id = %s
        """

        params = [version_id]

        if unidad:
            sql += " AND unidad = %s"
            params.append(unidad)

        if programa:
            sql += " AND programa = %s"
            params.append(programa)
        if tipo_presupuesto:
            sql += " AND tipo_presupuesto = %s"
            params.append(tipo_presupuesto)
        if partida:
            sql += " AND partida = %s"
            params.append(partida)

        if fuente:
            sql += " AND fuente = %s"
            params.append(fuente)

        sql += " ORDER BY unidad, programa, partida, id"

        cur.execute(sql, params)
        detalles = cur.fetchall()


        # ==============================
        # RESUMEN FILTRADO
        # ==============================
        total_registros = len(detalles)

        subtotal = sum(
            Decimal(str(r[9] or 0))
            for r in detalles
        )

        iva = sum(
            Decimal(str(r[10] or 0))
            for r in detalles
        )

        total = sum(
            Decimal(str(r[11] or 0))
            for r in detalles
        )

        return render_template(
            "planificacion/pac_version_detalle.html",
            version=version,
            detalles=detalles,
            unidades=unidades,
            tipos_presupuesto=tipos_presupuesto,
            tipo_presupuesto=tipo_presupuesto,
            programas=programas,
            partidas=partidas,
            fuentes=fuentes,
            unidad=unidad,
            programa=programa,
            partida=partida,
            fuente=fuente,
            total_registros=total_registros,
            subtotal=subtotal,
            iva=iva,
            total=total
        )

    finally:
        cur.close()
        conn.close()