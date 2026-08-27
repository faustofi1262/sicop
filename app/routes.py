from flask import Blueprint, render_template, request, redirect, url_for, session, flash, jsonify, current_app
import psycopg2
import os
from werkzeug.security import check_password_hash
from werkzeug.security import generate_password_hash
from datetime import datetime, date
from flask import render_template, abort
from num2words import num2words
from decimal import Decimal, ROUND_HALF_UP
from psycopg2.extras import RealDictCursor
from flask import jsonify
from io import BytesIO
from flask import send_file
from flask import send_file
from io import BytesIO
from app.services.pdf_orden_compra import generar_pdf_orden_compra
from app.services.pdf_certificaciones import (generar_pdf_cate, generar_pdf_pac)
from flask import request, send_file
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from flask import (Blueprint, render_template, request, redirect, url_for, flash, session, jsonify)
#############################
## IMPORTACIONES PARA EXCEL
############################
from openpyxl import load_workbook
from datetime import datetime, date
#############################
## IMPORTACIONES PARA TABLA DE PUBLICACIONES
############################
import os
import uuid

from datetime import date, datetime
from openpyxl import load_workbook

from werkzeug.utils import secure_filename
from app.decorators import login_required
from app.database import get_connection
from datetime import datetime, date, timedelta
from decimal import Decimal

def valor_en_letras_con_decimales(valor):

    if valor is None or str(valor).strip() == "":
        return ""

    # Convertir a texto y limpiar separadores de miles
    texto = str(valor).strip()

    # Si viene como 1.178.568
    if texto.count(".") > 1:
        texto = texto.replace(".", "")

    # Si viene como 1,178,568
    if texto.count(",") > 1:
        texto = texto.replace(",", "")

    # Si viene con coma decimal: 1178568,50
    elif "," in texto and "." not in texto:
        texto = texto.replace(",", ".")

    valor_decimal = Decimal(texto).quantize(
        Decimal("0.01"),
        rounding=ROUND_HALF_UP
    )

    entero = int(valor_decimal)

    centavos = int(
        (valor_decimal - Decimal(entero)) * 100
    )

    letras = num2words(
        entero,
        lang="es"
    ).upper()

    return (
        f"{letras} DÓLARES DE LOS ESTADOS UNIDOS "
        f"DE AMÉRICA CON {centavos:02d}/100"
    )

main = Blueprint(
    "main",
    __name__,
    template_folder="../templates"
)

@main.route("/", methods=["GET"])
def inicio():
    return render_template("inicio.html")


@main.route("/login", methods=["GET"])
def login_form():
    return render_template("login.html")

@main.route("/login", methods=["POST"])
def login():
    usuario = request.form.get("usuario")
    password = request.form.get("password")

    if not usuario or not password:
        flash("Faltan datos", "error")
        return redirect(url_for("main.login_form"))

    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, usuario, password_hash, rol
        FROM usuarios
        WHERE usuario = %s OR correo = %s
    """, (usuario, usuario))

    user = cur.fetchone()

    cur.close()
    conn.close()

    if not user:
        flash("Usuario no encontrado", "error")
        return redirect(url_for("main.login_form"))

    user_id, user_usuario, password_hash, rol = user

    if not check_password_hash(password_hash, password):
        flash("Contraseña incorrecta", "error")
        return redirect(url_for("main.login_form"))

    # ✔ Login exitoso
    session["user_id"] = user_id
    session["usuario"] = user_usuario
    session["rol"] = rol

# ==========================================
# TODOS INGRESAN AL CENTRO DE INTELIGENCIA
# ==========================================
    return redirect(
        url_for("inteligencia.centro_inteligencia")
    )


@main.route("/admin")
@login_required(role="Administrador")
def admin_dashboard():
    return render_template(
        "admin_dashboard.html",
        nombre=session.get("usuario")
    )

@main.route("/usuario")
@login_required(role="Usuario")
def user_dashboard():
    return render_template(
        "usuario_dashboard.html",
        nombre=session.get("usuario")
    )

@main.route("/admin/usuarios")
@login_required(role="Administrador")
def gestionar_usuarios():
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, usuario, nombre, correo, rol
        FROM usuarios
        ORDER BY id
    """)
    rows = cur.fetchall()

    cur.close()
    conn.close()

    usuarios = []
    for r in rows:
        usuarios.append({
            "id": r[0],
            "usuario": r[1],
            "nombre": r[2],
            "correo": r[3],
            "rol": r[4]
        })

    return render_template("admin_usuarios.html", usuarios=usuarios)

# ==========================
# PANEL PRINCIPAL SEGÚN ROL
# ==========================
@main.route("/panel")
@login_required()
def panel_principal():

    rol = session.get("rol", "").strip().lower()

    if rol == "administrador":
        return redirect(url_for("main.admin_dashboard"))

    elif rol == "analista":
        return redirect(url_for("main.analista_dashboard"))

    elif rol == "usuario":
        return redirect(url_for("main.user_dashboard"))

    return redirect(url_for("main.login_form"))
#===================
# CREAR USUARIO
#===================
@main.route("/admin/usuarios/nuevo", methods=["GET", "POST"])
@login_required(role="Administrador")
def crear_usuario():
    if request.method == "POST":
        usuario = request.form.get("usuario")
        nombre = request.form.get("nombre")
        correo = request.form.get("correo")
        password = request.form.get("password")
        rol = request.form.get("rol")

        password_hash = generate_password_hash(password)

        conn = get_connection()
        cur = conn.cursor()

        cur.execute("""
            INSERT INTO usuarios (usuario, nombre, correo, password_hash, rol)
            VALUES (%s, %s, %s, %s, %s)
        """, (usuario, nombre, correo, password_hash, rol))

        conn.commit()
        cur.close()
        conn.close()

        return redirect(url_for("main.gestionar_usuarios"))

    return render_template("admin_usuario_nuevo.html")

#===================
#ELIMINAR USUARIO
#===================

@main.route("/admin/usuarios/eliminar/<int:user_id>", methods=["POST"])
@login_required(role="Administrador")
def eliminar_usuario(user_id):
    # Evitar que el admin se elimine a sí mismo
    if user_id == session.get("user_id"):
        return redirect(url_for("main.gestionar_usuarios"))

    conn = get_connection()
    cur = conn.cursor()

    cur.execute("DELETE FROM usuarios WHERE id = %s", (user_id,))
    conn.commit()

    cur.close()
    conn.close()

    return redirect(url_for("main.gestionar_usuarios"))
#===================
# EDITAR USUARIO
#===================

@main.route("/admin/usuarios/editar/<int:user_id>", methods=["GET", "POST"])
@login_required(role="Administrador")
def editar_usuario(user_id):
    # Evitar editar tu propio rol
    if user_id == session.get("user_id"):
        return redirect(url_for("main.gestionar_usuarios"))

    conn = get_connection()
    cur = conn.cursor()

    if request.method == "POST":
        nuevo_rol = request.form.get("rol")

        cur.execute("""
            UPDATE usuarios
            SET rol = %s
            WHERE id = %s
        """, (nuevo_rol, user_id))

        conn.commit()
        cur.close()
        conn.close()

        return redirect(url_for("main.gestionar_usuarios"))

    # GET → cargar datos actuales
    cur.execute("""
        SELECT id, usuario, rol
        FROM usuarios
        WHERE id = %s
    """, (user_id,))

    row = cur.fetchone()
    cur.close()
    conn.close()

    if not row:
        return redirect(url_for("main.gestionar_usuarios"))

    usuario = {
        "id": row[0],
        "usuario": row[1],
        "rol": row[2]
    }

    return render_template("admin_usuario_editar.html", usuario=usuario)

#====================
#PERFIL ANALISTA Y USUARIO
#====================

@main.route("/analista")
@login_required(role="Analista")
def analista_dashboard():
    return render_template(
        "analista_dashboard.html",
        nombre=session.get("usuario")
    )
# ===============================
# INGRESOS DE REQUERIMIENTOS
# ===============================
@main.route("/requerimientos")
@login_required(role=None)
def listar_requerimientos():
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            r.id,
            r.mem_requi,
            r.fecha_memo_requi,

            CASE
                WHEN UPPER(TRIM(u.nombre_unidad))
                    IN ('DECANATO', 'SUBDECANATO')
                THEN
                    u.nombre_unidad
                    || ' - '
                    || COALESCE(u.bloque, '')

                ELSE
                    u.nombre_unidad
            END AS unidad,

            r.monto_req

        FROM requerimientos r

        LEFT JOIN unidades u
            ON u.id = r.unid_requirente

        ORDER BY r.id DESC
    """)

    rows = cur.fetchall()
    cur.close()
    conn.close()

    requerimientos = [
        {
            "id": r[0],
            "mem_requi": r[1],
            "fecha_memo_requi": r[2],
            "unidad": r[3],
            "monto_req": r[4],
        }
        for r in rows
    ]

    return render_template(
        "requerimientos/requerimientos_list.html",
        requerimientos=requerimientos
    )
# ===============================
# NUEVO REQUERIMIENTOS
# ===============================
@main.route("/requerimientos/nuevo")
@login_required(role=None)  # Admin y Analista
def nuevo_requerimiento():
    conn = get_connection()
    cur = conn.cursor()

   # ==========================================================
    # UNIDADES REQUIRIENTES ACTIVAS
    #
    # Se obtiene:
    # - ID de la unidad
    # - Nombre oficial
    # - Departamento principal
    # - Bloque
    #
    # Para DECANATO y SUBDECANATO se construye una etiqueta
    # especial para distinguir a qué Facultad pertenecen.
    # ==========================================================
    cur.execute("""
        SELECT
            id,
            nombre_unidad,
            departamento_principal,
            bloque,

            CASE
                WHEN UPPER(TRIM(nombre_unidad))
                    IN ('DECANATO', 'SUBDECANATO')
                THEN
                    nombre_unidad
                    || ' - '
                    || COALESCE(bloque, '')

                ELSE
                    nombre_unidad
            END AS etiqueta

        FROM unidades

        WHERE activo = TRUE

        ORDER BY
            bloque,
            departamento_principal,
            nombre_unidad
    """)

    unidades = cur.fetchall()

    # Funcionarios (usuarios)
    cur.execute("SELECT nombre FROM usuarios ORDER BY nombre")
    funcionarios = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "requerimientos/requerimiento_form.html",
        unidades=unidades,
        funcionarios=funcionarios
    )
# ===============================
# GUARDAR REQUERIMIENTOS
# ===============================
@main.route("/requerimientos/guardar", methods=["POST"])
@login_required(role=None)
def guardar_requerimiento():
    conn = get_connection()
    cur = conn.cursor()

    requerimiento_id = request.form.get("id")  # 👈 CLAVE

    if requerimiento_id:
        # =====================
        # UPDATE
        # =====================
        cur.execute("""
            UPDATE requerimientos SET
                mem_requi = %s,
                fecha_memo_requi = %s,
                unid_requirente = %s,
                funcionario_elaborador = %s,
                funcionario_encargado = %s,
                memo_vice_ad = %s,
                fecha_memo_vice_ad = %s,
                memo_dir_ad = %s,
                fecha_memo_dir_ad = %s,
                fecha_recep_req = %s,
                breve_descr = %s,
                descripcion = %s,
                monto_req = %s
            WHERE id = %s
        """, (
            request.form["mem_requi"],
            request.form["fecha_memo_requi"],
            request.form["unid_requirente"],
            request.form.get("funcionario_elaborador"),
            request.form.get("funcionario_encargado"),
            request.form.get("memo_vice_ad"),
            request.form.get("fecha_memo_vice_ad"),
            request.form.get("memo_dir_ad"),
            request.form.get("fecha_memo_dir_ad"),
            request.form.get("fecha_recep_req"),
            request.form.get("breve_descr"),
            request.form["descripcion"],
            request.form["monto_req"],
            requerimiento_id
        ))

        conn.commit()
        cur.close()
        conn.close()

        # 👉 vuelve al listado
        return redirect(url_for("main.listar_requerimientos"))

    else:
        # =====================
        # INSERT
        # =====================
        # ==========================================================
        # CREACIÓN DEL REQUERIMIENTO
        # ==========================================================
        # Registra también al funcionario de la unidad requirente
        # que elaboró la documentación del expediente.
        # ==========================================================

        cur.execute("""
            INSERT INTO requerimientos (
                mem_requi,
                fecha_memo_requi,
                unid_requirente,
                funcionario_elaborador,
                funcionario_encargado,
                memo_vice_ad,
                fecha_memo_vice_ad,
                memo_dir_ad,
                fecha_memo_dir_ad,
                fecha_recep_req,
                breve_descr,
                descripcion,
                monto_req
            )
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING id
        """, (
            request.form["mem_requi"],
            request.form["fecha_memo_requi"],
            request.form["unid_requirente"],
            request.form.get("funcionario_elaborador"),
            request.form.get("funcionario_encargado"),
            request.form.get("memo_vice_ad"),
            request.form.get("fecha_memo_vice_ad"),
            request.form.get("memo_dir_ad"),
            request.form.get("fecha_memo_dir_ad"),
            request.form.get("fecha_recep_req"),
            request.form.get("breve_descr"),
            request.form["descripcion"],
            request.form["monto_req"]
        ))

        nuevo_id = cur.fetchone()[0]

        conn.commit()
        cur.close()
        conn.close()

        # 👉 luego de crear, lo mandamos a editar (para partidas)
        return redirect(url_for("main.editar_requerimiento", id=nuevo_id))
# ===============================
# EDITAR REQUERIMIENTOS
# ===============================    
@main.route("/requerimientos/<int:id>")
@login_required(role=None)
def editar_requerimiento(id):
    conn = get_connection()
    cur = conn.cursor()

    # Requerimiento (columnas explícitas)
    # ==========================================================
    # DATOS DEL REQUERIMIENTO PARA EDICIÓN
    # ==========================================================
    # Recupera también al funcionario elaborador para conservarlo
    # y mostrarlo nuevamente al editar el requerimiento.
    # ==========================================================

    cur.execute("""
        SELECT
            id,
            mem_requi,
            fecha_memo_requi,
            unid_requirente,
            funcionario_elaborador,
            funcionario_encargado,
            memo_vice_ad,
            fecha_memo_vice_ad,
            memo_dir_ad,
            fecha_memo_dir_ad,
            fecha_recep_req,
            breve_descr,
            descripcion,
            monto_req
        FROM requerimientos
        WHERE id = %s
    """, (id,))

    row = cur.fetchone()

    if not row:
        cur.close()
        conn.close()
        return redirect(url_for("main.listar_requerimientos"))

    # 🔑 convertir a dict
    requerimiento = {
        "id": row[0],
        "mem_requi": row[1],
        "fecha_memo_requi": row[2],
        "unid_requirente": row[3],

        # Servidor de la unidad que elaboró los documentos
        "funcionario_elaborador": row[4],

        # Analista UCP asignado al requerimiento
        "funcionario_encargado": row[5],

        "memo_vice_ad": row[6],
        "fecha_memo_vice_ad": row[7],
        "memo_dir_ad": row[8],
        "fecha_memo_dir_ad": row[9],
        "fecha_recep_req": row[10],
        "breve_descr": row[11],
        "descripcion": row[12],
        "monto_req": row[13],
    }

    # ==========================================================
    # UNIDADES REQUIRIENTES ACTIVAS
    #
    # Se obtiene:
    # - ID de la unidad
    # - Nombre oficial
    # - Departamento principal
    # - Bloque
    #
    # Para DECANATO y SUBDECANATO se construye una etiqueta
    # especial para distinguir a qué Facultad pertenecen.
    # ==========================================================
    cur.execute("""
        SELECT
            id,
            nombre_unidad,
            departamento_principal,
            bloque,

            CASE
                WHEN UPPER(TRIM(nombre_unidad))
                    IN ('DECANATO', 'SUBDECANATO')
                THEN
                    nombre_unidad
                    || ' - '
                    || COALESCE(bloque, '')

                ELSE
                    nombre_unidad
            END AS etiqueta

        FROM unidades

        WHERE activo = TRUE

        ORDER BY
            bloque,
            departamento_principal,
            nombre_unidad
    """)

    unidades = cur.fetchall()

    # Funcionarios
    cur.execute("SELECT nombre FROM usuarios ORDER BY nombre")
    funcionarios = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "requerimientos/requerimiento_form.html",
        requerimiento=requerimiento,
        unidades=unidades,
        funcionarios=funcionarios
    )

# ===============================
# INGRESAR PARTIDAS A REQUERIMIENTOS
# ===============================
@main.route("/requerimientos/<int:requerimiento_id>/partidas")
@login_required(role=None)
def partidas_requerimiento(requerimiento_id):
    conn = get_connection()
    cur = conn.cursor()

    # Partidas
    cur.execute("""
        SELECT id, nombre_part, num_part, programa, fuente, monto
        FROM partidas
        WHERE requerimiento_id = %s
        ORDER BY id
    """, (requerimiento_id,))
    partidas = cur.fetchall()

    # 🔢 Total de partidas
    cur.execute("""
        SELECT COALESCE(SUM(monto), 0)
        FROM partidas
        WHERE requerimiento_id = %s
    """, (requerimiento_id,))
    total_partidas = cur.fetchone()[0]

    cur.close()
    conn.close()

    return render_template(
        "requerimientos/partidas_form.html",
        requerimiento_id=requerimiento_id,
        partidas=partidas,
        total_partidas=total_partidas
    )
@main.route('/partidas/editar/<int:id_partida>')
@login_required(role=None)
def partida_editar(id_partida):
    conn = get_connection()
    cur = conn.cursor()

    # Partida a editar
    cur.execute("""
        SELECT id, nombre_part, num_part, programa, fuente, monto, requerimiento_id
        FROM partidas
        WHERE id = %s
    """, (id_partida,))
    partida = cur.fetchone()

    # Partidas del mismo requerimiento
    cur.execute("""
        SELECT id, nombre_part, num_part, programa, fuente, monto
        FROM partidas
        WHERE requerimiento_id = %s
        ORDER BY id
    """, (partida[6],))
    partidas = cur.fetchall()

    # Total
    cur.execute("""
        SELECT COALESCE(SUM(monto), 0)
        FROM partidas
        WHERE requerimiento_id = %s
    """, (partida[6],))
    total_partidas = cur.fetchone()[0]

    cur.close()
    conn.close()

    return render_template(
        "requerimientos/partidas_form.html",
        partida=partida,
        partidas=partidas,
        total_partidas=total_partidas,
        requerimiento_id=partida[6]
    )
@main.route('/partidas/eliminar/<int:id_partida>')
@login_required(role=None)
def partida_eliminar(id_partida):
    conn = get_connection()
    cur = conn.cursor()

    # Obtener requerimiento_id
    cur.execute("""
        SELECT requerimiento_id
        FROM partidas
        WHERE id = %s
    """, (id_partida,))
    row = cur.fetchone()

    if not row:
        cur.close()
        conn.close()
        flash("❌ La partida no existe.", "danger")
        return redirect(url_for("main.partidas_requerimiento", requerimiento_id=0))

    requerimiento_id = row[0]

    # 🔒 VALIDACIÓN: ¿la partida está siendo usada en tareas?
    cur.execute("""
        SELECT COUNT(*)
        FROM tareas
        WHERE partida_id = %s
    """, (id_partida,))
    en_uso = cur.fetchone()[0]

    if en_uso > 0:
        cur.close()
        conn.close()
        flash(
            "⚠️ No se puede eliminar la partida porque ya está asociada a una o más tareas.",
            "warning"
        )
        return redirect(
            url_for("main.partidas_requerimiento", requerimiento_id=requerimiento_id)
        )

    # 🗑️ Eliminar partida (seguro)
    cur.execute("""
        DELETE FROM partidas
        WHERE id = %s
    """, (id_partida,))

    conn.commit()
    cur.close()
    conn.close()

    flash("✅ Partida eliminada correctamente.", "success")

    return redirect(
        url_for("main.partidas_requerimiento", requerimiento_id=requerimiento_id)
    )

# ===============================
# GUARDAR PARTIDAS REQUERIMIENTOS
# ===============================
@main.route("/requerimientos/<int:requerimiento_id>/partidas/guardar", methods=["POST"])
@login_required(role=None)
def guardar_partida(requerimiento_id):
    conn = get_connection()
    cur = conn.cursor()

    id_partida = request.form.get("id_partida")

    if id_partida:
        # 🔄 UPDATE (editar partida)
        cur.execute("""
            UPDATE partidas
            SET nombre_part = %s,
                num_part = %s,
                programa = %s,
                fuente = %s,
                monto = %s
            WHERE id = %s
        """, (
            request.form["nombre_part"],
            request.form["num_part"],
            request.form["programa"],
            request.form["fuente"],
            request.form["monto"],
            id_partida
        ))
    else:
        # ➕ INSERT (nueva partida)
        cur.execute("""
            INSERT INTO partidas (nombre_part, num_part, programa, fuente, monto, requerimiento_id)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (
            request.form["nombre_part"],
            request.form["num_part"],
            request.form["programa"],
            request.form["fuente"],
            request.form["monto"],
            requerimiento_id
        ))

    conn.commit()
    cur.close()
    conn.close()

    return redirect(
        url_for("main.partidas_requerimiento", requerimiento_id=requerimiento_id)
    )
# ==========================================
# EDITAR PARTIDA DESDE EL MODAL DE TAREAS
# ==========================================
@main.route("/partidas/<int:partida_id>/editar", methods=["POST"])
@login_required()
def editar_partida(partida_id):

    nombre_part = request.form.get("nombre_part", "").strip()
    num_part = request.form.get("num_part", "").strip()
    programa = request.form.get("programa", "").strip()
    fuente = request.form.get("fuente", "").strip()
    monto = request.form.get("monto", "").strip()

    if not num_part or not nombre_part or not monto:
        return jsonify({
            "ok": False,
            "mensaje": "Debe completar partida, nombre y monto."
        }), 400

    conn = get_connection()
    cur = conn.cursor()

    try:

        cur.execute("""
            UPDATE partidas
            SET
                nombre_part = %s,
                num_part = %s,
                programa = %s,
                fuente = %s,
                monto = %s,
                usuario_id = %s
            WHERE id = %s
        """, (
            nombre_part,
            num_part,
            programa,
            fuente,
            monto,
            session.get("user_id"),
            partida_id
        ))

        if cur.rowcount == 0:
            conn.rollback()

            return jsonify({
                "ok": False,
                "mensaje": "La partida no existe."
            }), 404

        conn.commit()

        return jsonify({
            "ok": True,
            "mensaje": "Partida actualizada correctamente."
        })

    except Exception as e:

        conn.rollback()

        return jsonify({
            "ok": False,
            "mensaje": str(e)
        }), 500

    finally:

        cur.close()
        conn.close()
# ==========================================
# API - OBTENER UNA PARTIDA
# ==========================================
@main.route("/api/partidas/<int:partida_id>")
@login_required()
def api_obtener_partida(partida_id):

    conn = get_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT
                id,
                nombre_part,
                num_part,
                programa,
                fuente,
                monto
            FROM partidas
            WHERE id = %s
        """, (partida_id,))

        partida = cur.fetchone()

        if not partida:
            return jsonify({
                "ok": False,
                "mensaje": "La partida no existe."
            }), 404

        return jsonify({
            "ok": True,
            "partida": {
                "id": partida[0],
                "nombre_part": partida[1] or "",
                "num_part": partida[2] or "",
                "programa": partida[3] or "",
                "fuente": partida[4] or "",
                "monto": float(partida[5] or 0)
            }
        })

    finally:
        cur.close()
        conn.close()
# ===============================
# ELIMINAR REQUERIMIENTOS
# ===============================
@main.route("/requerimientos/eliminar/<int:id>", methods=["POST"])
@login_required(role=None)
def eliminar_requerimiento(id):
    conn = get_connection()
    cur = conn.cursor()

    # 1️⃣ Eliminar partidas asociadas
    cur.execute("""
        DELETE FROM partidas
        WHERE requerimiento_id = %s
    """, (id,))

    # 2️⃣ Eliminar requerimiento
    cur.execute("""
        DELETE FROM requerimientos
        WHERE id = %s
    """, (id,))

    conn.commit()
    cur.close()
    conn.close()

    return redirect(url_for("main.listar_requerimientos"))

# =================
# NUEVA TAREAS
# ================
@main.route("/tareas/nueva")
@login_required()
def tareas_nueva():
    conn = get_connection()
    cur = conn.cursor()

    # Requerimientos (para memo)
    cur.execute("""
        SELECT id, memo_vice_ad
        FROM requerimientos
        ORDER BY memo_vice_ad
    """)
    requerimientos = cur.fetchall()

    # Tipos de proceso
    cur.execute("""
        SELECT id, nombre_proceso
        FROM tipo_procesos
        ORDER BY nombre_proceso
    """)
    tipos_proceso = cur.fetchall()

    # Tipos de régimen
    cur.execute("""
        SELECT id, nombre_regimen
        FROM tipo_regimen
        ORDER BY nombre_regimen
    """)
    tipos_regimen = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "tareas/tareas_form.html",
        requerimientos=requerimientos,
        tipos_proceso=tipos_proceso,
        tipos_regimen=tipos_regimen
    )
# =================
# EDITAR TAREAS
# =================
@main.route("/tareas/editar/<int:id>")
@login_required()
def tareas_editar(id):
    conn = get_connection()
    cur = conn.cursor()
    # 🔹 Requerimientos (para memo)
    cur.execute("""
    SELECT id, memo_vice_ad
    FROM requerimientos
    ORDER BY memo_vice_ad
    """)
    requerimientos = cur.fetchall()

    # 🔹 Traer la tarea
    cur.execute("""
    SELECT
        id,                         -- 0
        tipo_proceso,               -- 1
        estado_requerimiento,       -- 2
        objeto_contratacion,        -- 3
        codigo_proceso,             -- 4
        fecha_recepcion,            -- 5
        valor_sin_iva,              -- 6
        valor_exento,               -- 7
        valor_en_letras,            -- 8
        tipo_regimen,               -- 9
        base_legal,                 -- 10
        observaciones,              -- 11
        funcionario_encargado,      -- 12
        unidad_solicitante,         -- 13
        requerimiento_id,           -- 14

        -- 🔹 GESTIÓN DE OBSERVACIONES
        fecha_envio_observaciones,  -- 15
        fecha_correccion_observacion, -- 16
        nombre_jefe_compras,        -- 17

        -- 🔹 VERIFICACIÓN DOCUMENTAL (BOOLEANOS)
        presenta_estudio_previo,        -- 18
        presenta_terminos_referencia,   -- 19
        presenta_estudio_mercado,       -- 20
        presenta_especificaciones,      -- 21
        presenta_proformas,             -- 22
        determinacion_necesidad,        -- 23
        consta_catalogo_electronico,    -- 24
        consta_poa,                     -- 25
        consta_pac,                     -- 26
        presenta_errores,               -- 27
        cumple_normativa,                -- 28
          -- 🔹 NUEVOS CAMPOS
        presenta_planos,                -- 29
        presenta_apus,                  -- 30
        presenta_condiciones_contratacion, -- 31
        presenta_viabilidad_tecnico_economica -- 32
        tipo_compra                        -- 33
    FROM tareas
    WHERE id = %s
""", (id,))

    tarea = cur.fetchone()

    if not tarea:
        conn.close()
        flash("Tarea no encontrada", "danger")
        return redirect(url_for("main.tareas"))

    # 🔹 Traer combos (ESTO FALTABA)
    cur.execute("SELECT id, nombre_proceso FROM tipo_procesos ORDER BY nombre_proceso")
    tipos_proceso = cur.fetchall()

    cur.execute("SELECT id, nombre_regimen FROM tipo_regimen ORDER BY nombre_regimen")
    tipos_regimen = cur.fetchall()

    conn.close()

    return render_template(
        "tareas/tareas_form.html",
        tarea=tarea,
        requerimientos=requerimientos,
        tipos_proceso=tipos_proceso,
        tipos_regimen=tipos_regimen
    )
# =================
# GUARDAR / EDITAR TAREAS (SIN DUPLICAR)
# =================
def to_decimal(valor):
    try:
        if valor in (None, "", " "):
            return 0
        return float(valor)
    except Exception:
        return 0

def to_bool(name):
    return True if request.form.get(name) == "on" else False


@main.route("/tareas/guardar", methods=["POST"])
@login_required()
def guardar_tarea():
    conn = get_connection()
    cur = conn.cursor()

    # -------------------------
    # 1) TOMAR ID (SI VIENE)
    # -------------------------
    tarea_id_raw = request.form.get("id", "").strip()
    tarea_id = int(tarea_id_raw) if tarea_id_raw.isdigit() else None

    # Debug opcional:
    print("ID RECIBIDO:", tarea_id)

    # -------------------------
    # 2) CAMPOS PRINCIPALES
    # -------------------------
    codigo_proceso = (request.form.get("codigo_proceso") or "").strip()
    nombre_jefe = (request.form.get("nombre_jefe_compras") or "").strip()
    tipo_compra = (request.form.get("tipo_compra") or "").strip()
    # -------------------------
    # VALORES ECONÓMICOS
    # -------------------------
    valor_sin_iva = to_decimal(
        request.form.get("valor_sin_iva")
    )

    valor_exento = to_decimal(
        request.form.get("valor_exento")
    )

    valor_en_letras = valor_en_letras_con_decimales(
        valor_sin_iva
    )
    # -------------------------
    # 3) PROTEGER NOMBRE JEFE (EDICIÓN)
    # -------------------------
    if tarea_id and not nombre_jefe:
        cur.execute("SELECT nombre_jefe_compras FROM tareas WHERE id = %s", (tarea_id,))
        row = cur.fetchone()
        if row and row[0]:
            nombre_jefe = row[0]

    # Si es NUEVA tarea, jefe es obligatorio
    if not tarea_id and not nombre_jefe:
        flash("⚠️ Debe ingresar el nombre del Jefe de Compras Públicas.", "danger")
        cur.close()
        conn.close()
        return redirect(url_for("main.tareas_nueva"))

    # -------------------------
    # 4) CHECKBOX
    # -------------------------
    presenta_estudio_previo = to_bool("presenta_estudio_previo")
    presenta_terminos_referencia = to_bool("presenta_terminos_referencia")
    presenta_estudio_mercado = to_bool("presenta_estudio_mercado")
    presenta_especificaciones = to_bool("presenta_especificaciones")
    presenta_proformas = to_bool("presenta_proformas")
    determinacion_necesidad = to_bool("determinacion_necesidad")
    consta_catalogo_electronico = to_bool("consta_catalogo_electronico")
    consta_poa = to_bool("consta_poa")
    consta_pac = to_bool("consta_pac")
    presenta_errores = to_bool("presenta_errores")
    cumple_normativa = to_bool("cumple_normativa")

    # SOLO OBRAS
    presenta_planos = to_bool("presenta_planos")
    presenta_apus = to_bool("presenta_apus")
    presenta_condiciones_contratacion = to_bool("presenta_condiciones_contratacion")

    # SOLO CEP
    presenta_viabilidad_tecnico_economica = to_bool("presenta_viabilidad_tecnico_economica")

    # -------------------------
    # 5) VALIDAR CÓDIGO REPETIDO
    # -------------------------
    if codigo_proceso:
        if tarea_id:
            cur.execute(
                "SELECT 1 FROM tareas WHERE codigo_proceso = %s AND id <> %s",
                (codigo_proceso, tarea_id)
            )
        else:
            cur.execute(
                "SELECT 1 FROM tareas WHERE codigo_proceso = %s",
                (codigo_proceso,)
            )

        if cur.fetchone():
            flash("⚠️ Código de proceso repetido", "danger")
            cur.close()
            conn.close()
            return redirect(
                url_for("main.tareas_editar", id=tarea_id) if tarea_id else url_for("main.tareas_nueva")
            )

    # -------------------------
    # 6) SI HAY ID -> ASEGURAR QUE EXISTA (ANTI-DUPLICADO)
    # -------------------------
    if tarea_id:
        cur.execute("SELECT 1 FROM tareas WHERE id = %s", (tarea_id,))
        if not cur.fetchone():
            # Si llega un id inválido, NO INSERTES JAMÁS (evita duplicación)
            flash("⚠️ No se pudo editar: el ID de la tarea no existe o no llegó correctamente.", "danger")
            cur.close()
            conn.close()
            return redirect(url_for("main.tareas"))

        # =========================
        # UPDATE (EDITAR)
        # =========================
        cur.execute("""
            UPDATE tareas SET
                tipo_proceso = %s,
                estado_requerimiento = %s,
                objeto_contratacion = %s,
                codigo_proceso = %s,
                fecha_recepcion = %s,
                valor_sin_iva = %s,
                valor_exento = %s,
                valor_en_letras = %s,
                tipo_regimen = %s,
                tipo_compra = %s,
                base_legal = %s,
                observaciones = %s,
                funcionario_encargado = %s,
                nombre_jefe_compras = %s,
                unidad_solicitante = %s,
                requerimiento_id = %s,

                presenta_estudio_previo = %s,
                presenta_terminos_referencia = %s,
                presenta_estudio_mercado = %s,
                presenta_especificaciones = %s,
                presenta_proformas = %s,
                determinacion_necesidad = %s,
                consta_catalogo_electronico = %s,
                consta_poa = %s,
                consta_pac = %s,
                presenta_errores = %s,
                cumple_normativa = %s,

                presenta_planos = %s,
                presenta_apus = %s,
                presenta_condiciones_contratacion = %s,
                presenta_viabilidad_tecnico_economica = %s

            WHERE id = %s
        """, (
            request.form.get("tipo_proceso"),
            request.form.get("estado_requerimiento"),
            request.form.get("objeto_contratacion"),
            codigo_proceso,
            request.form.get("fecha_recepcion"),

            valor_sin_iva,
            valor_exento,
            valor_en_letras,

            request.form.get("tipo_regimen"),
            tipo_compra,
            request.form.get("base_legal"),
            request.form.get("observaciones"),
            request.form.get("funcionario_encargado"),
            nombre_jefe,
            request.form.get("unidad_solicitante"),
            request.form.get("requerimiento_id") or None,

            presenta_estudio_previo,
            presenta_terminos_referencia,
            presenta_estudio_mercado,
            presenta_especificaciones,
            presenta_proformas,
            determinacion_necesidad,
            consta_catalogo_electronico,
            consta_poa,
            consta_pac,
            presenta_errores,
            cumple_normativa,

            presenta_planos,
            presenta_apus,
            presenta_condiciones_contratacion,
            presenta_viabilidad_tecnico_economica,

            tarea_id
        ))

        flash("✅ Tarea actualizada correctamente", "success")

    else:
        
        # =========================
        # INSERT (NUEVA)
        # =========================
        cur.execute("""
            INSERT INTO tareas (
                tipo_proceso,
                estado_requerimiento,
                objeto_contratacion,
                codigo_proceso,
                fecha_recepcion,
                valor_sin_iva,
                valor_exento,
                valor_en_letras,
                tipo_regimen,
                tipo_compra,
                base_legal,
                observaciones,
                funcionario_encargado,
                nombre_jefe_compras,
                unidad_solicitante,
                requerimiento_id,

                presenta_estudio_previo,
                presenta_terminos_referencia,
                presenta_estudio_mercado,
                presenta_especificaciones,
                presenta_proformas,
                determinacion_necesidad,
                consta_catalogo_electronico,
                consta_poa,
                consta_pac,
                presenta_errores,
                cumple_normativa,

                presenta_planos,
                presenta_apus,
                presenta_condiciones_contratacion,
                presenta_viabilidad_tecnico_economica
            )
            VALUES (
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s,
                %s
            )
        """, (
            request.form.get("tipo_proceso"),
            request.form.get("estado_requerimiento"),
            request.form.get("objeto_contratacion"),
            codigo_proceso,
            request.form.get("fecha_recepcion"),

            valor_sin_iva,
            valor_exento,
            valor_en_letras,

            request.form.get("tipo_regimen"),
            tipo_compra,
            request.form.get("base_legal"),
            request.form.get("observaciones"),
            request.form.get("funcionario_encargado"),
            nombre_jefe,
            request.form.get("unidad_solicitante"),
            request.form.get("requerimiento_id") or None,

            presenta_estudio_previo,
            presenta_terminos_referencia,
            presenta_estudio_mercado,
            presenta_especificaciones,
            presenta_proformas,
            determinacion_necesidad,
            consta_catalogo_electronico,
            consta_poa,
            consta_pac,
            presenta_errores,
            cumple_normativa,

            presenta_planos,
            presenta_apus,
            presenta_condiciones_contratacion,
            presenta_viabilidad_tecnico_economica
        ))

        flash("✅ Tarea guardada correctamente", "success")

    conn.commit()
    cur.close()
    conn.close()
    return redirect(url_for("main.tareas"))

@main.route("/tareas")
@login_required()
def tareas():

    buscar = request.args.get("buscar", "").strip()
    codigo = request.args.get("codigo", "").strip()
    unidad = request.args.get("unidad", "").strip()
    funcionario = request.args.get("funcionario", "").strip()

    conn = get_connection()
    cur = conn.cursor()

    sql = """
        SELECT 
            t.id,
            t.codigo_proceso,
            t.objeto_contratacion,
            tp.nombre_proceso,
            t.estado_requerimiento,
            t.fecha_recepcion,
            t.numero_certificacion,
            t.funcionario_encargado,

            (
                SELECT sc.id
                FROM seguimiento_contratos sc
                WHERE sc.codigo_proceso = t.codigo_proceso
                ORDER BY sc.id DESC
                LIMIT 1
            ) AS contrato_id

        FROM tareas t

        LEFT JOIN tipo_procesos tp
            ON t.tipo_proceso = tp.id::TEXT

        WHERE 1=1
    """

    params = []

    # ==========================================
    # BUSCADOR GENERAL
    # ==========================================
    if buscar:

        patron = f"%{buscar}%"

        sql += """
            AND (
                COALESCE(t.codigo_proceso, '') ILIKE %s
                OR COALESCE(t.objeto_contratacion, '') ILIKE %s
                OR COALESCE(tp.nombre_proceso, '') ILIKE %s
                OR COALESCE(t.estado_requerimiento, '') ILIKE %s
                OR COALESCE(t.unidad_solicitante, '') ILIKE %s
                OR COALESCE(t.funcionario_encargado, '') ILIKE %s
                OR COALESCE(t.numero_certificacion, '') ILIKE %s
            )
        """

        params.extend([
            patron,
            patron,
            patron,
            patron,
            patron,
            patron,
            patron
        ])

    # ==========================================
    # FILTROS INDIVIDUALES
    # ==========================================
    if codigo:
        sql += " AND t.codigo_proceso ILIKE %s"
        params.append(f"%{codigo}%")

    if unidad:
        sql += " AND t.unidad_solicitante ILIKE %s"
        params.append(f"%{unidad}%")

    if funcionario:
        sql += " AND t.funcionario_encargado ILIKE %s"
        params.append(f"%{funcionario}%")

    sql += " ORDER BY t.id DESC"

    cur.execute(sql, params)

    tareas = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "tareas_list.html",
        tareas=tareas,
        buscar=buscar,
        codigo=codigo,
        unidad=unidad,
        funcionario=funcionario
    )

# =================
# ELIMINAR  TAREAS
# ================
@main.route("/tareas/eliminar/<int:id>", methods=["POST"])
@login_required()
def tareas_eliminar(id):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("DELETE FROM tareas WHERE id = %s", (id,))
    conn.commit()

    cur.close()
    conn.close()

    flash("🗑️ Tarea eliminada correctamente", "success")
    return redirect(url_for("main.tareas"))

# ==========================================
# API - DATOS DEL REQUERIMIENTO PARA TAREAS
# ==========================================
@main.route("/api/requerimiento/<int:requerimiento_id>")
@login_required()
def api_requerimiento(requerimiento_id):

    conn = get_connection()
    cur = conn.cursor()

    try:

        # ==========================================
        # CONSULTAR REQUERIMIENTO + UNIDAD OFICIAL
        # ==========================================
        cur.execute("""
            SELECT
                u.nombre_unidad,
                u.departamento_principal,
                u.bloque,
                r.funcionario_encargado
            FROM requerimientos r

            LEFT JOIN unidades u
                ON u.id = r.unid_requirente

            WHERE r.id = %s
        """, (requerimiento_id,))

        row = cur.fetchone()

        if not row:
            return jsonify({
                "ok": False,
                "unidad": "",
                "departamento_principal": "",
                "bloque": "",
                "funcionario": ""
            }), 404


        # ==========================================
        # RESPUESTA JSON
        # ==========================================
        return jsonify({
            "ok": True,
            "unidad": row[0] or "",
            "departamento_principal": row[1] or "",
            "bloque": row[2] or "",
            "funcionario": row[3] or ""
        })

    finally:

        cur.close()
        conn.close()

# ==========================================
# API - DETALLE DE TAREA Y PARTIDAS
# ==========================================
@main.route("/api/tareas/<int:tarea_id>/detalle")
@login_required()
def api_tarea_detalle(tarea_id):

    conn = get_connection()
    cur = conn.cursor()

    try:

        cur.execute("""
            SELECT
                t.id,
                t.codigo_proceso,
                t.objeto_contratacion,
                t.estado_requerimiento,
                t.unidad_solicitante,
                t.funcionario_encargado,
                t.requerimiento_id
            FROM tareas t
            WHERE t.id = %s
        """, (tarea_id,))

        tarea = cur.fetchone()

        if not tarea:
            return jsonify({
                "ok": False,
                "mensaje": "La tarea no existe."
            }), 404

        requerimiento_id = tarea[6]

        # ==========================================
        # PARTIDAS + ADJUDICACIÓN EXISTENTE
        # ==========================================
        cur.execute("""
            SELECT
                p.id,
                p.nombre_part,
                p.num_part,
                p.programa,
                p.fuente,
                p.monto,

                COALESCE(
                    ap.monto_adjudicado,
                    0
                ) AS monto_adjudicado

            FROM partidas p

            LEFT JOIN LATERAL (
                SELECT
                    a.id
                FROM adjudicaciones a
                WHERE a.tarea_id = %s
                ORDER BY a.id DESC
                LIMIT 1
            ) a ON TRUE

            LEFT JOIN adjudicacion_partidas ap
                ON ap.adjudicacion_id = a.id
            AND ap.partida_id = p.id

            WHERE p.requerimiento_id = %s

            ORDER BY p.id
        """, (
            tarea_id,
            requerimiento_id
        ))

        filas_partidas = cur.fetchall()

        partidas = []
        total_partidas = 0

        for p in filas_partidas:

            monto = float(p[5] or 0)

            total_partidas += monto

            partidas.append({
                "id": p[0],
                "nombre": p[1] or "",
                "numero": p[2] or "",
                "programa": p[3] or "",
                "fuente": p[4] or "",
                "monto": monto,
                "monto_adjudicado": float(p[6] or 0)
            })

        return jsonify({
            "ok": True,

            "tarea": {
                "id": tarea[0],
                "codigo": tarea[1] or "",
                "objeto": tarea[2] or "",
                "estado": tarea[3] or "",
                "unidad": tarea[4] or "",
                "analista": tarea[5] or "",
                "requerimiento_id": tarea[6]
            },

            "partidas": partidas,
            "total_partidas": total_partidas
        })

    finally:

        cur.close()
        conn.close()
# =========================
# LISTADO ORDENES DE COMPRA
# =========================
@main.route("/ordenes_compra")
@login_required()
def ordenes_compra():
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            id,
            numero_oc,
            fecha,
            proveedor,
            total
        FROM ordenes_compra
        ORDER BY fecha DESC
    """)

    ordenes = cur.fetchall()
    cur.close()
    conn.close()

    return render_template(
        "ordenes_compra/ordenes_compra_list.html",
        ordenes=ordenes
    )
# =========================
# PDF ORDEN DE COMPRA
# =========================
@main.route("/ordenes_compra/pdf/<int:id>")
@login_required()
def orden_compra_pdf(id):

    conn = get_connection()
    cur = conn.cursor()

    try:

        # ==========================================
        # 1. ORDEN DE COMPRA + ESTRUCTURA
        #    INSTITUCIONAL DE LA UNIDAD
        # ==========================================
        cur.execute("""
            SELECT
                oc.*,

                u.nombre_unidad,
                u.departamento_principal,
                u.bloque

            FROM ordenes_compra oc

            LEFT JOIN tareas t
                ON t.id = oc.tarea_id

            LEFT JOIN requerimientos r
                ON r.id = t.requerimiento_id

            LEFT JOIN unidades u
                ON u.id = r.unid_requirente

            WHERE oc.id = %s
        """, (id,))

        orden = cur.fetchone()


        # ==========================================
        # 2. PRODUCTOS DE LA ORDEN
        # ==========================================
        cur.execute("""
            SELECT
                id,
                descripcion,
                unidad,
                cantidad,
                valor_uni,
                cantidad * valor_uni AS valor_total,
                cpc
            FROM productos
            WHERE orden_compra_id = %s
            ORDER BY id
        """, (id,))

        productos = cur.fetchall()

    finally:

        cur.close()
        conn.close()


    # ==========================================
    # 3. VALIDAR EXISTENCIA
    # ==========================================
    if not orden:
        abort(404)


    # ==========================================
    # 4. GENERAR PDF
    # ==========================================
    buffer = generar_pdf_orden_compra(
        orden,
        productos
    )


    # ==========================================
    # 5. DESCARGAR PDF
    # ==========================================
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"orden_compra_{id}.pdf",
        mimetype="application/pdf"
    )
# =========================
# NUEVA ORDEN DE COMPRA
# ÍNFIMA CUANTÍA
# =========================
@main.route("/ordenes_compra/nueva")
@main.route("/ordenes_compra/nueva/<int:tarea_id>")
@login_required()
def ordenes_compra_nueva(tarea_id=None):

    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            t.id,
            t.codigo_proceso,
            r.memo_vice_ad
        FROM tareas t
        JOIN requerimientos r
            ON r.id = t.requerimiento_id
        WHERE UPPER(TRIM(COALESCE(t.codigo_proceso, '')))
              LIKE 'IC-%'
        ORDER BY t.codigo_proceso
    """)

    tareas = cur.fetchall()

    # Verificar que la tarea enviada corresponda a Ínfima Cuantía
    if tarea_id is not None:

        tarea_valida = any(
            tarea[0] == tarea_id
            for tarea in tareas
        )

        if not tarea_valida:

            cur.close()
            conn.close()

            flash(
                "La tarea seleccionada no corresponde a un proceso de Ínfima Cuantía.",
                "danger"
            )

            return redirect(
                url_for("main.tareas")
            )

    cur.close()
    conn.close()

    return render_template(
        "ordenes_compra/ordenes_compra_form.html",
        tareas=tareas,
        tarea_id_seleccionada=tarea_id,
        orden=None,
        items=[]
    )
# =========================
# GUARDAR ORDEN DE COMPRA
# =========================
@main.route("/ordenes_compra/guardar", methods=["POST"])
@login_required()
def ordenes_compra_guardar():

    conn = get_connection()
    cur = conn.cursor()

    orden_id = request.form.get("id")  # 🔑 CLAVE

    # ============================
    # 1️⃣ VALIDACIÓN DE CAMPOS OBLIGATORIOS
    # ============================
    campos_obligatorios = {
        "tarea_id": "Proceso asociado",
        "numero_oc": "Número de Orden de Compra",
        "fecha": "Fecha",
        "area_requirente": "Área requirente",
        "objeto": "Objeto de contratación",
        "proveedor": "Proveedor",
        "ruc": "RUC"
    }

    faltantes = []
    for campo, nombre in campos_obligatorios.items():
        valor = request.form.get(campo)
        if not valor or not valor.strip():
            faltantes.append(nombre)

    if faltantes:
        mensaje = "🚨 FALTA DE LLENAR LOS SIGUIENTES CAMPOS:<br>" + "<br>".join(
            f"• {c}" for c in faltantes
        )
        flash(mensaje, "danger")
        cur.close()
        conn.close()
        return redirect(request.referrer)

    try:
        # ============================
        # 2️⃣ INSERT o UPDATE CABECERA
        # ============================
        if orden_id:  # ✏️ EDITAR
            cur.execute("""
                UPDATE ordenes_compra SET
                    numero_oc=%s, fecha=%s, area_requirente=%s, cert_presupuestaria=%s,
                    objeto=%s, proveedor=%s, ruc=%s, telefono=%s, direccion=%s, correo=%s,
                    proforma_num=%s, proforma_fecha=%s, contacto=%s, vigencia=%s,
                    forma_pago=%s, plazo_ejecucion=%s, lugar_entrega=%s,
                    administrador_orden=%s,
                    maxima_autoridad=%s,
                    cargo_maxima_autoridad=%s,
                    subtotal=%s, iva=%s, total=%s,
                    observaciones=%s, tarea_id=%s
                WHERE id=%s
            """, (
                request.form["numero_oc"],
                request.form["fecha"],
                request.form["area_requirente"],
                request.form["cert_presupuestaria"],
                request.form["objeto"],
                request.form["proveedor"],
                request.form["ruc"],
                request.form["telefono"],
                request.form["direccion"],
                request.form["correo"],
                request.form["proforma_num"],
                request.form["proforma_fecha"],
                request.form["contacto"],
                request.form["vigencia"],
                request.form["forma_pago"],
                request.form["plazo_ejecucion"],
                request.form["lugar_entrega"],
                request.form["administrador_orden"],
                request.form.get("maxima_autoridad"),
                request.form.get("cargo_maxima_autoridad"),
                request.form.get("subtotal", 0),
                request.form.get("iva", 0),
                request.form.get("total", 0),
                request.form.get("observaciones"),
                request.form["tarea_id"],
                orden_id
            ))

            # 🧹 borrar ítems anteriores
            cur.execute("DELETE FROM productos WHERE orden_compra_id = %s", (orden_id,))
            orden_compra_id = orden_id

        else:  # ➕ NUEVA
            cur.execute("""
                INSERT INTO ordenes_compra (
                    numero_oc, fecha, area_requirente, cert_presupuestaria,
                    objeto, proveedor, ruc, telefono, direccion, correo,
                    proforma_num, proforma_fecha, contacto, vigencia,
                    forma_pago, plazo_ejecucion, lugar_entrega,
                    administrador_orden,
                    maxima_autoridad,
                    cargo_maxima_autoridad,
                    subtotal, iva, total,
                    observaciones, tarea_id
                )
                VALUES (
                    %s,%s,%s,%s,
                    %s,%s,%s,%s,%s,%s,
                    %s,%s,%s,%s,
                    %s,%s,%s,
                    %s,
                    %s,
                    %s,    
                    %s,%s,%s,
                    %s,%s
                )
                RETURNING id
            """, (
                request.form["numero_oc"],
                request.form["fecha"],
                request.form["area_requirente"],
                request.form["cert_presupuestaria"],
                request.form["objeto"],
                request.form["proveedor"],
                request.form["ruc"],
                request.form["telefono"],
                request.form["direccion"],
                request.form["correo"],
                request.form["proforma_num"],
                request.form["proforma_fecha"],
                request.form["contacto"],
                request.form["vigencia"],
                request.form["forma_pago"],
                request.form["plazo_ejecucion"],
                request.form["lugar_entrega"],
                request.form["administrador_orden"],
                request.form.get("maxima_autoridad"),
                request.form.get("cargo_maxima_autoridad"),
                request.form.get("subtotal", 0),
                request.form.get("iva", 0),
                request.form.get("total", 0),
                request.form.get("observaciones"),
                request.form["tarea_id"]
            ))

            orden_compra_id = cur.fetchone()[0]

        # ============================
        # 3️⃣ GUARDAR ÍTEMS
        # ============================
        descripciones = request.form.getlist("descripcion[]")
        unidades = request.form.getlist("unidad[]")
        cantidades = request.form.getlist("cantidad[]")
        valores = request.form.getlist("valor_unitario[]")

        cpcs = request.form.getlist("cpc[]")

        for i in range(len(descripciones)):

            if not descripciones[i].strip():
                continue

            cantidad = cantidades[i] if i < len(cantidades) else 0
            valor = valores[i] if i < len(valores) else 0
            unidad = unidades[i] if i < len(unidades) else None
            cpc = cpcs[i] if i < len(cpcs) else None

          
            cur.execute("""
               INSERT INTO productos (
                    descripcion,
                    unidad,
                    cantidad,
                    valor_uni,
                    orden_compra_id,
                    cpc
                )
                VALUES (%s,%s,%s,%s,%s,%s)                
            """, (
                descripciones[i],
                unidad,
                cantidad,
                valor,                
                orden_compra_id,
                cpc
            ))

        conn.commit()
        flash("✅ Orden de Compra guardada correctamente", "success")
        return redirect(url_for("main.ordenes_compra"))

    except Exception as e:
        conn.rollback()
        print("🔥 ERROR OC:", e)   # 👈 AGREGA EST
        flash(f"❌ Error al guardar la Orden de Compra: {e}", "danger")
        return redirect(request.referrer)

    finally:
        cur.close()
        conn.close()


@main.route("/ordenes_compra/eliminar/<int:id>", methods=["POST"])
@login_required()
def ordenes_compra_eliminar(id):
    
    try:
        conn = get_connection()
        print("✔ conexión creada")

        cur = conn.cursor()
        print("✔ cursor creado")

        cur.execute(
            "DELETE FROM productos WHERE orden_compra_id = %s",
            (id,)
        )
        print("Productos eliminados:", cur.rowcount)

        cur.execute(
            "DELETE FROM ordenes_compra WHERE id = %s",
            (id,)
        )
        print("Orden eliminada:", cur.rowcount)

        conn.commit()
        print("✔ commit realizado")

    except Exception as e:
        print("❌ ERROR:", e)
        conn.rollback()

    finally:
        cur.close()
        conn.close()

    return redirect(url_for("main.ordenes_compra"))

# ==========================================
# API - DATOS DE LA TAREA PARA
# ÓRDENES DE COMPRA
# ==========================================
@main.route("/api/tarea/<int:id>")
@login_required()
def api_tarea(id):

    conn = get_connection()
    cur = conn.cursor()

    try:

        # ==========================================
        # TAREA + REQUERIMIENTO + UNIDAD OFICIAL
        # ==========================================
        cur.execute("""
            SELECT
                t.codigo_proceso,
                t.objeto_contratacion,

                u.nombre_unidad,
                u.departamento_principal,
                u.bloque

            FROM tareas t

            JOIN requerimientos r
                ON r.id = t.requerimiento_id

            JOIN unidades u
                ON u.id = r.unid_requirente

            WHERE t.id = %s
        """, (id,))

        row = cur.fetchone()

        if not row:
            return jsonify({
                "ok": False
            }), 404


        # ==========================================
        # RESPUESTA PARA EL FORMULARIO
        # ==========================================
        return jsonify({
            "ok": True,

            "codigo_proceso": row[0] or "",
            "objeto": row[1] or "",

            "unidad": row[2] or "",
            "departamento_principal": row[3] or "",
            "bloque": row[4] or ""
        })

    finally:

        cur.close()
        conn.close()

# ================================
# EDITAR ORDEN DE COMPRA
# ================================
@main.route("/ordenes_compra/editar/<int:id>")
@login_required()
def ordenes_compra_editar(id):
    conn = get_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    # Cabecera
    cur.execute("""
        SELECT *
        FROM ordenes_compra
        WHERE id = %s
    """, (id,))
    orden = cur.fetchone()

    # Productos
    cur.execute("""
        SELECT
            id,
            descripcion,
            unidad,
            cantidad,
            valor_uni,
            cpc
        FROM productos
        WHERE orden_compra_id = %s
        ORDER BY id
    """, (id,))
    productos = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "ordenes_compra/ordenes_compra_form.html",
        orden=orden,
        productos=productos
    )
# ==========================================
# NUEVA ORDEN DE CATÁLOGO ELECTRÓNICO
# ==========================================
@main.route("/ordenes_catalogo/nueva/<int:tarea_id>")
@login_required()
def orden_catalogo_nueva(tarea_id):

    conn = get_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    cur.execute("""
        SELECT
            t.id AS tarea_id,
            t.codigo_proceso,
            t.objeto_contratacion,
            r.memo_vice_ad,
            r.fecha_memo_vice_ad,
            r.monto_req AS presupuesto_referencial,
            u.nombre_unidad AS unidad_requirente
        FROM tareas t
        JOIN requerimientos r
            ON r.id = t.requerimiento_id
        LEFT JOIN unidades u
            ON u.id = r.unid_requirente
        WHERE t.id = %s
    """, (tarea_id,))

    proceso = cur.fetchone()

    if not proceso:
        cur.close()
        conn.close()

        flash(
            "No se encontró la tarea seleccionada.",
            "danger"
        )
        return redirect(url_for("main.tareas"))

    codigo = (
        proceso["codigo_proceso"] or ""
    ).strip().upper()

    if not codigo.startswith("CATE-"):
        cur.close()
        conn.close()

        flash(
            "La tarea seleccionada no corresponde a Catálogo Electrónico.",
            "danger"
        )
        return redirect(url_for("main.tareas"))

    # ==========================================
    # BUSCAR CATÁLOGO Y ÓRDENES YA REGISTRADAS
    # ==========================================
    cur.execute("""
        SELECT
            id,
            presupuesto_referencial
        FROM catalogos_electronicos
        WHERE tarea_id = %s
    """, (tarea_id,))

    catalogo_existente = cur.fetchone()

    ordenes_registradas = []
    total_adjudicado = Decimal("0.00")

    if catalogo_existente:

        catalogo_id = catalogo_existente["id"]

        cur.execute("""
            SELECT
                id,
                numero_orden,
                fecha_aceptacion,
                ruc,
                proveedor,
                monto_adjudicado,
                administrador_orden,
                plazo_dias,
                fecha_vencimiento,
                observaciones
            FROM ordenes_catalogo
            WHERE catalogo_id = %s
            ORDER BY id DESC
        """, (catalogo_id,))

        ordenes_registradas = cur.fetchall()

        total_adjudicado = sum(
            Decimal(str(orden["monto_adjudicado"] or 0))
            for orden in ordenes_registradas
        )

    presupuesto_referencial = Decimal(
        str(proceso["presupuesto_referencial"] or 0)
    )

    saldo_no_adjudicado = (
        presupuesto_referencial - total_adjudicado
    )

    cur.close()
    conn.close()

    return render_template(
    "ordenes_catalogo/orden_catalogo_form.html",
    proceso=proceso,
    catalogo_existente=catalogo_existente,
    ordenes_registradas=ordenes_registradas,
    total_adjudicado=total_adjudicado,
    saldo_no_adjudicado=saldo_no_adjudicado
)
# ==========================================
# GUARDAR ORDEN DE CATÁLOGO ELECTRÓNICO
# ==========================================
@main.route("/ordenes_catalogo/guardar", methods=["POST"])
@login_required()
def orden_catalogo_guardar():

    tarea_id = request.form.get("tarea_id")
    numero_orden = request.form.get("numero_orden", "").strip()
    fecha_aceptacion = request.form.get("fecha_aceptacion")
    ruc = request.form.get("ruc", "").strip()
    proveedor = request.form.get("proveedor", "").strip()
    monto_adjudicado = request.form.get("monto_adjudicado", 0)
    administrador_orden = request.form.get(
        "administrador_orden", ""
    ).strip()
    plazo_dias = request.form.get("plazo_dias", 0)
    observaciones = request.form.get("observaciones", "").strip()

    if not tarea_id or not numero_orden or not fecha_aceptacion:
        flash(
            "Debe completar la tarea, número de orden y fecha de aceptación.",
            "danger"
        )
        return redirect(request.referrer)

    try:
        plazo_dias = int(plazo_dias or 0)
        monto_adjudicado = Decimal(
            str(monto_adjudicado or 0)
        )

        fecha_aceptacion_obj = datetime.strptime(
            fecha_aceptacion,
            "%Y-%m-%d"
        ).date()

        fecha_vencimiento = (
            fecha_aceptacion_obj
            + timedelta(days=plazo_dias)
        )

        conn = get_connection()
        cur = conn.cursor()

        # Datos del proceso
        cur.execute("""
            SELECT
                t.codigo_proceso,
                t.objeto_contratacion,
                r.memo_vice_ad,
                r.fecha_memo_vice_ad,
                r.monto_req,
                u.nombre_unidad
            FROM tareas t
            JOIN requerimientos r
                ON r.id = t.requerimiento_id
            LEFT JOIN unidades u
                ON u.id = r.unid_requirente
            WHERE t.id = %s
        """, (tarea_id,))

        proceso = cur.fetchone()

        if not proceso:
            flash("No se encontró la tarea seleccionada.", "danger")
            return redirect(url_for("main.tareas"))

        # Crear cabecera si todavía no existe
        cur.execute("""
            INSERT INTO catalogos_electronicos (
                tarea_id,
                codigo_proceso,
                objeto_contratacion,
                unidad_requirente,
                memo_vice_ad,
                fecha_memo_vice_ad,
                presupuesto_referencial
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (tarea_id)
            DO UPDATE SET
                codigo_proceso = EXCLUDED.codigo_proceso,
                objeto_contratacion = EXCLUDED.objeto_contratacion,
                unidad_requirente = EXCLUDED.unidad_requirente,
                memo_vice_ad = EXCLUDED.memo_vice_ad,
                fecha_memo_vice_ad = EXCLUDED.fecha_memo_vice_ad,
                presupuesto_referencial =
                    EXCLUDED.presupuesto_referencial
            RETURNING id
        """, (
            tarea_id,
            proceso[0],
            proceso[1],
            proceso[5],
            proceso[2],
            proceso[3],
            proceso[4]
        ))

        catalogo_id = cur.fetchone()[0]

        # Guardar la orden
        cur.execute("""
            INSERT INTO ordenes_catalogo (
                catalogo_id,
                numero_orden,
                fecha_aceptacion,
                ruc,
                proveedor,
                monto_adjudicado,
                administrador_orden,
                plazo_dias,
                fecha_vencimiento,
                observaciones
            )
            VALUES (
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s
            )
        """, (
            catalogo_id,
            numero_orden,
            fecha_aceptacion_obj,
            ruc,
            proveedor,
            monto_adjudicado,
            administrador_orden,
            plazo_dias,
            fecha_vencimiento,
            observaciones
        ))

        conn.commit()

        flash(
            "✅ Orden de Catálogo guardada correctamente.",
            "success"
        )

        return redirect(
            url_for(
                "main.orden_catalogo_nueva",
                tarea_id=tarea_id
            )
        )

    except Exception as e:

        if "conn" in locals():
            conn.rollback()

        print("ERROR ORDEN CATÁLOGO:", e)

        flash(
            f"❌ Error al guardar la orden: {e}",
            "danger"
        )

        return redirect(request.referrer)

    finally:

        if "cur" in locals():
            cur.close()

        if "conn" in locals():
            conn.close()

# ==========================================
# EDITAR ORDEN DE CATÁLOGO
# ==========================================
@main.route(
    "/ordenes_catalogo/editar/<int:orden_id>",
    methods=["GET", "POST"]
)
@login_required()
def editar_orden_catalogo(orden_id):

    conn = get_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    try:

        cur.execute("""
            SELECT
                oc.*,
                ce.tarea_id,
                ce.codigo_proceso,
                ce.objeto_contratacion,
                ce.unidad_requirente,
                ce.memo_vice_ad,
                ce.fecha_memo_vice_ad,
                ce.presupuesto_referencial
            FROM ordenes_catalogo oc
            JOIN catalogos_electronicos ce
                ON ce.id = oc.catalogo_id
            WHERE oc.id = %s
        """, (orden_id,))

        orden = cur.fetchone()

        if not orden:
            flash(
                "La orden de catálogo no existe.",
                "danger"
            )
            return redirect(url_for("main.tareas"))

        if request.method == "POST":

            numero_orden = request.form.get(
                "numero_orden", ""
            ).strip()

            fecha_aceptacion = request.form.get(
                "fecha_aceptacion"
            )

            ruc = request.form.get(
                "ruc", ""
            ).strip()

            proveedor = request.form.get(
                "proveedor", ""
            ).strip()

            monto_adjudicado = Decimal(
                str(
                    request.form.get(
                        "monto_adjudicado", 0
                    ) or 0
                )
            )

            administrador_orden = request.form.get(
                "administrador_orden", ""
            ).strip()

            plazo_dias = int(
                request.form.get(
                    "plazo_dias", 0
                ) or 0
            )

            observaciones = request.form.get(
                "observaciones", ""
            ).strip()

            fecha_aceptacion_obj = datetime.strptime(
                fecha_aceptacion,
                "%Y-%m-%d"
            ).date()

            fecha_vencimiento = (
                fecha_aceptacion_obj
                + timedelta(days=plazo_dias)
            )

            # Total de las otras órdenes
            cur.execute("""
                SELECT
                    COALESCE(
                        SUM(monto_adjudicado),
                        0
                    ) AS total_otras
                FROM ordenes_catalogo
                WHERE catalogo_id = %s
                AND id <> %s
            """, (
                orden["catalogo_id"],
                orden_id
            ))

            fila = cur.fetchone()

            total_otras = Decimal(
                str(fila["total_otras"] or 0)
            )

            presupuesto = Decimal(
                str(
                    orden[
                        "presupuesto_referencial"
                    ] or 0
                )
            )

            if total_otras + monto_adjudicado > presupuesto:

                flash(
                    "El monto supera el presupuesto disponible.",
                    "danger"
                )

                return redirect(
                    url_for(
                        "main.editar_orden_catalogo",
                        orden_id=orden_id
                    )
                )

            cur.execute("""
                UPDATE ordenes_catalogo
                SET
                    numero_orden = %s,
                    fecha_aceptacion = %s,
                    ruc = %s,
                    proveedor = %s,
                    monto_adjudicado = %s,
                    administrador_orden = %s,
                    plazo_dias = %s,
                    fecha_vencimiento = %s,
                    observaciones = %s
                WHERE id = %s
            """, (
                numero_orden,
                fecha_aceptacion_obj,
                ruc,
                proveedor,
                monto_adjudicado,
                administrador_orden,
                plazo_dias,
                fecha_vencimiento,
                observaciones,
                orden_id
            ))

            conn.commit()

            flash(
                "✅ Orden de catálogo actualizada.",
                "success"
            )

            return redirect(
                url_for(
                    "main.orden_catalogo_nueva",
                    tarea_id=orden["tarea_id"]
                )
            )

        return render_template(
            "ordenes_catalogo/orden_catalogo_editar.html",
            orden=orden
        )

    finally:

        cur.close()
        conn.close()


# =========================
# LISTAR SEGUIMIENTO CONTRATOS CON ALERTAS
# =========================
@main.route("/seguimiento_contratos")
@login_required()
def seguimiento_contratos():

    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            id,
            codigo_proceso,
            objeto_contratacion,
            numero_contrato,
            administrador_contrato,
            fecha_fin_estimada,
            estado
        FROM seguimiento_contratos
        ORDER BY id DESC
    """)

    contratos_db = cur.fetchall()

    contratos = []

    hoy = date.today()

    for c in contratos_db:

        dias_restantes = None
        alerta = "Sin fecha"
        color_alerta = "secondary"

        fecha_fin = c[5]

        if fecha_fin:
            dias_restantes = (fecha_fin - hoy).days

            if dias_restantes < 0:
                alerta = f"Vencido hace {abs(dias_restantes)} días"
                color_alerta = "danger"

            elif dias_restantes <= 5:
                alerta = f"Vence en {dias_restantes} días"
                color_alerta = "danger"

            elif dias_restantes <= 15:
                alerta = f"Vence en {dias_restantes} días"
                color_alerta = "warning"

            elif dias_restantes <= 30:
                alerta = f"Vence en {dias_restantes} días"
                color_alerta = "info"

            else:
                alerta = f"Vence en {dias_restantes} días"
                color_alerta = "success"

        contratos.append({
            "id": c[0],
            "codigo_proceso": c[1],
            "objeto_contratacion": c[2],
            "numero_contrato": c[3],
            "administrador_contrato": c[4],
            "fecha_fin_estimada": c[5],
            "estado": c[6],
            "dias_restantes": dias_restantes,
            "alerta": alerta,
            "color_alerta": color_alerta
        })

    cur.close()
    conn.close()

    return render_template(
        "seguimiento_contratos/seguimiento_contratos_list.html",
        contratos=contratos
    )
# =========================
# DASHBOARD SEGUIMIENTO CONTRACTUAL
# =========================
@main.route("/seguimiento_contratos/dashboard")
@login_required()
def seguimiento_contratos_dashboard():

    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            id,
            codigo_proceso,
            objeto_contratacion,
            numero_contrato,
            administrador_contrato,
            fecha_fin_estimada,
            estado
        FROM seguimiento_contratos
        ORDER BY fecha_fin_estimada ASC
    """)

    contratos_db = cur.fetchall()

    cur.close()
    conn.close()

    hoy = date.today()

    total_contratos = len(contratos_db)
    activos = 0
    por_vencer_30 = 0
    por_vencer_15 = 0
    por_vencer_5 = 0
    vencidos = 0
    finalizados = 0

    contratos_alerta = []

    for c in contratos_db:

        estado = (c[6] or "").upper()
        fecha_fin = c[5]

        dias_restantes = None
        alerta = "Sin fecha"
        color_alerta = "secondary"

        if estado == "FINALIZADO":
            finalizados += 1
        else:
            activos += 1

        if fecha_fin:

            dias_restantes = (fecha_fin - hoy).days

            if dias_restantes < 0 and estado != "FINALIZADO":
                vencidos += 1
                alerta = f"Vencido hace {abs(dias_restantes)} días"
                color_alerta = "danger"

            elif dias_restantes <= 5 and estado != "FINALIZADO":
                por_vencer_5 += 1
                por_vencer_15 += 1
                por_vencer_30 += 1
                alerta = f"Vence en {dias_restantes} días"
                color_alerta = "danger"

            elif dias_restantes <= 15 and estado != "FINALIZADO":
                por_vencer_15 += 1
                por_vencer_30 += 1
                alerta = f"Vence en {dias_restantes} días"
                color_alerta = "warning"

            elif dias_restantes <= 30 and estado != "FINALIZADO":
                por_vencer_30 += 1
                alerta = f"Vence en {dias_restantes} días"
                color_alerta = "info"

            else:
                alerta = f"Vence en {dias_restantes} días"
                color_alerta = "success"

        if estado != "FINALIZADO" and (
            dias_restantes is not None and dias_restantes <= 30
        ):
            contratos_alerta.append({
                "id": c[0],
                "codigo_proceso": c[1],
                "objeto_contratacion": c[2],
                "numero_contrato": c[3],
                "administrador_contrato": c[4],
                "fecha_fin_estimada": c[5],
                "estado": c[6],
                "dias_restantes": dias_restantes,
                "alerta": alerta,
                "color_alerta": color_alerta
            })

    return render_template(
        "seguimiento_contratos/dashboard_contratos.html",
        total_contratos=total_contratos,
        activos=activos,
        por_vencer_30=por_vencer_30,
        por_vencer_15=por_vencer_15,
        por_vencer_5=por_vencer_5,
        vencidos=vencidos,
        finalizados=finalizados,
        contratos_alerta=contratos_alerta
    )


# =========================
# NUEVO CONTRATO
# =========================
@main.route("/seguimiento_contratos/nuevo")
@login_required()
def seguimiento_contratos_nuevo():

    tarea_id = request.args.get(
        "tarea_id",
        type=int
    )

    conn = get_connection()
    cur = conn.cursor(
        cursor_factory=RealDictCursor
    )

    # ==========================================
    # DATOS DE LA TAREA DE ORIGEN
    # ==========================================
    tarea = None
    if tarea_id:
        cur.execute("""
            SELECT
                t.id,
                t.codigo_proceso,
                t.objeto_contratacion,
                t.unidad_solicitante,
                t.tipo_proceso,
                tp.nombre_proceso
            FROM tareas t
            LEFT JOIN tipo_procesos tp
                ON t.tipo_proceso = tp.id::TEXT

            WHERE t.id = %s
        """, (tarea_id,))

        tarea = cur.fetchone()
    # ==========================================
    # TIPOS DE PROCESO
    # ==========================================
    cur.execute("""
        SELECT
            id,
            nombre_proceso
        FROM tipo_procesos
        ORDER BY nombre_proceso
    """)

    tipos_proceso = cur.fetchall()
    # ==========================================
    # UNIDADES
    # ==========================================
    cur.execute("""
        SELECT
            id,
            nombre_unidad
        FROM unidades
        ORDER BY nombre_unidad
    """)
    unidades = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "seguimiento_contratos/seguimiento_contratos_form.html",

        tarea=tarea,
        tarea_id=tarea_id,

        tipos_proceso=tipos_proceso,
        unidades=unidades
    )

@main.route("/seguimiento_contratos/guardar", methods=["POST"])
@login_required()
def seguimiento_contratos_guardar():

    contrato_id = request.form.get("id")

    conn = get_connection()
    cur = conn.cursor()

    try:
        if contrato_id:
            cur.execute("""
                UPDATE seguimiento_contratos
                SET
                    codigo_proceso = %s,
                    objeto_contratacion = %s,
                    numero_contrato = %s,
                    proveedor = %s,
                    ruc = %s,
                    administrador_contrato = %s,
                    correo_administrador = %s,
                    fecha_suscripcion = %s,
                    fecha_inicio = %s,
                    fecha_fin_estimada = %s,
                    plazo_contractual = %s,
                    monto_contractual = %s,
                    unidad_requirente = %s,
                    tipo_procedimiento = %s,
                    estado = %s,
                    observaciones = %s
                WHERE id = %s
            """, (
                request.form.get("codigo_proceso"),
                request.form.get("objeto_contratacion"),
                request.form.get("numero_contrato"),
                request.form.get("proveedor"),
                request.form.get("ruc"),
                request.form.get("administrador_contrato"),
                request.form.get("correo_administrador"),
                request.form.get("fecha_suscripcion") or None,
                request.form.get("fecha_inicio") or None,
                request.form.get("fecha_fin_estimada") or None,
                request.form.get("plazo_contractual") or None,
                request.form.get("monto_contractual") or 0,
                request.form.get("unidad_requirente"),
                request.form.get("tipo_procedimiento"),
                request.form.get("estado"),
                request.form.get("observaciones"),
                contrato_id
            ))

            flash("✅ Contrato actualizado correctamente", "success")

        else:
            cur.execute("""
                INSERT INTO seguimiento_contratos (
                    codigo_proceso,
                    objeto_contratacion,
                    numero_contrato,
                    proveedor,
                    ruc,
                    administrador_contrato,
                    correo_administrador,
                    fecha_suscripcion,
                    fecha_inicio,
                    fecha_fin_estimada,
                    plazo_contractual,
                    monto_contractual,
                    unidad_requirente,
                    tipo_procedimiento,
                    estado,
                    observaciones
                )
                VALUES (
                    %s,%s,%s,%s,%s,%s,%s,%s,
                    %s,%s,%s,%s,%s,%s,%s,%s
                )
            """, (
                request.form.get("codigo_proceso"),
                request.form.get("objeto_contratacion"),
                request.form.get("numero_contrato"),
                request.form.get("proveedor"),
                request.form.get("ruc"),
                request.form.get("administrador_contrato"),
                request.form.get("correo_administrador"),
                request.form.get("fecha_suscripcion") or None,
                request.form.get("fecha_inicio") or None,
                request.form.get("fecha_fin_estimada") or None,
                request.form.get("plazo_contractual") or None,
                request.form.get("monto_contractual") or 0,
                request.form.get("unidad_requirente"),
                request.form.get("tipo_procedimiento"),
                request.form.get("estado"),
                request.form.get("observaciones")
            ))

            flash("✅ Contrato registrado correctamente", "success")

        conn.commit()

        return redirect(url_for("main.seguimiento_contratos"))

    except Exception as e:
        conn.rollback()
        flash(f"❌ Error al guardar contrato: {e}", "danger")
        return redirect(request.referrer)

    finally:
        cur.close()
        conn.close()
# =========================
# DETALLE PROFESIONAL DEL CONTRATO
# =========================
@main.route("/seguimiento_contratos/<int:contrato_id>/seguimientos")
@login_required()
def seguimiento_contratos_detalle(contrato_id):

    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            sc.id,
            sc.codigo_proceso,
            sc.objeto_contratacion,
            sc.numero_contrato,
            sc.proveedor,
            sc.ruc,
            sc.administrador_contrato,
            sc.correo_administrador,
            sc.fecha_suscripcion,
            sc.fecha_inicio,
            sc.fecha_fin_estimada,
            sc.plazo_contractual,
            sc.monto_contractual,
            sc.unidad_requirente,
            sc.tipo_procedimiento,
            sc.estado,
            sc.observaciones
        FROM seguimiento_contratos sc
        WHERE sc.id = %s
    """, (contrato_id,))
    contrato = cur.fetchone()
    cur.execute("""
        SELECT
            id,
            fecha_seguimiento,
            tipo_gestion,
            descripcion,
            compromiso,
            fecha_compromiso,
            estado
        FROM seguimiento_contrato_detalle
        WHERE contrato_id = %s
        ORDER BY fecha_seguimiento DESC, id DESC
    """, (contrato_id,))
    seguimientos = cur.fetchall()

    cur.execute("""
        SELECT
            id,
            numero_memorando,
            fecha_memorando,
            asunto,
            descripcion,
            archivo_pdf,
            fecha_registro
        FROM contrato_memorandos
        WHERE contrato_id = %s
        ORDER BY fecha_memorando DESC, id DESC
    """, (contrato_id,))

    memorandos = cur.fetchall()

    cur.execute("""
        SELECT
            id,
            tipo_comunicacion,
            fecha_comunicacion,
            asunto,
            participantes,
            descripcion,
            archivo_pdf
        FROM contrato_comunicaciones
        WHERE contrato_id = %s
        ORDER BY fecha_comunicacion DESC, id DESC
    """, (contrato_id,))

    comunicaciones = cur.fetchall()


    cur.execute("""
        SELECT
            id,
            tipo_informe,
            numero_informe,
            fecha_informe,
            asunto,
            descripcion,
            archivo_pdf
        FROM contrato_informes
        WHERE contrato_id = %s
        ORDER BY fecha_informe DESC, id DESC
    """, (contrato_id,))

    informes = cur.fetchall()

    cur.close()
    conn.close()

    if not contrato:
        abort(404)
   
    return render_template(
        "seguimiento_contratos/seguimiento_contratos_detalle.html",
        contrato=contrato,
        seguimientos=seguimientos,
        memorandos=memorandos,
        comunicaciones=comunicaciones,
        informes=informes
    )
# =========================
# NUEVO SEGUIMIENTO
# =========================
@main.route("/seguimiento_contratos/<int:contrato_id>/seguimientos/nuevo")
@login_required()
def seguimiento_nuevo(contrato_id):

    return render_template(
        "seguimiento_contratos/seguimiento_nuevo.html",
        contrato_id=contrato_id
    )
# =========================
# EDITAR SEGUIMIENTO CONTRATO
# =========================
@main.route("/seguimiento_contratos/editar/<int:contrato_id>")
@login_required()
def seguimiento_contratos_editar(contrato_id):
    conn = get_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    cur.execute("""
        SELECT *
        FROM seguimiento_contratos
        WHERE id = %s
    """, (contrato_id,))
    contrato = cur.fetchone()

    cur.execute("""
        SELECT id, nombre_unidad
        FROM unidades
        ORDER BY nombre_unidad
    """)
    unidades = cur.fetchall()

    cur.execute("""
        SELECT id, nombre_proceso
        FROM tipo_procesos
        ORDER BY nombre_proceso
    """)
    tipos_proceso = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "seguimiento_contratos/seguimiento_contratos_form.html",
        contrato=contrato,
        unidades=unidades,
        tipos_proceso=tipos_proceso
    )

# =========================
# NUEVO MEMORANDO
# =========================
@main.route("/seguimiento_contratos/<int:contrato_id>/memorando/nuevo")
@login_required()
def memorando_nuevo(contrato_id):

    return render_template(
        "seguimiento_contratos/memorando_form.html",
        contrato_id=contrato_id
    )
# =========================
# NUEVA COMUNICACIÓN
# =========================
@main.route("/seguimiento_contratos/<int:contrato_id>/comunicacion/nueva")
@login_required()
def comunicacion_nueva(contrato_id):

    return render_template(
        "seguimiento_contratos/comunicacion_form.html",
        contrato_id=contrato_id
    )
# =========================
# NUEVO INFORME
# =========================
@main.route("/seguimiento_contratos/<int:contrato_id>/informe/nuevo")
@login_required()
def informe_nuevo(contrato_id):

    return render_template(
        "seguimiento_contratos/informe_form.html",
        contrato_id=contrato_id
    )


# =========================
# GUARDAR INFORME
# =========================
@main.route("/seguimiento_contratos/<int:contrato_id>/informe/guardar", methods=["POST"])
@login_required()
def informe_guardar(contrato_id):

    conn = get_connection()
    cur = conn.cursor()

    try:
        archivo_pdf = None

        if "archivo_pdf" in request.files:
            archivo = request.files["archivo_pdf"]

            if archivo.filename:
                from werkzeug.utils import secure_filename

                carpeta = os.path.join(
                    "app",
                    "static",
                    "uploads",
                    "informes"
                )

                os.makedirs(carpeta, exist_ok=True)

                nombre_archivo = secure_filename(archivo.filename)

                archivo.save(
                    os.path.join(carpeta, nombre_archivo)
                )

                archivo_pdf = nombre_archivo

        cur.execute("""
            INSERT INTO contrato_informes (
                contrato_id,
                tipo_informe,
                numero_informe,
                fecha_informe,
                asunto,
                descripcion,
                archivo_pdf
            )
            VALUES (%s,%s,%s,%s,%s,%s,%s)
        """, (
            contrato_id,
            request.form.get("tipo_informe"),
            request.form.get("numero_informe"),
            request.form.get("fecha_informe"),
            request.form.get("asunto"),
            request.form.get("descripcion"),
            archivo_pdf
        ))

        conn.commit()
        flash("✅ Informe registrado correctamente", "success")

        return redirect(
            url_for(
                "main.seguimiento_contratos_detalle",
                contrato_id=contrato_id
            )
        )

    except Exception as e:
        conn.rollback()
        flash(f"❌ Error al guardar informe: {e}", "danger")
        return redirect(request.referrer)

    finally:
        cur.close()
        conn.close()
# =========================
# GUARDAR MEMORANDO
# =========================
@main.route("/seguimiento_contratos/<int:contrato_id>/memorando/guardar", methods=["POST"])
@login_required()
def memorando_guardar(contrato_id):

    conn = get_connection()
    cur = conn.cursor()

    try:
        archivo_binario = None
        archivo_nombre = None
        archivo_tipo = None

        archivo = request.files.get("archivo_pdf")

        if archivo and archivo.filename:
            from werkzeug.utils import secure_filename

            archivo_nombre = secure_filename(archivo.filename)
            archivo_tipo = archivo.content_type
            archivo_binario = archivo.read()

        cur.execute("""
            INSERT INTO contrato_memorandos (
                contrato_id,
                numero_memorando,
                fecha_memorando,
                asunto,
                descripcion,
                archivo_pdf,
                archivo_binario,
                archivo_nombre,
                archivo_tipo
            )
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            contrato_id,
            request.form.get("numero_memorando"),
            request.form.get("fecha_memorando"),
            request.form.get("asunto"),
            request.form.get("descripcion"),
            archivo_nombre,
            archivo_binario,
            archivo_nombre,
            archivo_tipo
        ))

        conn.commit()

        flash("✅ Memorando registrado correctamente", "success")

        return redirect(url_for(
            "main.seguimiento_contratos_detalle",
            contrato_id=contrato_id
        ))

    except Exception as e:
        conn.rollback()
        flash(f"❌ Error: {e}", "danger")
        return redirect(request.referrer)

    finally:
        cur.close()
        conn.close()
# =========================
# GUARDAR SEGUIMIENTO
# =========================
@main.route("/seguimiento_contratos/<int:contrato_id>/seguimientos/guardar", methods=["POST"])
@login_required()
def seguimiento_guardar(contrato_id):

    conn = get_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            INSERT INTO seguimiento_contrato_detalle (
                contrato_id,
                fecha_seguimiento,
                tipo_gestion,
                descripcion,
                compromiso,
                fecha_compromiso,
                estado
            )
            VALUES (%s,%s,%s,%s,%s,%s,%s)
        """, (
            contrato_id,
            request.form.get("fecha_seguimiento"),
            request.form.get("tipo_gestion"),
            request.form.get("descripcion"),
            request.form.get("compromiso"),
            request.form.get("fecha_compromiso") or None,
            request.form.get("estado")
        ))

        conn.commit()
        flash("✅ Seguimiento registrado correctamente", "success")

        return redirect(
            url_for("main.seguimiento_contratos_detalle", contrato_id=contrato_id)
        )

    except Exception as e:
        conn.rollback()
        flash(f"❌ Error al guardar seguimiento: {e}", "danger")
        return redirect(request.referrer)

    finally:
        cur.close()
        conn.close()

# =========================
# GUARDAR COMUNICACIÓN
# =========================
@main.route(
    "/seguimiento_contratos/<int:contrato_id>/comunicacion/guardar",
    methods=["POST"]
)
@login_required()
def comunicacion_guardar(contrato_id):

    conn = get_connection()
    cur = conn.cursor()

    try:

        archivo_pdf = None

        if "archivo_pdf" in request.files:

            archivo = request.files["archivo_pdf"]

            if archivo.filename:

                from werkzeug.utils import secure_filename

                carpeta = os.path.join(
                    "app",
                    "static",
                    "uploads",
                    "comunicaciones"
                )

                os.makedirs(carpeta, exist_ok=True)

                nombre_archivo = secure_filename(
                    archivo.filename
                )

                archivo.save(
                    os.path.join(carpeta, nombre_archivo)
                )

                archivo_pdf = nombre_archivo

        cur.execute("""
            INSERT INTO contrato_comunicaciones (
                contrato_id,
                tipo_comunicacion,
                fecha_comunicacion,
                asunto,
                participantes,
                descripcion,
                archivo_pdf
            )
            VALUES (%s,%s,%s,%s,%s,%s,%s)
        """, (

            contrato_id,
            request.form.get("tipo_comunicacion"),
            request.form.get("fecha_comunicacion"),
            request.form.get("asunto"),
            request.form.get("participantes"),
            request.form.get("descripcion"),
            archivo_pdf

        ))

        conn.commit()

        flash(
            "✅ Comunicación registrada correctamente",
            "success"
        )

        return redirect(
            url_for(
                "main.seguimiento_contratos_detalle",
                contrato_id=contrato_id
            )
        )

    except Exception as e:

        conn.rollback()

        flash(f"❌ Error: {e}", "danger")

        return redirect(request.referrer)

    finally:

        cur.close()
        conn.close()

# ================================
# INFORME DE VERIFICACIÓN (AUTOMÁTICO)
# ================================
@main.route('/informe/verificacion/<int:id_tarea>')
@login_required()
def informe_verificacion(id_tarea):

    conn = get_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    try:

        # ==========================================
        # TAREA + TIPO DE PROCESO
        # + ESTRUCTURA INSTITUCIONAL
        # ==========================================
        cur.execute("""
            SELECT
                t.*,

                tp.nombre_proceso
                    AS tipo_proceso_nombre,

                u.nombre_unidad
                    AS unidad_requirente_oficial,

                u.departamento_principal,
                u.bloque

            FROM tareas t

            LEFT JOIN tipo_procesos tp
                ON t.tipo_proceso::INTEGER = tp.id

            LEFT JOIN requerimientos r
                ON r.id = t.requerimiento_id

            LEFT JOIN unidades u
                ON u.id = r.unid_requirente

            WHERE t.id = %s

        """, (id_tarea,))

        tarea = cur.fetchone()

    finally:

        cur.close()
        conn.close()


    # ==========================================
    # VALIDAR TAREA
    # ==========================================
    if not tarea:
        abort(404)


    # ==========================================
    # CÓDIGO DEL INFORME
    # ==========================================
    year = datetime.now().year

    codigo_verificacion = (
        f"UCP-VERF-{year}-"
        f"{str(id_tarea).zfill(4)}"
    )


    # ==========================================
    # GENERAR INFORME
    # ==========================================
    return render_template(
        'verificaciones/informe_verificacion.html',

        fecha=datetime.now().strftime('%d/%m/%Y'),

        codigo_verificacion=
            codigo_verificacion,


        # ======================================
        # ESTRUCTURA INSTITUCIONAL
        # ======================================
        unidad_solicitante=
            tarea['unidad_requirente_oficial']
            or tarea['unidad_solicitante'],

        departamento_principal=
            tarea['departamento_principal']
            or "",

        bloque=
            tarea['bloque']
            or "",

        # ======================================
        # DATOS DEL PROCESO
        # ======================================
        funcionario_encargado=
            tarea['funcionario_encargado'],

        objeto_contratacion=
            tarea['objeto_contratacion'],

        codigo_proceso=
            tarea['codigo_proceso'],

        tipo_proceso=
            tarea['tipo_proceso_nombre'],

        tipo_compra=
            tarea['tipo_compra'],

        valor_sin_iva=
            tarea['valor_sin_iva'],

        valor_exento=
            tarea['valor_exento'],

        valor_en_letras=
            tarea['valor_en_letras'],


        # ======================================
        # INFORMACIÓN COMPLEMENTARIA
        # ======================================
        base_legal=
            tarea['base_legal']
            or "No registrada",

        observaciones=
            "La documentación cumple con "
            "los requisitos formales.",

        nombre_jefe_compras=
            tarea['nombre_jefe_compras'],


        # ======================================
        # VERIFICACIÓN DOCUMENTAL
        # ======================================
        presenta_estudio_previo=
            tarea['presenta_estudio_previo'],

        presenta_terminos_referencia=
            tarea['presenta_terminos_referencia'],

        presenta_estudio_mercado=
            tarea['presenta_estudio_mercado'],

        presenta_especificaciones=
            tarea['presenta_especificaciones'],

        presenta_proformas=
            tarea['presenta_proformas'],

        determinacion_necesidad=
            tarea['determinacion_necesidad'],

        consta_catalogo_electronico=
            tarea['consta_catalogo_electronico'],

        consta_poa=
            tarea['consta_poa'],

        consta_pac=
            tarea['consta_pac'],

        presenta_errores=
            tarea['presenta_errores'],

        cumple_normativa=
            tarea['cumple_normativa'],


        # ======================================
        # SOLO PARA OBRAS
        # ======================================
        presenta_planos=
            tarea['presenta_planos'],

        presenta_apus=
            tarea['presenta_apus'],

        presenta_condiciones_contratacion=
            tarea[
                'presenta_condiciones_contratacion'
            ],


        # ======================================
        # SOLO CEP
        # ======================================
        presenta_viabilidad_tecnico_economica=
            tarea[
                'presenta_viabilidad_tecnico_economica'
            ],


        # ======================================
        # PARA CONDICIONAR VISTAS
        # ======================================
        tipo_proceso_nombre=
            tarea['tipo_proceso_nombre'],

        tipo_regimen=
            tarea['tipo_regimen']
    )

# ==========================================
# CERRAR SESIÓN
# ==========================================
@main.route("/logout")
def logout():

    session.clear()

    response = redirect(
        url_for("main.inicio")
    )

    response.headers[
        "Cache-Control"
    ] = "no-store, no-cache, must-revalidate, max-age=0"

    response.headers[
        "Pragma"
    ] = "no-cache"

    response.headers[
        "Expires"
    ] = "0"

    return response
# =========================
# EXPEDIENTE ELECTRÓNICO
# =========================
@main.route("/seguimiento_contratos/<int:contrato_id>/expediente")
@login_required()
def expediente_contrato(contrato_id):

    conn = get_connection()
    cur = conn.cursor()

    # CONTRATO
    cur.execute("""
        SELECT *
        FROM seguimiento_contratos
        WHERE id = %s
    """, (contrato_id,))
    contrato = cur.fetchone()
    
    # MEMORANDOS
    cur.execute("""
        SELECT *
        FROM contrato_memorandos
        WHERE contrato_id = %s
        ORDER BY fecha_memorando DESC
    """, (contrato_id,))
    memorandos = cur.fetchall()
   
    # COMUNICACIONES
    cur.execute("""
        SELECT *
        FROM contrato_comunicaciones
        WHERE contrato_id = %s
        ORDER BY fecha_comunicacion DESC
    """, (contrato_id,))
    comunicaciones = cur.fetchall()

    # INFORMES
    cur.execute("""
        SELECT *
        FROM contrato_informes
        WHERE contrato_id = %s
        ORDER BY fecha_informe DESC
    """, (contrato_id,))
    informes = cur.fetchall()

    # SEGUIMIENTOS
    cur.execute("""
        SELECT *
        FROM seguimiento_contrato_detalle
        WHERE contrato_id = %s
        ORDER BY fecha_seguimiento DESC
    """, (contrato_id,))
    seguimientos = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "seguimiento_contratos/expediente_contrato.html",
        contrato=contrato,
        memorandos=memorandos,
        comunicaciones=comunicaciones,
        informes=informes,
        seguimientos=seguimientos
    )
# =========================
# LISTAR UNIDADES
# =========================
@main.route("/unidades")
@login_required()
def unidades_listar():
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, nombre_unidad
        FROM unidades
        ORDER BY nombre_unidad
    """)
    unidades = cur.fetchall()

    cur.close()
    conn.close()

    return render_template("catalogos/unidades_list.html", unidades=unidades)


# =========================
# NUEVA UNIDAD
# =========================
@main.route("/unidades/nueva")
@login_required()
def unidades_nueva():
    return render_template("catalogos/unidades_form.html", unidad=None)


# =========================
# GUARDAR UNIDAD
# =========================
@main.route("/unidades/guardar", methods=["POST"])
@login_required()
def unidades_guardar():
    nombre_unidad = request.form.get("nombre_unidad")

    conn = get_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            INSERT INTO unidades (nombre_unidad)
            VALUES (%s)
        """, (nombre_unidad,))

        conn.commit()
        flash("✅ Unidad registrada correctamente", "success")

    except Exception as e:
        conn.rollback()
        flash(f"❌ Error al guardar unidad: {e}", "danger")

    finally:
        cur.close()
        conn.close()

    return redirect(url_for("main.unidades_listar"))
# =========================
# EDITAR UNIDAD
# =========================
@main.route("/unidades/editar/<int:id>")
@login_required()
def unidades_editar(id):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, nombre_unidad
        FROM unidades
        WHERE id = %s
    """, (id,))
    unidad = cur.fetchone()

    cur.close()
    conn.close()

    if not unidad:
        abort(404)

    return render_template(
        "catalogos/unidades_form.html",
        unidad=unidad
    )


# =========================
# ACTUALIZAR UNIDAD
# =========================
@main.route("/unidades/actualizar/<int:id>", methods=["POST"])
@login_required()
def unidades_actualizar(id):
    nombre_unidad = request.form.get("nombre_unidad")

    conn = get_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            UPDATE unidades
            SET nombre_unidad = %s
            WHERE id = %s
        """, (nombre_unidad, id))

        conn.commit()
        flash("✅ Unidad actualizada correctamente", "success")

    except Exception as e:
        conn.rollback()
        flash(f"❌ Error al actualizar unidad: {e}", "danger")

    finally:
        cur.close()
        conn.close()

    return redirect(url_for("main.unidades_listar"))


# =========================
# ELIMINAR UNIDAD
# =========================
@main.route("/unidades/eliminar/<int:id>")
@login_required()
def unidades_eliminar(id):
    conn = get_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            DELETE FROM unidades
            WHERE id = %s
        """, (id,))

        conn.commit()
        flash("✅ Unidad eliminada correctamente", "success")

    except Exception as e:
        conn.rollback()
        flash(f"❌ No se pudo eliminar la unidad: {e}", "danger")

    finally:
        cur.close()
        conn.close()

    return redirect(url_for("main.unidades_listar"))
# =========================
# LISTAR TIPOS DE PROCESO
# =========================
@main.route("/tipo_procesos")
@login_required()
def tipo_procesos_listar():
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, nombre_proceso
        FROM tipo_procesos
        ORDER BY nombre_proceso
    """)

    tipos = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "catalogos/tipo_procesos_list.html",
        tipos=tipos
    )


# =========================
# NUEVO TIPO DE PROCESO
# =========================
@main.route("/tipo_procesos/nuevo")
@login_required()
def tipo_procesos_nuevo():
    return render_template(
        "catalogos/tipo_procesos_form.html",
        tipo=None
    )


# =========================
# GUARDAR TIPO DE PROCESO
# =========================
@main.route("/tipo_procesos/guardar", methods=["POST"])
@login_required()
def tipo_procesos_guardar():

    nombre_proceso = request.form.get("nombre_proceso")

    conn = get_connection()
    cur = conn.cursor()

    try:

        cur.execute("""
            INSERT INTO tipo_procesos (
                nombre_proceso
            )
            VALUES (%s)
        """, (nombre_proceso,))

        conn.commit()

        flash(
            "✅ Tipo de proceso registrado correctamente",
            "success"
        )

    except Exception as e:

        conn.rollback()

        flash(
            f"❌ Error: {e}",
            "danger"
        )

    finally:

        cur.close()
        conn.close()

    return redirect(
        url_for("main.tipo_procesos_listar")
    )


# =========================
# EDITAR TIPO DE PROCESO
# =========================
@main.route("/tipo_procesos/editar/<int:id>")
@login_required()
def tipo_procesos_editar(id):

    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, nombre_proceso
        FROM tipo_procesos
        WHERE id = %s
    """, (id,))

    tipo = cur.fetchone()

    cur.close()
    conn.close()

    return render_template(
        "catalogos/tipo_procesos_form.html",
        tipo=tipo
    )


# =========================
# ACTUALIZAR TIPO DE PROCESO
# =========================
@main.route("/tipo_procesos/actualizar/<int:id>", methods=["POST"])
@login_required()
def tipo_procesos_actualizar(id):

    nombre_proceso = request.form.get("nombre_proceso")

    conn = get_connection()
    cur = conn.cursor()

    try:

        cur.execute("""
            UPDATE tipo_procesos
            SET nombre_proceso = %s
            WHERE id = %s
        """, (
            nombre_proceso,
            id
        ))

        conn.commit()

        flash(
            "✅ Tipo de proceso actualizado correctamente",
            "success"
        )

    except Exception as e:

        conn.rollback()

        flash(
            f"❌ Error: {e}",
            "danger"
        )

    finally:

        cur.close()
        conn.close()

    return redirect(
        url_for("main.tipo_procesos_listar")
    )


# =========================
# ELIMINAR TIPO DE PROCESO
# =========================
@main.route("/tipo_procesos/eliminar/<int:id>")
@login_required()
def tipo_procesos_eliminar(id):

    conn = get_connection()
    cur = conn.cursor()

    try:

        cur.execute("""
            DELETE FROM tipo_procesos
            WHERE id = %s
        """, (id,))

        conn.commit()

        flash(
            "✅ Tipo de proceso eliminado correctamente",
            "success"
        )

    except Exception as e:

        conn.rollback()

        flash(
            f"❌ Error: {e}",
            "danger"
        )

    finally:

        cur.close()
        conn.close()

    return redirect(
        url_for("main.tipo_procesos_listar")
    )
# ===============================
# SEGUIMIENTO DE TAREAS
# ===============================
@main.route("/seguimiento_tareas")
@login_required()
def seguimiento_tareas():

    codigo = request.args.get("codigo", "").strip()
    estado = request.args.get("estado", "").strip()
    unidad = request.args.get("unidad", "").strip()
    funcionario = request.args.get("funcionario", "").strip()

    conn = get_connection()
    cur = conn.cursor()

    sql = """
        SELECT
            id,
            codigo_proceso,
            objeto_contratacion,
            unidad_solicitante,
            funcionario_encargado,
            estado_requerimiento,
            fecha_recepcion,
            CURRENT_DATE - fecha_recepcion AS dias_tramite,
            tipo_proceso
        FROM tareas
        WHERE 1=1
    """

    params = []

    if estado:
        sql += " AND estado_requerimiento ILIKE %s"
        params.append(f"%{estado}%")

    if unidad:
        sql += " AND unidad_solicitante ILIKE %s"
        params.append(f"%{unidad}%")

    if funcionario:
        sql += " AND funcionario_encargado ILIKE %s"
        params.append(f"%{funcionario}%")

    if codigo:
        sql += " AND codigo_proceso ILIKE %s"
        params.append(f"%{codigo}%")

    sql += """
        ORDER BY
            fecha_recepcion DESC NULLS LAST,
            id DESC
    """

    cur.execute(sql, params)
    tareas = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "seguimiento_tareas/seguimiento_tareas.html",
        tareas=tareas,
        codigo=codigo,
        estado=estado,
        unidad=unidad,
        funcionario=funcionario
    )
# ==========================================
# API - PARTIDAS PRESUPUESTARIAS DE UNA TAREA
# ==========================================
@main.route("/api/tarea/<int:tarea_id>/partidas")
@login_required()
def api_tarea_partidas(tarea_id):

    conn = get_connection()
    cur = conn.cursor()

    try:

        cur.execute("""
            SELECT
                p.id,
                p.nombre_part,
                p.num_part,
                p.programa,
                p.fuente,
                p.monto,

                COALESCE(
                    ap.monto_adjudicado,
                    0
                ) AS monto_adjudicado,
            ap.numero_orden_compra

            FROM tareas t

            INNER JOIN partidas p
                ON p.requerimiento_id = t.requerimiento_id

            LEFT JOIN LATERAL (
                SELECT
                    a.id
                FROM adjudicaciones a
                WHERE a.tarea_id = t.id
                ORDER BY a.id DESC
                LIMIT 1
            ) a ON TRUE

            LEFT JOIN adjudicacion_partidas ap
                ON ap.adjudicacion_id = a.id
               AND ap.partida_id = p.id

            WHERE t.id = %s

            ORDER BY p.id
        """, (tarea_id,))

        filas = cur.fetchall()

        partidas = []

        for p in filas:

            partidas.append({
                "id": p[0],
                "nombre": p[1] or "",
                "numero": p[2] or "",
                "programa": p[3] or "",
                "fuente": p[4] or "",
                "monto": float(p[5] or 0),
                "monto_adjudicado": float(p[6] or 0),
                "numero_orden_compra": p[7] or ""
            })

        return jsonify(partidas)

    finally:

        cur.close()
        conn.close()
# ===============================
# GUARDAR SEGUIMIENTO DE TAREA
# ===============================
@main.route("/seguimiento_tareas/guardar", methods=["POST"])
@login_required()
def seguimiento_tareas_guardar():

    tarea_id = request.form.get("tarea_id")
    estado = request.form.get("estado")
    observacion = request.form.get("observacion")
    numero_certificacion = request.form.get("numero_certificacion")
    fecha_certificacion = request.form.get("fecha_certificacion")
    valor_certificacion = request.form.get("valor_certificacion")

    certificacion_plurianual = (
        request.form.get("certificacion_plurianual") == "1"
    )

    anios_plurianuales = request.form.getlist("anio_plurianual[]")
    montos_plurianuales = request.form.getlist("monto_plurianual[]")

    # Partidas mostradas en el formulario de adjudicación
    partida_ids = request.form.getlist("partida_id[]")
    montos_adjudicados = request.form.getlist("monto_adjudicado[]")

    conn = get_connection()
    cur = conn.cursor()

    try:

        # ==========================================
        # 1. CONSULTAR DATOS DE LA TAREA
        # ==========================================
        cur.execute("""
            SELECT
                codigo_proceso,
                tipo_proceso,
                estado_requerimiento,
                numero_certificacion
            FROM tareas
            WHERE id = %s
        """, (tarea_id,))

        tarea = cur.fetchone()

        if not tarea:
            raise ValueError("No se encontró la tarea.")

        codigo_proceso = tarea[0] or ""
        tipo_proceso = tarea[1] or ""
        estado_anterior = tarea[2] or ""
        numero_certificacion_anterior = tarea[3] or ""

        # ==========================================
        # DETECTAR EDICIÓN DEL MISMO ESTADO
        # ==========================================
        editando_mismo_estado = (
            estado == estado_anterior
            and estado in (
                "CON CERTIFICACION",
                "ADJUDICADA",
                "ORDEN DE COMPRA ENVIADA"
            )
        )
        codigo_normalizado = codigo_proceso.upper().strip()

        es_catalogo = codigo_normalizado.startswith("CATE-")
        es_infima = codigo_normalizado.startswith("IC-")


        # ==========================================
        # 2. GUARDAR HISTORIAL DEL SEGUIMIENTO
        # ==========================================
        # Si estamos editando una certificación existente,
        # NO crear otro movimiento de seguimiento.
        if not editando_mismo_estado:

            cur.execute("""
                INSERT INTO seguimiento_tareas (
                    tarea_id,
                    estado,
                    observacion,
                    usuario_id
                )
                VALUES (%s, %s, %s, %s)
            """, (
                tarea_id,
                estado,
                observacion,
                session.get("user_id")
            ))
        # ==========================================
        # 3. ACTUALIZAR ESTADO / CERTIFICACIÓN
        # ==========================================
        if estado == "CON CERTIFICACION":

            # ------------------------------------------
            # VALIDAR VALOR DE CERTIFICACIÓN
            # ------------------------------------------
            try:
                valor_certificacion_decimal = Decimal(
                    str(valor_certificacion or 0)
                )
            except Exception:
                raise ValueError(
                    "El valor de la certificación no es válido."
                )

            if valor_certificacion_decimal < 0:
                raise ValueError(
                    "El valor de la certificación no puede ser negativo."
                )


            # ------------------------------------------
            # SI ES PLURIANUAL, VALIDAR AÑOS Y MONTOS
            # ------------------------------------------
            detalle_plurianual = []

            if certificacion_plurianual:

                if not anios_plurianuales:
                    raise ValueError(
                        "Debe ingresar al menos un año "
                        "para la certificación plurianual."
                    )

                total_plurianual = Decimal("0.00")
                anios_usados = set()

                for indice, anio_texto in enumerate(anios_plurianuales):

                    anio_texto = (anio_texto or "").strip()

                    monto_texto = (
                        montos_plurianuales[indice]
                        if indice < len(montos_plurianuales)
                        else ""
                    )

                    monto_texto = (monto_texto or "").strip()

                    # Ignorar una fila completamente vacía
                    if not anio_texto and not monto_texto:
                        continue

                    if not anio_texto:
                        raise ValueError(
                            "Existe un monto plurianual sin año."
                        )

                    try:
                        anio = int(anio_texto)
                    except ValueError:
                        raise ValueError(
                            f"El año '{anio_texto}' no es válido."
                        )

                    if anio in anios_usados:
                        raise ValueError(
                            f"El año {anio} está repetido."
                        )

                    anios_usados.add(anio)

                    try:
                        monto = Decimal(
                            str(monto_texto or 0)
                        )
                    except Exception:
                        raise ValueError(
                            f"El monto del año {anio} no es válido."
                        )

                    if monto < 0:
                        raise ValueError(
                            f"El monto del año {anio} "
                            f"no puede ser negativo."
                        )

                    detalle_plurianual.append(
                        (anio, monto)
                    )

                    total_plurianual += monto


                if not detalle_plurianual:
                    raise ValueError(
                        "Debe registrar al menos un año "
                        "con su monto plurianual."
                    )


                # ==========================================
                # TOTAL GENERAL DE LA CERTIFICACIÓN
                # Ejercicio actual + otros ejercicios
                # ==========================================
                total_certificacion_general = (
                    valor_certificacion_decimal
                    + total_plurianual
                )


            # ------------------------------------------
            # GUARDAR / EDITAR CERTIFICACIÓN EN TAREAS
            # ------------------------------------------
            cur.execute("""
                UPDATE tareas
                SET
                    estado_requerimiento = %s,
                    numero_certificacion = %s,
                    fecha_certificacion = %s,
                    valor_certificacion = %s,
                    certificacion_plurianual = %s
                WHERE id = %s
            """, (
                estado,
                numero_certificacion,
                fecha_certificacion or None,
                valor_certificacion_decimal,
                certificacion_plurianual,
                tarea_id
            ))


            # ------------------------------------------
            # BORRAR DISTRIBUCIÓN ANTERIOR
            # Esto permite EDITAR sin duplicar años
            # ------------------------------------------
            cur.execute("""
                DELETE FROM certificaciones_plurianuales
                WHERE tarea_id = %s
            """, (tarea_id,))


            # ------------------------------------------
            # VOLVER A GUARDAR DISTRIBUCIÓN ACTUAL
            # ------------------------------------------
            if certificacion_plurianual:

                for anio, monto in detalle_plurianual:

                    cur.execute("""
                        INSERT INTO certificaciones_plurianuales (
                            tarea_id,
                            anio,
                            monto,
                            usuario_id
                        )
                        VALUES (%s, %s, %s, %s)
                    """, (
                        tarea_id,
                        anio,
                        monto,
                        session.get("user_id")
                    ))


        else:

            cur.execute("""
                UPDATE tareas
                SET estado_requerimiento = %s
                WHERE id = %s
            """, (
                estado,
                tarea_id
            ))
     
        # ==========================================
        # 4. DETERMINAR SI DEBE GUARDAR ADJUDICACIÓN
        # ==========================================
        guardar_adjudicacion = False
        tipo_formalizacion = None

        # Catálogo e Ínfima:
        # guardamos cuando ya existe Orden de Compra
        if (
            (es_catalogo or es_infima)
            and estado == "ORDEN DE COMPRA ENVIADA"
        ):
            guardar_adjudicacion = True

            if es_catalogo:
                tipo_formalizacion = "CATALOGO_ELECTRONICO"
            else:
                tipo_formalizacion = "ORDEN_COMPRA"

        # Subasta, Licitación y demás:
        # guardamos cuando se adjudica
        elif (
            not es_catalogo
            and not es_infima
            and estado == "ADJUDICADA"
        ):
            guardar_adjudicacion = True
            tipo_formalizacion = "ADJUDICACION"


        # ==========================================
        # 5. GUARDAR CABECERA Y PARTIDAS
        # ==========================================
        if guardar_adjudicacion and partida_ids:

            # Buscar si ya existe una adjudicación
            # para evitar duplicarla si se corrige el registro.
            cur.execute("""
                SELECT id
                FROM adjudicaciones
                WHERE tarea_id = %s
                ORDER BY id DESC
                LIMIT 1
            """, (tarea_id,))

            adjudicacion_existente = cur.fetchone()

            if adjudicacion_existente:

                adjudicacion_id = adjudicacion_existente[0]

                cur.execute("""
                    UPDATE adjudicaciones
                    SET
                        tipo_formalizacion = %s,
                        usuario_id = %s
                    WHERE id = %s
                """, (
                    tipo_formalizacion,
                    session.get("user_id"),
                    adjudicacion_id
                ))

                # Borramos el detalle anterior para volver
                # a registrar la información corregida.
                cur.execute("""
                    DELETE FROM adjudicacion_partidas
                    WHERE adjudicacion_id = %s
                """, (adjudicacion_id,))

            else:

                cur.execute("""
                    INSERT INTO adjudicaciones (
                        tarea_id,
                        tipo_formalizacion,
                        usuario_id
                    )
                    VALUES (%s, %s, %s)
                    RETURNING id
                """, (
                    tarea_id,
                    tipo_formalizacion,
                    session.get("user_id")
                ))

                adjudicacion_id = cur.fetchone()[0]


            # ==========================================
            # 6. GUARDAR CADA PARTIDA
            # ==========================================
            for indice, partida_id in enumerate(partida_ids):

                monto_adjudicado = 0

                if indice < len(montos_adjudicados):
                    try:
                        monto_adjudicado = float(
                            montos_adjudicados[indice] or 0
                        )
                    except (ValueError, TypeError):
                        monto_adjudicado = 0


                # ======================================
                # ÍNFIMA CUANTÍA
                # Una sola orden por partida
                # ======================================
                numero_orden = None

                if es_infima:

                    ordenes_infima = request.form.getlist(
                        "numero_orden_compra[]"
                    )

                    if indice < len(ordenes_infima):
                        numero_orden = (
                            ordenes_infima[indice].strip()
                            or None
                        )


                cur.execute("""
                    INSERT INTO adjudicacion_partidas (
                        adjudicacion_id,
                        partida_id,
                        monto_adjudicado,
                        numero_orden_compra
                    )
                    VALUES (%s, %s, %s, %s)
                    RETURNING id
                """, (
                    adjudicacion_id,
                    partida_id,
                    monto_adjudicado,
                    numero_orden
                ))

                adjudicacion_partida_id = cur.fetchone()[0]


                # ======================================
                # CATÁLOGO ELECTRÓNICO
                # Varias órdenes por una misma partida
                # ======================================
                if es_catalogo:

                    numeros_orden = request.form.getlist(
                        f"orden_catalogo_numero_{partida_id}[]"
                    )

                    montos_orden = request.form.getlist(
                        f"orden_catalogo_monto_{partida_id}[]"
                    )

                    for pos, numero in enumerate(numeros_orden):

                        numero = (numero or "").strip()

                        if not numero:
                            continue

                        monto_orden = 0

                        if pos < len(montos_orden):
                            try:
                                monto_orden = float(
                                    montos_orden[pos] or 0
                                )
                            except (ValueError, TypeError):
                                monto_orden = 0

                        cur.execute("""
                            INSERT INTO adjudicacion_ordenes (
                                adjudicacion_partida_id,
                                numero_orden_compra,
                                monto_orden
                            )
                            VALUES (%s, %s, %s)
                        """, (
                            adjudicacion_partida_id,
                            numero,
                            monto_orden
                        ))


        conn.commit()

        flash(
            "✅ Seguimiento registrado correctamente",
            "success"
        )

       
    except Exception as e:
    
        conn.rollback()

        print("ERROR GUARDANDO SEGUIMIENTO:", e)

        flash(
            f"❌ Error al registrar seguimiento: {e}",
            "danger"
        )

    finally:

        cur.close()
        conn.close()

    return redirect(
        url_for("main.seguimiento_tareas")
    )

# ===============================
# HISTORIAL SEGUIMIENTO DE TAREA
# ===============================

@main.route("/seguimiento_tareas/historial/<int:tarea_id>")
@login_required()
def seguimiento_tareas_historial(tarea_id):

    conn = get_connection()
    cur = conn.cursor()

    try:

        # ==========================================
        # 1. DATOS DE LA TAREA
        # ==========================================
        cur.execute("""
            SELECT
                id,
                codigo_proceso,
                objeto_contratacion,
                unidad_solicitante,
                funcionario_encargado,
                estado_requerimiento,
                numero_certificacion,
                fecha_certificacion,
                valor_certificacion,
                certificacion_plurianual
            FROM tareas
            WHERE id = %s
        """, (tarea_id,))

        tarea = cur.fetchone()

        if not tarea:
            flash(
                "No se encontró la tarea.",
                "danger"
            )

            return redirect(
                url_for("main.seguimiento_tareas")
            )


        # ==========================================
        # 2. CERTIFICACIÓN PLURIANUAL
        # ==========================================
        cur.execute("""
            SELECT
                anio,
                monto
            FROM certificaciones_plurianuales
            WHERE tarea_id = %s
            ORDER BY anio
        """, (tarea_id,))

        plurianuales = cur.fetchall()


        # ==========================================
        # 3. TOTALES DE LA CERTIFICACIÓN
        # ==========================================
        otros_ejercicios = sum(
            Decimal(str(p[1] or 0))
            for p in plurianuales
        )

        valor_actual = Decimal(
            str(tarea[8] or 0)
        )

        total_certificacion = (
            valor_actual + otros_ejercicios
        )


        # ==========================================
        # 4. HISTORIAL DE MOVIMIENTOS
        # ==========================================
        cur.execute("""
            SELECT
                s.fecha,
                s.estado,
                s.observacion,
                u.nombre
            FROM seguimiento_tareas s
            LEFT JOIN usuarios u
                ON s.usuario_id = u.id
            WHERE s.tarea_id = %s
            ORDER BY s.fecha DESC
        """, (tarea_id,))

        seguimientos = cur.fetchall()
        # ==========================================
        # ÓRDENES DE COMPRA - ÍNFIMA CUANTÍA
        # ==========================================
        cur.execute("""
            SELECT
                fecha,
                numero_oc,
                total,
                proveedor
            FROM ordenes_compra
            WHERE tarea_id = %s
            ORDER BY fecha DESC, id DESC
        """, (tarea_id,))

        ordenes_compra = cur.fetchall()
        # ==========================================
        # ÓRDENES DE CATÁLOGO ELECTRÓNICO
        # ==========================================
        cur.execute("""
            SELECT
                oc.fecha_aceptacion,
                oc.numero_orden,
                oc.monto_adjudicado,
                oc.proveedor
            FROM ordenes_catalogo oc
            JOIN catalogos_electronicos ce
                ON ce.id = oc.catalogo_id
            WHERE ce.tarea_id = %s
            ORDER BY oc.fecha_aceptacion DESC, oc.id DESC
        """, (tarea_id,))

        ordenes_catalogo = cur.fetchall()
        # ==========================================
        # 7. CONTRATOS VINCULADOS AL PROCESO
        # ==========================================
        cur.execute("""
            SELECT
                fecha_suscripcion,
                numero_contrato,
                proveedor,
                monto_contractual,
                estado
            FROM seguimiento_contratos
            WHERE codigo_proceso = %s
            ORDER BY fecha_suscripcion DESC, id DESC
        """, (tarea[1],))

        contratos = cur.fetchall()
        
        
        
        # ==========================================
        # 5. MOSTRAR HISTORIAL
        # ==========================================
        return render_template(
            "seguimiento_tareas/historial.html",

            tarea=tarea,
            seguimientos=seguimientos,

            plurianuales=plurianuales,
            otros_ejercicios=otros_ejercicios,
            valor_actual=valor_actual,
            total_certificacion=total_certificacion,
            ordenes_compra=ordenes_compra,
            ordenes_catalogo=ordenes_catalogo,
            contratos=contratos
        )

    finally:

        cur.close()
        conn.close()
@main.route("/seguimiento_tareas/dashboard")
@login_required()
def seguimiento_tareas_dashboard():
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            COUNT(*) AS total,

            COUNT(*) FILTER (
                WHERE estado_requerimiento NOT ILIKE '%FINALIZADA%'
                AND estado_requerimiento NOT ILIKE '%ANULADA%'
            ) AS en_tramite,

            COUNT(*) FILTER (
                WHERE estado_requerimiento ILIKE '%FINALIZADA%'
            ) AS finalizadas,

            COUNT(*) FILTER (
                WHERE fecha_recepcion IS NOT NULL
                AND CURRENT_DATE - fecha_recepcion > 10
                AND estado_requerimiento NOT ILIKE '%FINALIZADA%'
                AND estado_requerimiento NOT ILIKE '%ANULADA%'
            ) AS atrasadas,

            COUNT(*) FILTER (
                WHERE estado_requerimiento ILIKE '%CERTIFICACIÓN%'
                OR estado_requerimiento ILIKE '%CERTIFICACION%'
            ) AS certificacion,

            COUNT(*) FILTER (
                WHERE estado_requerimiento ILIKE '%ORDEN DE COMPRA%'
            ) AS orden_compra,

            COUNT(*) FILTER (
                WHERE estado_requerimiento ILIKE '%PLIEGO%'
            ) AS pliegos,

            COUNT(*) FILTER (
                WHERE estado_requerimiento ILIKE '%ADJUDICADA%'
            ) AS adjudicadas,

            COUNT(*) FILTER (
                WHERE estado_requerimiento ILIKE '%OBSERVADA%'
            ) AS observadas,

            COUNT(*) FILTER (
                WHERE estado_requerimiento ILIKE '%ANULADA%'
            ) AS anuladas
        FROM tareas
    """)
    stats = cur.fetchone()

    cur.execute("""
        SELECT
            id,
            codigo_proceso,
            objeto_contratacion,
            estado_requerimiento,
            fecha_recepcion,
            CURRENT_DATE - fecha_recepcion AS dias
        FROM tareas
        WHERE fecha_recepcion IS NOT NULL
          AND CURRENT_DATE - fecha_recepcion > 10
          AND estado_requerimiento NOT ILIKE '%FINALIZADA%'
          AND estado_requerimiento NOT ILIKE '%ANULADA%'
        ORDER BY dias DESC
        LIMIT 20
    """)
    alertas = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "seguimiento_tareas/dashboard.html",
        stats=stats,
        alertas=alertas
    )
# ===============================
# DASHBOARD EJECUTIVO SICOP
# ===============================
@main.route("/dashboard_ejecutivo")
@login_required()
def dashboard_ejecutivo():

    unidad = (request.args.get("unidad") or "").strip()
    fecha_inicio = (request.args.get("fecha_inicio") or "").strip()
    fecha_fin = (request.args.get("fecha_fin") or "").strip()

    conn = get_connection()
    cur = conn.cursor()

    try:

        # =====================================================
        # 1. FILTROS GENERALES
        # =====================================================
        condiciones_generales = [
            "t.fecha_recepcion IS NOT NULL"
        ]

        parametros = []

        if unidad:
            condiciones_generales.append(
                "t.unidad_solicitante = %s"
            )
            parametros.append(unidad)

        if fecha_inicio:
            condiciones_generales.append(
                "t.fecha_recepcion >= %s"
            )
            parametros.append(fecha_inicio)

        if fecha_fin:
            condiciones_generales.append(
                "t.fecha_recepcion <= %s"
            )
            parametros.append(fecha_fin)

        where_general = (
            " WHERE "
            + " AND ".join(condiciones_generales)
        )


        # =====================================================
        # 2. FILTRO ECONÓMICO
        #
        # NO PARTICIPAN:
        # - DEVUELTOS
        # - VERIFICACION PRODUCCION NACIONAL
        # - ARRENDAMIENTOS BIENES MUEBLES
        # =====================================================
        condiciones_economicas = list(
            condiciones_generales
        )

        condiciones_economicas.append("""
            UPPER(
                TRIM(
                    COALESCE(
                        t.estado_requerimiento,
                        ''
                    )
                )
            ) <> 'DEVUELTO'
        """)

        condiciones_economicas.append("""
            UPPER(
                TRIM(
                    COALESCE(
                        tp.nombre_proceso,
                        ''
                    )
                )
            ) NOT IN (
                'VERIFICACION DE PRODUCCION NACIONAL',
                'ARRENDAMIENTOS DE BIENES MUEBLES'
            )
        """)

        where_economico = (
            " WHERE "
            + " AND ".join(condiciones_economicas)
        )


        # =====================================================
        # 3. RESUMEN ECONÓMICO GENERAL
        # =====================================================
        #
        # adjudicado_tarea:
        # suma una sola vez todo lo adjudicado por tarea.
        #
        # IMPORTANTE:
        # adjudicacion_partidas es nuestra fuente económica
        # central para evitar contar nuevamente órdenes.
        # =====================================================
        sql_resumen = f"""
            WITH adjudicado_tarea AS (
                SELECT
                    a.tarea_id,
                    COALESCE(
                        SUM(ap.monto_adjudicado),
                        0
                    ) AS monto_adjudicado
                FROM adjudicaciones a
                LEFT JOIN adjudicacion_partidas ap
                    ON ap.adjudicacion_id = a.id
                GROUP BY a.tarea_id
            ),

            base AS (
                SELECT
                    t.id,
                    t.fecha_recepcion,
                    t.estado_requerimiento,
                    t.unidad_solicitante,

                    COALESCE(
                        t.valor_sin_iva,
                        0
                    )
                    +
                    COALESCE(
                        t.valor_exento,
                        0
                    ) AS monto_ingresado,
                    COALESCE(
                        t.valor_certificacion,
                        0
                    ) AS monto_certificado,
                    COALESCE(
                        at.monto_adjudicado,
                        0
                    ) AS monto_adjudicado,

                    CASE
                        WHEN at.tarea_id IS NOT NULL
                        THEN TRUE
                        ELSE FALSE
                    END AS tiene_adjudicacion

                FROM tareas t

                LEFT JOIN tipo_procesos tp
                    ON t.tipo_proceso = tp.id::TEXT

                LEFT JOIN adjudicado_tarea at
                    ON at.tarea_id = t.id

                {where_economico}
            )

            SELECT
                COUNT(*) AS procesos_validos,

                COALESCE(
                    SUM(monto_ingresado),
                    0
                ) AS monto_ingresado,
                COALESCE(
                    SUM(monto_certificado),
                    0
                ) AS monto_certificado,
                COALESCE(
                    SUM(monto_adjudicado),
                    0
                ) AS monto_adjudicado,

                COALESCE(
                    SUM(monto_ingresado)
                    -
                    SUM(monto_adjudicado),
                    0
                ) AS diferencia_total,

                CASE
                    WHEN COALESCE(
                        SUM(monto_ingresado),
                        0
                    ) > 0
                    THEN ROUND(
                        (
                            SUM(monto_adjudicado)
                            /
                            SUM(monto_ingresado)
                        ) * 100,
                        2
                    )
                    ELSE 0
                END AS porcentaje_adjudicado,

                ROUND(
                    AVG(
                        CURRENT_DATE
                        - fecha_recepcion
                    ),
                    2
                ) AS promedio_dias,

                COUNT(*) FILTER (
                    WHERE tiene_adjudicacion
                ) AS procesos_adjudicados,

                COALESCE(
                    SUM(
                        CASE
                            WHEN tiene_adjudicacion
                            THEN (
                                monto_ingresado
                                -
                                monto_adjudicado
                            )
                            ELSE 0
                        END
                    ),
                    0
                ) AS diferencia_solo_adjudicados

            FROM base
        """

        cur.execute(
            sql_resumen,
            parametros
        )

        resumen = cur.fetchone()


        # =====================================================
        # 4. PROCESOS DEVUELTOS
        # Se muestran aparte pero NO participan económicamente
        # =====================================================
        sql_devueltos = f"""
            SELECT
                COUNT(*) AS total,

                COALESCE(
                    SUM(
                        COALESCE(
                            t.valor_sin_iva,
                            0
                        )
                        +
                        COALESCE(
                            t.valor_exento,
                            0
                        )
                    ),
                    0
                ) AS monto

            FROM tareas t

            LEFT JOIN tipo_procesos tp
                ON t.tipo_proceso = tp.id::TEXT

            {where_general}

            AND UPPER(
                TRIM(
                    COALESCE(
                        t.estado_requerimiento,
                        ''
                    )
                )
            ) = 'DEVUELTO'
        """

        cur.execute(
            sql_devueltos,
            parametros
        )

        devueltos = cur.fetchone()


        # =====================================================
        # 5. PROCESOS EXCLUIDOS POR TIPO
        # VPN + ARRENDAMIENTOS
        # =====================================================
        sql_excluidos = f"""
            SELECT
                COUNT(*) AS total,

                COALESCE(
                    SUM(
                        COALESCE(
                            t.valor_sin_iva,
                            0
                        )
                        +
                        COALESCE(
                            t.valor_exento,
                            0
                        )
                    ),
                    0
                ) AS monto

            FROM tareas t

            LEFT JOIN tipo_procesos tp
                ON t.tipo_proceso = tp.id::TEXT

            {where_general}

            AND UPPER(
                TRIM(
                    COALESCE(
                        tp.nombre_proceso,
                        ''
                    )
                )
            ) IN (
                'VERIFICACION DE PRODUCCION NACIONAL',
                'ARRENDAMIENTOS DE BIENES MUEBLES'
            )
        """

        cur.execute(
            sql_excluidos,
            parametros
        )

        excluidos = cur.fetchone()


        # =====================================================
        # 6. MONTO POR UNIDAD
        # =====================================================
        sql_unidad = f"""
            WITH adjudicado_tarea AS (
                SELECT
                    a.tarea_id,
                    COALESCE(
                        SUM(ap.monto_adjudicado),
                        0
                    ) AS adjudicado
                FROM adjudicaciones a
                LEFT JOIN adjudicacion_partidas ap
                    ON ap.adjudicacion_id = a.id
                GROUP BY a.tarea_id
            )

            SELECT
                COALESCE(
                    t.unidad_solicitante,
                    'SIN UNIDAD'
                ) AS unidad,

                COUNT(*) AS procesos,

                COALESCE(
                    SUM(
                        COALESCE(
                            t.valor_sin_iva,
                            0
                        )
                        +
                        COALESCE(
                            t.valor_exento,
                            0
                        )
                    ),
                    0
                ) AS ingresado,

                COALESCE(
                    SUM(
                        COALESCE(
                            at.adjudicado,
                            0
                        )
                    ),
                    0
                ) AS adjudicado,

                COALESCE(
                    SUM(
                        COALESCE(
                            t.valor_sin_iva,
                            0
                        )
                        +
                        COALESCE(
                            t.valor_exento,
                            0
                        )
                    ),
                    0
                )
                -
                COALESCE(
                    SUM(
                        COALESCE(
                            at.adjudicado,
                            0
                        )
                    ),
                    0
                ) AS diferencia

            FROM tareas t

            LEFT JOIN tipo_procesos tp
                ON t.tipo_proceso = tp.id::TEXT

            LEFT JOIN adjudicado_tarea at
                ON at.tarea_id = t.id

            {where_economico}

            GROUP BY
                t.unidad_solicitante

            ORDER BY
                ingresado DESC
        """

        cur.execute(
            sql_unidad,
            parametros
        )

        por_unidad = cur.fetchall()


        # =====================================================
        # 7. MONTO POR PROCEDIMIENTO
        # =====================================================
        sql_procedimiento = f"""
            WITH adjudicado_tarea AS (
                SELECT
                    a.tarea_id,
                    COALESCE(
                        SUM(ap.monto_adjudicado),
                        0
                    ) AS adjudicado
                FROM adjudicaciones a
                LEFT JOIN adjudicacion_partidas ap
                    ON ap.adjudicacion_id = a.id
                GROUP BY a.tarea_id
            )

            SELECT
                COALESCE(
                    tp.nombre_proceso,
                    'SIN TIPO DE PROCESO'
                ) AS procedimiento,

                COUNT(*) AS procesos,

                COALESCE(
                    SUM(
                        COALESCE(
                            t.valor_sin_iva,
                            0
                        )
                        +
                        COALESCE(
                            t.valor_exento,
                            0
                        )
                    ),
                    0
                ) AS ingresado,

                COALESCE(
                    SUM(
                        COALESCE(
                            at.adjudicado,
                            0
                        )
                    ),
                    0
                ) AS adjudicado,

                COALESCE(
                    SUM(
                        CASE
                            WHEN at.tarea_id IS NOT NULL
                            THEN
                                COALESCE(t.valor_sin_iva, 0)
                                +
                                COALESCE(t.valor_exento, 0)
                                -
                                COALESCE(at.adjudicado, 0)
                            ELSE 0
                        END
                    ),
                    0
                ) AS diferencia

            FROM tareas t

            LEFT JOIN tipo_procesos tp
                ON t.tipo_proceso = tp.id::TEXT

            LEFT JOIN adjudicado_tarea at
                ON at.tarea_id = t.id

            {where_economico}

            GROUP BY
                tp.nombre_proceso

            ORDER BY
                ingresado DESC
        """

        cur.execute(
            sql_procedimiento,
            parametros
        )

        por_procedimiento = cur.fetchall()


        # =====================================================
        # 8. DISTRIBUCIÓN POR PROGRAMA
        #
        # Aquí usamos partidas.monto porque una tarea puede
        # tener varias partidas y no debemos repetir el monto
        # total de la tarea en cada una.
        # =====================================================
        sql_programa = f"""
            SELECT
                COALESCE(
                    p.programa,
                    'SIN PROGRAMA'
                ) AS programa,

                COUNT(
                    DISTINCT t.id
                ) AS procesos,

                COALESCE(
                    SUM(p.monto),
                    0
                ) AS ingresado,

                COALESCE(
                    SUM(
                        ap.monto_adjudicado
                    ),
                    0
                ) AS adjudicado,

                COALESCE(
                    SUM(p.monto),
                    0
                )
                -
                COALESCE(
                    SUM(
                        ap.monto_adjudicado
                    ),
                    0
                ) AS diferencia

            FROM tareas t

            LEFT JOIN tipo_procesos tp
                ON t.tipo_proceso = tp.id::TEXT

            INNER JOIN partidas p
                ON p.requerimiento_id =
                   t.requerimiento_id

            LEFT JOIN adjudicaciones a
                ON a.tarea_id = t.id

            LEFT JOIN adjudicacion_partidas ap
                ON ap.adjudicacion_id = a.id
               AND ap.partida_id = p.id

            {where_economico}

            GROUP BY
                p.programa

            ORDER BY
                ingresado DESC
        """

        cur.execute(
            sql_programa,
            parametros
        )

        por_programa = cur.fetchall()


        # =====================================================
        # 9. DISTRIBUCIÓN POR PARTIDA
        # =====================================================
        sql_partida = f"""
            SELECT
                COALESCE(
                    p.num_part,
                    'SIN PARTIDA'
                ) AS partida,

                COALESCE(
                    p.nombre_part,
                    ''
                ) AS nombre,

                COUNT(
                    DISTINCT t.id
                ) AS procesos,

                COALESCE(
                    SUM(p.monto),
                    0
                ) AS ingresado,

                COALESCE(
                    SUM(
                        ap.monto_adjudicado
                    ),
                    0
                ) AS adjudicado,

                COALESCE(
                    SUM(p.monto),
                    0
                )
                -
                COALESCE(
                    SUM(
                        ap.monto_adjudicado
                    ),
                    0
                ) AS diferencia

            FROM tareas t

            LEFT JOIN tipo_procesos tp
                ON t.tipo_proceso = tp.id::TEXT

            INNER JOIN partidas p
                ON p.requerimiento_id =
                   t.requerimiento_id

            LEFT JOIN adjudicaciones a
                ON a.tarea_id = t.id

            LEFT JOIN adjudicacion_partidas ap
                ON ap.adjudicacion_id = a.id
               AND ap.partida_id = p.id

            {where_economico}

            GROUP BY
                p.num_part,
                p.nombre_part

            ORDER BY
                ingresado DESC
        """

        cur.execute(
            sql_partida,
            parametros
        )

        por_partida = cur.fetchall()


        # =====================================================
        # 10. DISTRIBUCIÓN POR FUENTE
        # =====================================================
        sql_fuente = f"""
            SELECT
                COALESCE(
                    p.fuente,
                    'SIN FUENTE'
                ) AS fuente,

                COUNT(
                    DISTINCT t.id
                ) AS procesos,

                COALESCE(
                    SUM(p.monto),
                    0
                ) AS ingresado,

                COALESCE(
                    SUM(
                        ap.monto_adjudicado
                    ),
                    0
                ) AS adjudicado,

                COALESCE(
                    SUM(p.monto),
                    0
                )
                -
                COALESCE(
                    SUM(
                        ap.monto_adjudicado
                    ),
                    0
                ) AS diferencia

            FROM tareas t

            LEFT JOIN tipo_procesos tp
                ON t.tipo_proceso = tp.id::TEXT

            INNER JOIN partidas p
                ON p.requerimiento_id =
                   t.requerimiento_id

            LEFT JOIN adjudicaciones a
                ON a.tarea_id = t.id

            LEFT JOIN adjudicacion_partidas ap
                ON ap.adjudicacion_id = a.id
               AND ap.partida_id = p.id

            {where_economico}

            GROUP BY
                p.fuente

            ORDER BY
                ingresado DESC
        """

        cur.execute(
            sql_fuente,
            parametros
        )

        por_fuente = cur.fetchall()


        # =====================================================
        # 11. ESTADOS
        # Aquí mostramos TODOS, incluso DEVUELTOS.
        # Es información operativa, no cálculo económico.
        # =====================================================
        sql_estado = f"""
            SELECT
                COALESCE(
                    t.estado_requerimiento,
                    'SIN ESTADO'
                ) AS estado,

                COUNT(*) AS total

            FROM tareas t

            LEFT JOIN tipo_procesos tp
                ON t.tipo_proceso = tp.id::TEXT

            {where_general}

            GROUP BY
                t.estado_requerimiento

            ORDER BY
                total DESC
        """

        cur.execute(
            sql_estado,
            parametros
        )

        por_estado = cur.fetchall()


        # =====================================================
        # 12. ANALISTAS
        # =====================================================
        sql_funcionario = f"""
            SELECT
                COALESCE(
                    t.funcionario_encargado,
                    'SIN FUNCIONARIO'
                ) AS funcionario,

                COUNT(*) AS total

            FROM tareas t

            LEFT JOIN tipo_procesos tp
                ON t.tipo_proceso = tp.id::TEXT

            {where_general}

            GROUP BY
                t.funcionario_encargado

            ORDER BY
                total DESC
        """

        cur.execute(
            sql_funcionario,
            parametros
        )

        por_funcionario = cur.fetchall()


        # =====================================================
        # 13. UNIDADES PARA EL FILTRO
        # =====================================================
        cur.execute("""
            SELECT DISTINCT
                unidad_solicitante

            FROM tareas

            WHERE unidad_solicitante
                IS NOT NULL

              AND TRIM(
                    unidad_solicitante
                  ) <> ''

            ORDER BY
                unidad_solicitante
        """)

        unidades_filtro = [
            fila[0]
            for fila in cur.fetchall()
        ]


        # =====================================================
        # 14. ENVIAR AL DASHBOARD
        # =====================================================
        return render_template(
            "dashboard_ejecutivo.html",

            resumen=resumen,

            devueltos=devueltos,
            excluidos=excluidos,

            por_estado=por_estado,
            por_funcionario=por_funcionario,

            por_unidad=por_unidad,
            por_procedimiento=por_procedimiento,
            por_programa=por_programa,
            por_partida=por_partida,
            por_fuente=por_fuente,

            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            unidad=unidad,
            unidades_filtro=unidades_filtro
        )

    finally:

        cur.close()
        conn.close()
    # ==========================================
    # TAREAS POR ESTADO
    # ==========================================
    sql_estado = f"""
        SELECT
            COALESCE(
                estado_requerimiento,
                'SIN ESTADO'
            ) AS estado,

            COUNT(*) AS total

        FROM tareas

        {where_sql}

        GROUP BY estado_requerimiento
        ORDER BY total DESC
    """

    cur.execute(
        sql_estado,
        parametros
    )

    por_estado = cur.fetchall()

    # ==========================================
    # TAREAS POR ANALISTA
    # ==========================================
    sql_funcionario = f"""
        SELECT
            COALESCE(
                funcionario_encargado,
                'SIN FUNCIONARIO'
            ) AS funcionario,

            COUNT(*) AS total

        FROM tareas

        {where_sql}

        GROUP BY funcionario_encargado
        ORDER BY total DESC
        LIMIT 20
    """

    cur.execute(
        sql_funcionario,
        parametros
    )

    por_funcionario = cur.fetchall()

    # ==========================================
    # TAREAS POR UNIDAD
    # ==========================================
    sql_unidad = f"""
        SELECT
            COALESCE(
                unidad_solicitante,
                'SIN UNIDAD'
            ) AS unidad,

            COUNT(*) AS total

        FROM tareas

        {where_sql}

        GROUP BY unidad_solicitante
        ORDER BY total DESC
        LIMIT 20
    """

    cur.execute(
        sql_unidad,
        parametros
    )

    por_unidad = cur.fetchall()

    cur.execute("""
        SELECT DISTINCT unidad_solicitante
        FROM tareas
        WHERE unidad_solicitante IS NOT NULL
        AND TRIM(unidad_solicitante) <> ''
        ORDER BY unidad_solicitante
    """)

    unidades_filtro = [
        fila[0]
        for fila in cur.fetchall()
    ]

    cur.close()
    conn.close()

    return render_template(
        "dashboard_ejecutivo.html",
        resumen=resumen,
        por_estado=por_estado,
        por_funcionario=por_funcionario,
        por_unidad=por_unidad,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        unidad=unidad,
        unidades_filtro=unidades_filtro
    )
# =========================================
# DASHBOARD DE REQUERIMIENTOS POR UNIDAD
# =========================================
@main.route("/dashboard_requerimientos")
@login_required()
def dashboard_requerimientos():

    conn = get_connection()
    cur = conn.cursor()

    # Totales generales
    cur.execute("""
        SELECT
            COUNT(*) AS total_requerimientos,
            COALESCE(SUM(monto_req), 0) AS monto_total
        FROM requerimientos
    """)

    resumen = cur.fetchone()

    # Cantidad y monto por unidad requirente
    cur.execute("""
        SELECT
            COALESCE(
                u.nombre_unidad,
                'SIN UNIDAD'
            ) AS unidad,

            COUNT(r.id) AS cantidad_requerimientos,

            COALESCE(
                SUM(r.monto_req),
                0
            ) AS monto_total

        FROM requerimientos r

        LEFT JOIN unidades u
            ON u.id = r.unid_requirente

        GROUP BY
            u.id,
            u.nombre_unidad

        ORDER BY
            cantidad_requerimientos DESC,
            unidad ASC
    """)

    por_unidad = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "dashboard_requerimientos.html",
        resumen=resumen,
        por_unidad=por_unidad
    )

# ===============================
# REPORTES
# ===============================
@main.route("/reportes")
@login_required()
def reportes():
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, nombre_unidad
        FROM unidades
        ORDER BY nombre_unidad
    """)
    unidades = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "reportes/reportes.html",
        unidades=unidades
    )

@main.route("/reporte/procesos_periodo/pdf")
@login_required()
def reporte_procesos_periodo_pdf():

    fecha_desde = request.args.get("fecha_desde")
    fecha_hasta = request.args.get("fecha_hasta")

    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            codigo_proceso,
            objeto_contratacion,
            monto_contractual,
            estado
        FROM seguimiento_contratos
        WHERE fecha_registro::date BETWEEN %s AND %s
        ORDER BY fecha_registro DESC
    """, (fecha_desde, fecha_hasta))

    procesos = cur.fetchall()

    cur.close()
    conn.close()

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4)
    )

    elementos = []
    styles = getSampleStyleSheet()

    elementos.append(
        Paragraph(
            f"Reporte de Procesos del {fecha_desde} al {fecha_hasta}",
            styles["Title"]
        )
    )

    data = [[
        "Código Proceso",
        "Objeto Contratación",
        "Monto",
        "Estado"
    ]]

    for proceso in procesos:
        data.append([
            Paragraph(str(proceso[0]), styles["Normal"]),
            Paragraph(str(proceso[1]), styles["Normal"]),
            Paragraph(f"${proceso[2]:,.2f}", styles["Normal"]),
            Paragraph(str(proceso[3]), styles["Normal"])
        ])

    tabla = Table(data, colWidths=[110, 520, 90, 90])

    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))

    elementos.append(tabla)

    doc.build(elementos)

    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="reporte_procesos.pdf",
        mimetype="application/pdf"
    )
@main.route("/reporte/procesos_ingresados/pdf")
@login_required()
def reporte_procesos_ingresados_pdf():

    fecha_desde = request.args.get("fecha_desde")
    fecha_hasta = request.args.get("fecha_hasta")
    unidad_id = request.args.get("unidad_id")

    conn = get_connection()
    cur = conn.cursor()

    consulta = """
        SELECT
            r.memo_vice_ad,
            r.descripcion,
            u.nombre_unidad,
            r.fecha_recep_req,
            r.monto_req
        FROM requerimientos r
        LEFT JOIN unidades u
            ON u.id = r.unid_requirente
        WHERE r.fecha_recep_req BETWEEN %s AND %s
    """

    parametros = [fecha_desde, fecha_hasta]

    if unidad_id:
        consulta += """
            AND r.unid_requirente = %s
        """
        parametros.append(unidad_id)

    consulta += """
        ORDER BY r.fecha_recep_req ASC
    """

    cur.execute(consulta, parametros)
    procesos = cur.fetchall()
    cur.close()
    conn.close()

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=25,
        leftMargin=25,
        topMargin=25,
        bottomMargin=25
    )

    styles = getSampleStyleSheet()

    estilo_titulo = ParagraphStyle(
        "TituloProcesosIngresados",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=12,
        leading=15,
        alignment=1,
        spaceAfter=10
    )

    estilo_celda = ParagraphStyle(
        "CeldaProcesosIngresados",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=10
    )

    elementos = []

    elementos.append(
        Paragraph(
            "PROCESOS INGRESADOS A LA UNIDAD DE COMPRAS PÚBLICAS",
            estilo_titulo
        )
    )

    elementos.append(
        Paragraph(
            f"Período: {fecha_desde} al {fecha_hasta}",
            styles["Normal"]
        )
    )

    elementos.append(Spacer(1, 12))

    data = [[
        "#",
        "Oficio Vicerrectorado",
        "Descripción",
        "Unidad requirente",
        "Fecha de recepción",
        "Monto total"
    ]]

    total_general = 0

    for i, proceso in enumerate(procesos, start=1):
        monto = float(proceso[4] or 0)
        total_general += monto

        data.append([
            Paragraph(str(i), estilo_celda),
            Paragraph(str(proceso[0] or ""), estilo_celda),
            Paragraph(str(proceso[1] or ""), estilo_celda),
            Paragraph(str(proceso[2] or ""), estilo_celda),
            Paragraph(
                proceso[3].strftime("%d/%m/%Y") if proceso[3] else "",
                estilo_celda
            ),
            Paragraph(f"${monto:,.2f}", estilo_celda)
        ])

    data.append([
        "",
        "",
        "",
        "",
        Paragraph("<b>TOTAL GENERAL</b>", estilo_celda),
        Paragraph(f"<b>${total_general:,.2f}</b>", estilo_celda)
    ])

    tabla = Table(
        data,
        colWidths=[30, 135, 255, 160, 90, 90],
        repeatRows=1
    )

    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0b2f4f")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (4, 1), (4, -1), "RIGHT"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#eaf2f8")),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
    ]))

    elementos.append(tabla)
    doc.build(elementos)

    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="procesos_ingresados_compras_publicas.pdf",
        mimetype="application/pdf"
    )
# ==================================
# CÓDIGOS DE PROCESO OCUPADOS
# ==================================
@main.route("/tareas/codigos_ocupados")
@login_required()
def tareas_codigos_ocupados():
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            codigo_proceso,
            objeto_contratacion,
            funcionario_encargado,
            tipo_proceso,
            estado_requerimiento,
            fecha_recepcion,
            COALESCE(valor_sin_iva, 0) + COALESCE(valor_exento, 0) AS monto_total
        FROM tareas
        WHERE codigo_proceso IS NOT NULL
          AND TRIM(codigo_proceso) <> ''
        ORDER BY codigo_proceso ASC
    """)

    codigos = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "tareas/codigos_ocupados.html",
        codigos=codigos
    )
# ==========================================
# CERTIFICACIONES DE UNA TAREA
# ==========================================
@main.route("/tareas/<int:tarea_id>/certificaciones")
@login_required()
def tarea_certificaciones(tarea_id):
    conn = get_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    # Datos de la tarea
    cur.execute("""
        SELECT
            id,
            codigo_proceso,
            objeto_contratacion,
            funcionario_encargado,
            fecha_recepcion
        FROM tareas
        WHERE id = %s
    """, (tarea_id,))

    tarea = cur.fetchone()

    if not tarea:
        cur.close()
        conn.close()
        flash("❌ La tarea indicada no existe.", "danger")
        return redirect(url_for("main.tareas"))

    # Certificaciones guardadas
    cur.execute("""
        SELECT
            id,
            tipo_certificacion,
            fecha_certificacion,
            nombre_archivo,
            tipo_mime,
            fecha_registro
        FROM certificaciones_tareas
        WHERE tarea_id = %s
        ORDER BY fecha_registro DESC
    """, (tarea_id,))

    certificaciones = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "tareas/certificaciones.html",
        tarea=tarea,
        certificaciones=certificaciones
    )
# ==========================================
# GUARDAR CERTIFICACIÓN
# ==========================================
@main.route("/tareas/<int:tarea_id>/certificaciones/guardar", methods=["POST"])
@login_required()
def tarea_certificacion_guardar(tarea_id):

    conn = get_connection()
    cur = conn.cursor()

    try:

        tipo = request.form.get("tipo_certificacion")
        fecha = request.form.get("fecha_certificacion")
        archivos = request.files.getlist("imagenes")

        if not tipo:
            flash("Debe seleccionar el tipo de certificación.", "danger")
            return redirect(
                url_for("main.tarea_certificaciones", tarea_id=tarea_id)
            )

        archivos_validos = [
            archivo for archivo in archivos
            if archivo and archivo.filename
        ]

        if not archivos_validos:
            flash("Debe seleccionar al menos una imagen.", "danger")
            return redirect(
                url_for("main.tarea_certificaciones", tarea_id=tarea_id)
            )

        cur.execute("""
            INSERT INTO certificaciones_tareas (
                tarea_id,
                tipo_certificacion,
                fecha_certificacion,
                usuario_id
            )
            VALUES (%s, %s, %s, %s)
            RETURNING id
        """, (
            tarea_id,
            tipo,
            fecha,
            session.get("user_id")
        ))

        certificacion_id = cur.fetchone()[0]

        for archivo in archivos_validos:

            tipo_mime = archivo.mimetype or ""

            if tipo_mime not in ("image/png", "image/jpeg"):
                raise ValueError(
                    f"El archivo {archivo.filename} no es una imagen JPG o PNG válida."
                )

            imagen_bytes = archivo.read()

            cur.execute("""
                INSERT INTO certificaciones_imagenes (
                    certificacion_id,
                    nombre_archivo,
                    tipo_mime,
                    imagen
                )
                VALUES (%s, %s, %s, %s)
            """, (
                certificacion_id,
                archivo.filename,
                tipo_mime,
                psycopg2.Binary(imagen_bytes)
            ))

        conn.commit()

        flash("✅ Certificación guardada correctamente.", "success")

    except Exception as e:

        conn.rollback()
        print(e)

        flash(f"❌ Error: {e}", "danger")

    finally:

        cur.close()
        conn.close()

    return redirect(
        url_for(
            "main.tarea_certificaciones",
            tarea_id=tarea_id
        )
    )
# ==========================================
# ELIMINAR CERTIFICACIÓN
# ==========================================
@main.route(
    "/certificaciones/<int:certificacion_id>/eliminar",
    methods=["POST"]
)
@login_required()
def certificacion_eliminar(certificacion_id):

    conn = get_connection()
    cur = conn.cursor()

    try:
        # Primero obtenemos la tarea para poder regresar
        cur.execute("""
            SELECT tarea_id
            FROM certificaciones_tareas
            WHERE id = %s
        """, (certificacion_id,))

        fila = cur.fetchone()

        if not fila:
            flash("❌ La certificación no existe.", "danger")
            return redirect(url_for("main.tareas"))

        tarea_id = fila[0]

        # Las capturas se eliminan automáticamente por ON DELETE CASCADE,
        # pero lo dejamos explícito para mayor claridad.
        cur.execute("""
            DELETE FROM certificaciones_imagenes
            WHERE certificacion_id = %s
        """, (certificacion_id,))

        cur.execute("""
            DELETE FROM certificaciones_tareas
            WHERE id = %s
        """, (certificacion_id,))

        conn.commit()

        flash("✅ Certificación eliminada correctamente.", "success")

    except Exception as e:
        conn.rollback()
        print("ERROR AL ELIMINAR CERTIFICACIÓN:", e)

        flash(
            f"❌ No fue posible eliminar la certificación: {e}",
            "danger"
        )

        tarea_id = None

    finally:
        cur.close()
        conn.close()

    if tarea_id:
        return redirect(
            url_for(
                "main.tarea_certificaciones",
                tarea_id=tarea_id
            )
        )

    return redirect(url_for("main.tareas"))

# ==========================================
# PDF CERTIFICACIÓN CATE
# ==========================================
@main.route("/certificaciones/<int:certificacion_id>/cate/pdf")
@login_required()
def certificacion_cate_pdf(certificacion_id):

    conn = get_connection()
    cur = conn.cursor()

    # Datos generales de la certificación y de la tarea
    cur.execute("""
        SELECT
            c.id,
            c.fecha_certificacion,
            t.codigo_proceso,
            t.objeto_contratacion,

            COALESCE(
                tp.nombre_proceso,
                t.tipo_proceso
            ) AS tipo_proceso,

            t.funcionario_encargado,
            t.nombre_jefe_compras,
            t.consta_pac,

            -- ==========================================
            -- ESTRUCTURA INSTITUCIONAL
            -- ==========================================
            u.nombre_unidad AS unidad_requirente,
            u.departamento_principal,
            u.bloque

        FROM certificaciones_tareas c

        JOIN tareas t
            ON t.id = c.tarea_id

        LEFT JOIN tipo_procesos tp
            ON tp.id::text = TRIM(t.tipo_proceso)

        LEFT JOIN requerimientos r
            ON r.id = t.requerimiento_id

        LEFT JOIN unidades u
            ON u.id = r.unid_requirente

        WHERE c.id = %s
        AND c.tipo_certificacion = 'CATE'
    """, (certificacion_id,))

    datos = cur.fetchone()

    # Todas las capturas asociadas a la certificación
    cur.execute("""
        SELECT
            id,
            nombre_archivo,
            tipo_mime,
            imagen
        FROM certificaciones_imagenes
        WHERE certificacion_id = %s
        ORDER BY id ASC
    """, (certificacion_id,))

    capturas = cur.fetchall()

    cur.close()
    conn.close()

    if not datos:
        flash(
            "❌ No se encontró la Certificación de Catálogo Electrónico.",
            "danger"
        )
        return redirect(url_for("main.tareas"))

    return generar_pdf_cate(datos, capturas)
# ==========================================
# PDF CERTIFICACIÓN PAC
# ==========================================
@main.route("/certificaciones/<int:certificacion_id>/pac/pdf")
@login_required()
def certificacion_pac_pdf(certificacion_id):

    conn = get_connection()
    cur = conn.cursor()

    # ==========================================
    # DATOS DE LA CERTIFICACIÓN
    # ==========================================
    cur.execute("""
        SELECT
            c.id,
            c.fecha_certificacion,
            t.codigo_proceso,
            t.objeto_contratacion,
            COALESCE(tp.nombre_proceso, t.tipo_proceso) AS tipo_proceso,
            t.funcionario_encargado,
            t.nombre_jefe_compras,
            t.consta_pac,
            u.nombre_unidad,
            u.departamento_principal,
            u.bloque

        FROM certificaciones_tareas c

        JOIN tareas t
            ON t.id = c.tarea_id

        LEFT JOIN tipo_procesos tp
            ON tp.id::text = TRIM(t.tipo_proceso)

        LEFT JOIN requerimientos r
            ON r.id = t.requerimiento_id

        LEFT JOIN unidades u
            ON u.id = r.unid_requirente

        WHERE c.id = %s
          AND c.tipo_certificacion = 'PAC'
    """, (certificacion_id,))

    datos = cur.fetchone()


    # ==========================================
    # CAPTURAS DE RESPALDO
    # ==========================================
    cur.execute("""
        SELECT
            id,
            nombre_archivo,
            tipo_mime,
            imagen
        FROM certificaciones_imagenes
        WHERE certificacion_id = %s
        ORDER BY id ASC
    """, (certificacion_id,))

    capturas = cur.fetchall()


    cur.close()
    conn.close()


    if not datos:
        flash(
            "❌ No se encontró la Certificación PAC.",
            "danger"
        )
        return redirect(
            url_for("main.tareas")
        )


    return generar_pdf_pac(
        datos,
        capturas
    )
# ==========================================
# TRAZABILIDAD INTEGRAL DE PROCESOS
# ==========================================
@main.route("/trazabilidad")
@login_required()
def trazabilidad_procesos():
    return render_template(
        "trazabilidad/trazabilidad.html"
    )
# ==========================================
# API - BÚSQUEDA DINÁMICA DE TRAZABILIDAD
# ==========================================
@main.route("/api/trazabilidad/procesos")
@login_required()
def api_trazabilidad_procesos():

    texto = request.args.get("q", "").strip()

    if len(texto) < 2:
        return jsonify({
            "procesos": []
        })

    patron = f"%{texto}%"

    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            t.id,
            t.codigo_proceso,
            t.objeto_contratacion,
            t.estado_requerimiento,
            t.funcionario_encargado,
            t.fecha_recepcion,

            COALESCE(t.valor_sin_iva, 0)
            + COALESCE(t.valor_exento, 0) AS monto_total,

            COALESCE(u.nombre_unidad, '') AS unidad,

            oc.numero_oc,
            oc.fecha AS fecha_orden,
            oc.administrador_orden,
            oc.proveedor AS proveedor_orden,
            oc.ruc AS ruc_orden,
            oc.total AS total_orden,

            sc.numero_contrato,
            sc.fecha_suscripcion,
            sc.administrador_contrato,
            sc.proveedor AS proveedor_contrato,
            sc.ruc AS ruc_contrato,
            sc.monto_contractual,
            sc.estado AS estado_contrato,
                
            EXISTS (
                SELECT 1
                FROM certificaciones_tareas ct
                WHERE ct.tarea_id = t.id
                AND ct.tipo_certificacion = 'PAC'
            ) AS tiene_pac,

            EXISTS (
                SELECT 1
                FROM certificaciones_tareas ct
                WHERE ct.tarea_id = t.id
                AND ct.tipo_certificacion = 'CATE'
            ) AS tiene_cate
        FROM tareas t

        LEFT JOIN requerimientos r
            ON r.id = t.requerimiento_id

        LEFT JOIN unidades u
            ON u.id = r.unid_requirente

        LEFT JOIN LATERAL (
            SELECT
                o.numero_oc,
                o.fecha,
                o.administrador_orden,
                o.proveedor,
                o.ruc,
                o.total
            FROM ordenes_compra o
            WHERE o.tarea_id = t.id
            ORDER BY o.id DESC
            LIMIT 1
        ) oc ON TRUE

        LEFT JOIN LATERAL (
            SELECT
                c.numero_contrato,
                c.fecha_suscripcion,
                c.administrador_contrato,
                c.proveedor,
                c.ruc,
                c.monto_contractual,
                c.estado
            FROM seguimiento_contratos c
            WHERE UPPER(TRIM(c.codigo_proceso))
                = UPPER(TRIM(t.codigo_proceso))
            ORDER BY c.id DESC
            LIMIT 1
        ) sc ON TRUE

        WHERE
            t.codigo_proceso ILIKE %s
            OR t.objeto_contratacion ILIKE %s
            OR t.funcionario_encargado ILIKE %s
            OR u.nombre_unidad ILIKE %s
            OR oc.numero_oc ILIKE %s
            OR sc.numero_contrato ILIKE %s
            OR sc.administrador_contrato ILIKE %s
            OR sc.proveedor ILIKE %s

        ORDER BY t.fecha_recepcion DESC NULLS LAST
        LIMIT 30
    """, (
        patron,
        patron,
        patron,
        patron,
        patron,
        patron,
        patron,
        patron
    ))

    filas = cur.fetchall()

    cur.close()
    conn.close()

    procesos = []

    for fila in filas:

        fecha_recepcion = fila[5]

        monto = float(fila[6] or 0)

        numero_orden = fila[8]
        fecha_orden = fila[9]
        administrador_orden = fila[10]
        proveedor_orden = fila[11]
        ruc_orden = fila[12]
        total_orden = fila[13]

        numero_contrato = fila[14]
        fecha_suscripcion = fila[15]
        administrador_contrato = fila[16]
        proveedor_contrato = fila[17]
        ruc_contrato = fila[18]
        monto_contractual = fila[19]
        estado_contrato = fila[20]

        tiene_pac = bool(fila[21])
        tiene_cate = bool(fila[22])

        # ==========================================
        # PRIORIDAD DE DATOS:
        # 1. CONTRATO
        # 2. ORDEN DE COMPRA
        # ==========================================

        administrador = (
            administrador_contrato
            or administrador_orden
            or ""
        )

        proveedor = (
            proveedor_contrato
            or proveedor_orden
            or ""
        )

        ruc = (
            ruc_contrato
            or ruc_orden
            or ""
        )

        fecha_documento = (
            fecha_suscripcion
            or fecha_orden
        )

        if numero_contrato:
            tipo_documento = "CONTRATO"
            numero_documento = numero_contrato

        elif numero_orden:
            tipo_documento = "ORDEN DE COMPRA"
            numero_documento = numero_orden

        else:
            tipo_documento = ""
            numero_documento = ""


        procesos.append({
            "id": fila[0],
            "codigo_proceso": fila[1] or "",
            "objeto": fila[2] or "",
            "estado": fila[3] or "SIN ESTADO",
            "analista": fila[4] or "",
            "fecha_recepcion": (
                fecha_recepcion.strftime("%d/%m/%Y")
                if fecha_recepcion
                else ""
            ),
            "monto": monto,
            "monto_formateado": f"${monto:,.2f}",
            "unidad": fila[7] or "",

            "tiene_orden": bool(numero_orden),
            "numero_orden": numero_orden or "",

            "tiene_contrato": bool(numero_contrato),
            "numero_contrato": numero_contrato or "",

           "administrador": administrador,
            "proveedor": proveedor,
            "ruc": ruc,

            "estado_contrato": estado_contrato or "",

            "tipo_documento": tipo_documento,
            "numero_documento": numero_documento,

            "fecha_documento": (
                fecha_documento.strftime("%d/%m/%Y")
                if fecha_documento
                else ""
            ),

            "fecha_orden": (
                fecha_orden.strftime("%d/%m/%Y")
                if fecha_orden
                else ""
            ),

            "fecha_contrato": (
                fecha_suscripcion.strftime("%d/%m/%Y")
                if fecha_suscripcion
                else ""
            ),

            "tiene_pac": tiene_pac,
            "tiene_cate": tiene_cate
        })

    return jsonify({
        "procesos": procesos
    })
# ===============================
# VOLVER AL PANEL SEGÚN ROL
# ===============================
@main.route("/volver_panel")
@login_required()
def volver_panel():

    rol = session.get("rol", "").strip().lower()

    if rol == "administrador":
        return redirect(url_for("main.admin_dashboard"))

    elif rol == "analista":
        return redirect(url_for("main.analista_dashboard"))

    return redirect(url_for("main.user_dashboard"))
# ==========================================
# PUBLICACIONES DE NECESIDAD
# ==========================================
@main.route("/publicaciones_necesidad")
@login_required()
def publicaciones_necesidad():

    codigo_publicacion = request.args.get(
        "codigo_publicacion",
        ""
    ).strip()

    unidad_requirente = request.args.get(
        "unidad_requirente",
        ""
    ).strip()

    estado = request.args.get(
        "estado",
        ""
    ).strip()

    conn = get_connection()
    cur = conn.cursor()

    condiciones = []
    parametros = []

    if codigo_publicacion:

        condiciones.append(
            "p.codigo_publicacion ILIKE %s"
        )

        parametros.append(
            f"%{codigo_publicacion}%"
        )

    if unidad_requirente:

        condiciones.append(
            "p.unidad_requirente = %s"
        )

        parametros.append(
            unidad_requirente
        )

    if estado:

        condiciones.append(
            "p.estado = %s"
        )

        parametros.append(
            estado
        )

    where_sql = ""

    if condiciones:

        where_sql = (
            "WHERE "
            + " AND ".join(condiciones)
        )

    consulta = f"""
        SELECT
            p.id,
            p.codigo_publicacion,
            p.objeto_compra,
            p.unidad_requirente,
            p.fecha_publicacion,
            p.fecha_limite,
            p.numero_publicacion,
            p.estado,
            COUNT(pr.id) AS total_proformas
        FROM publicaciones_necesidad p

        LEFT JOIN proformas_publicacion pr
            ON pr.publicacion_id = p.id

        {where_sql}

        GROUP BY
            p.id,
            p.codigo_publicacion,
            p.objeto_compra,
            p.unidad_requirente,
            p.fecha_publicacion,
            p.fecha_limite,
            p.numero_publicacion,
            p.estado

        ORDER BY
            p.fecha_publicacion DESC,
            p.id DESC
    """

    cur.execute(
        consulta,
        tuple(parametros)
    )

    publicaciones = cur.fetchall()

    # ==========================================
    # UNIDADES PARA EL FILTRO
    # ==========================================
    cur.execute("""
        SELECT DISTINCT unidad_requirente
        FROM publicaciones_necesidad
        WHERE unidad_requirente IS NOT NULL
          AND unidad_requirente <> ''
        ORDER BY unidad_requirente
    """)

    unidades_filtro = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "publicaciones/publicaciones_list.html",
        publicaciones=publicaciones,
        unidades_filtro=unidades_filtro,
        codigo_publicacion=codigo_publicacion,
        unidad_requirente=unidad_requirente,
        estado=estado
    )

# ==========================================
# NUEVA PUBLICACIÓN DE NECESIDAD
# ==========================================
@main.route("/publicaciones_necesidad/nueva")
@login_required()
def publicacion_necesidad_nueva():

    conn = get_connection()
    cur = conn.cursor()
    # ==========================================
    # OBTENER EL SIGUIENTE NÚMERO DE SOLICITUD
    # ==========================================
    cur.execute("""
        SELECT COALESCE(
            MAX(CAST(numero_solicitud AS INTEGER)),
            0
        )
        FROM publicaciones_necesidad
        WHERE numero_solicitud ~ '^[0-9]+$'
    """)

    ultimo_numero = cur.fetchone()[0]

    numero_solicitud = ultimo_numero + 1
    # Unidades institucionales
    cur.execute("""
        SELECT
            id,
            nombre_unidad,
            departamento_principal,
            bloque,
            CASE
                WHEN UPPER(TRIM(nombre_unidad)) IN ('DECANATO', 'SUBDECANATO')
                    AND bloque IS NOT NULL
                    AND TRIM(bloque) <> ''
                THEN nombre_unidad || ' - ' || bloque
                ELSE nombre_unidad
            END AS nombre_visible
        FROM unidades
        ORDER BY nombre_visible
    """)

    unidades = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "publicaciones/publicacion_form.html",
        unidades=unidades,
        publicacion=None,
        items=[],
        numero_solicitud=numero_solicitud
    )
# ==========================================
# GUARDAR PUBLICACIÓN DE NECESIDAD
# ==========================================
@main.route("/publicaciones_necesidad/guardar", methods=["POST"])
@login_required()
def guardar_publicacion_necesidad():

    conn = get_connection()
    cur = conn.cursor()

    try:

        # ==========================================
        # DATOS DE LA PUBLICACIÓN
        # ==========================================
        numero_solicitud = request.form.get("numero_solicitud")
        objeto_compra = request.form.get("objeto_compra")
        fecha_publicacion = request.form.get("fecha_publicacion")
        fecha_limite = request.form.get("fecha_limite") or None

        encargado = request.form.get("encargado")
        correo = request.form.get("correo")
        tipo_publicacion = request.form.get("tipo_publicacion")
        unidad_requirente = request.form.get("unidad_requirente")

        codigo_publicacion = request.form.get("codigo_publicacion")
        numero_publicacion = request.form.get("numero_publicacion") or 1

        estado = request.form.get("estado")
        observaciones = request.form.get("observaciones") 

        notificado = request.form.get("notificado") == "on"
        oc_subida = request.form.get("oc_subida") == "on"


        # ==========================================
        # GUARDAR CABECERA DE LA PUBLICACIÓN
        # ==========================================
        cur.execute("""
            INSERT INTO publicaciones_necesidad (
                numero_solicitud,
                objeto_compra,
                fecha_publicacion,
                fecha_limite,
                encargado,
                correo,
                tipo_publicacion,
                unidad_requirente,
                codigo_publicacion,
                numero_publicacion,
                notificado,
                oc_subida,
                estado,
                observaciones,
                usuario_id
            )
            VALUES (
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s
            )
            RETURNING id
        """, (
            numero_solicitud,
            objeto_compra,
            fecha_publicacion,
            fecha_limite,
            encargado,
            correo,
            tipo_publicacion,
            unidad_requirente,
            codigo_publicacion,
            numero_publicacion,
            notificado,
            oc_subida,
            estado,
            observaciones,
            session.get("user_id")
        ))

        publicacion_id = cur.fetchone()[0]


        # ==========================================
        # OBTENER ITEMS DEL FORMULARIO
        # ==========================================
        cpcs = request.form.getlist("item_cpc[]")
        descripciones = request.form.getlist("item_descripcion[]")
        cantidades = request.form.getlist("item_cantidad[]")
        unidades_items = request.form.getlist("item_unidad[]")
        formas_pago = request.form.getlist("item_forma_pago[]")


        # ==========================================
        # GUARDAR ITEMS DE LA PUBLICACIÓN
        # ==========================================
        for i, descripcion in enumerate(descripciones):

            # No guardar filas completamente vacías
            if not descripcion.strip():
                continue

            cpc = cpcs[i] if i < len(cpcs) else None
            cantidad = cantidades[i] if i < len(cantidades) else None
            unidad_item = (
                unidades_items[i]
                if i < len(unidades_items)
                else None
            )
            forma_pago = (
                formas_pago[i]
                if i < len(formas_pago)
                else None
            )

            # Evitar error NUMERIC cuando cantidad venga vacía
            cantidad = cantidad or None

            cur.execute("""
                INSERT INTO publicacion_items (
                    publicacion_id,
                    cpc,
                    descripcion_producto,
                    cantidad,
                    unidad,
                    forma_pago
                )
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (
                publicacion_id,
                cpc,
                descripcion,
                cantidad,
                unidad_item,
                forma_pago
            ))


        # ==========================================
        # CONFIRMAR TODO
        # ==========================================
        conn.commit()

        flash(
            "✅ Publicación registrada correctamente.",
            "success"
        )


    except Exception as e:

        conn.rollback()

        print("ERROR GUARDANDO PUBLICACIÓN:", e)

        flash(
            f"❌ Error al registrar la publicación: {e}",
            "danger"
        )


    finally:

        cur.close()
        conn.close()


    return redirect(
        url_for("main.publicaciones_necesidad")
    )
# ==========================================
# VER DETALLE DE PUBLICACIÓN
# ==========================================
@main.route("/publicaciones_necesidad/<int:publicacion_id>")
@login_required()
def publicacion_necesidad_detalle(publicacion_id):

    conn = get_connection()
    cur = conn.cursor()

    # ==========================================
    # CABECERA
    # ==========================================
    cur.execute("""
        SELECT
            id,
            numero_solicitud,
            objeto_compra,
            fecha_publicacion,
            fecha_limite,
            encargado,
            correo,
            tipo_publicacion,
            unidad_requirente,
            codigo_publicacion,
            numero_publicacion,
            notificado,
            oc_subida,
            estado,
            observaciones
        FROM publicaciones_necesidad
        WHERE id = %s
    """, (publicacion_id,))

    publicacion = cur.fetchone()

    if not publicacion:
        cur.close()
        conn.close()

        flash("❌ La publicación no existe.", "danger")

        return redirect(
            url_for("main.publicaciones_necesidad")
        )

    # ==========================================
    # PRODUCTOS / ITEMS
    # ==========================================
    cur.execute("""
        SELECT
            id,
            cpc,
            descripcion_producto,
            cantidad,
            unidad,
            forma_pago
        FROM publicacion_items
        WHERE publicacion_id = %s
        ORDER BY id
    """, (publicacion_id,))

    items = cur.fetchall()

    # ==========================================
    # PROFORMAS
    # ==========================================
    cur.execute("""
        SELECT
            id,
            proveedor,
            ruc,
            fecha_recepcion,
            monto_proforma,
            observaciones
        FROM proformas_publicacion
        WHERE publicacion_id = %s
        ORDER BY fecha_recepcion DESC NULLS LAST, id DESC
    """, (publicacion_id,))

    proformas = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "publicaciones/publicacion_detalle.html",
        publicacion=publicacion,
        items=items,
        proformas=proformas
    )
# ==========================================
# GUARDAR PROFORMA DE PUBLICACIÓN
# ==========================================
@main.route(
    "/publicaciones_necesidad/<int:publicacion_id>/proformas/guardar",
    methods=["POST"]
)
@login_required()
def guardar_proforma_publicacion(publicacion_id):

    conn = get_connection()
    cur = conn.cursor()

    try:

        proveedor = request.form.get("proveedor")
        ruc = request.form.get("ruc")
        fecha_recepcion = request.form.get("fecha_recepcion") or None
        monto_proforma = request.form.get("monto_proforma") or None
        observaciones = request.form.get("observaciones") or None

        cur.execute("""
            INSERT INTO proformas_publicacion (
                publicacion_id,
                proveedor,
                ruc,
                fecha_recepcion,
                monto_proforma,
                observaciones,
                usuario_id
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (
            publicacion_id,
            proveedor,
            ruc,
            fecha_recepcion,
            monto_proforma,
            observaciones,
            session.get("user_id")
        ))

        conn.commit()

        flash(
            "✅ Proforma registrada correctamente.",
            "success"
        )

    except Exception as e:

        conn.rollback()

        print("ERROR GUARDANDO PROFORMA:", e)

        flash(
            f"❌ Error al registrar la proforma: {e}",
            "danger"
        )

    finally:

        cur.close()
        conn.close()

    return redirect(
        url_for(
            "main.publicacion_necesidad_detalle",
            publicacion_id=publicacion_id
        )
    )
# ==========================================
# ELIMINAR PROFORMA
# ==========================================
@main.route(
    "/proformas_publicacion/<int:proforma_id>/eliminar",
    methods=["POST"]
)
@login_required()
def eliminar_proforma_publicacion(proforma_id):

    conn = get_connection()
    cur = conn.cursor()

    try:

        cur.execute("""
            SELECT publicacion_id
            FROM proformas_publicacion
            WHERE id = %s
        """, (proforma_id,))

        row = cur.fetchone()

        if not row:

            flash(
                "❌ La proforma no existe.",
                "danger"
            )

            return redirect(
                url_for("main.publicaciones_necesidad")
            )

        publicacion_id = row[0]

        cur.execute("""
            DELETE FROM proformas_publicacion
            WHERE id = %s
        """, (proforma_id,))

        conn.commit()

        flash(
            "✅ Proforma eliminada correctamente.",
            "success"
        )

    except Exception as e:

        conn.rollback()

        print("ERROR ELIMINANDO PROFORMA:", e)

        flash(
            f"❌ Error al eliminar la proforma: {e}",
            "danger"
        )

        publicacion_id = None

    finally:

        cur.close()
        conn.close()

    if publicacion_id:

        return redirect(
            url_for(
                "main.publicacion_necesidad_detalle",
                publicacion_id=publicacion_id
            )
        )

    return redirect(
        url_for("main.publicaciones_necesidad")
    )
# ==========================================
# EDITAR PROFORMA
# ==========================================
@main.route(
    "/proformas_publicacion/<int:proforma_id>/editar",
    methods=["POST"]
)
@login_required()
def editar_proforma_publicacion(proforma_id):

    conn = get_connection()
    cur = conn.cursor()

    try:

        cur.execute("""
            SELECT publicacion_id
            FROM proformas_publicacion
            WHERE id = %s
        """, (proforma_id,))

        row = cur.fetchone()

        if not row:
            flash("❌ La proforma no existe.", "danger")

            return redirect(
                url_for("main.publicaciones_necesidad")
            )

        publicacion_id = row[0]

        proveedor = request.form.get("proveedor")
        ruc = request.form.get("ruc")
        fecha_recepcion = (
            request.form.get("fecha_recepcion") or None
        )
        monto_proforma = (
            request.form.get("monto_proforma") or None
        )
        observaciones = request.form.get("observaciones")

        cur.execute("""
            UPDATE proformas_publicacion
            SET
                proveedor = %s,
                ruc = %s,
                fecha_recepcion = %s,
                monto_proforma = %s,
                observaciones = %s
            WHERE id = %s
        """, (
            proveedor,
            ruc,
            fecha_recepcion,
            monto_proforma,
            observaciones,
            proforma_id
        ))

        conn.commit()

        flash(
            "✅ Proforma actualizada correctamente.",
            "success"
        )

    except Exception as e:

        conn.rollback()

        print("ERROR EDITANDO PROFORMA:", e)

        flash(
            f"❌ Error al editar la proforma: {e}",
            "danger"
        )

        publicacion_id = None

    finally:

        cur.close()
        conn.close()

    if publicacion_id:

        return redirect(
            url_for(
                "main.publicacion_necesidad_detalle",
                publicacion_id=publicacion_id
            )
        )

    return redirect(
        url_for("main.publicaciones_necesidad")
    )
# ==========================================
# EDITAR PUBLICACIÓN - MOSTRAR FORMULARIO
# ==========================================
@main.route("/publicaciones_necesidad/<int:publicacion_id>/editar")
@login_required()
def editar_publicacion_necesidad(publicacion_id):

    conn = get_connection()
    cur = conn.cursor()

    # Publicación
    cur.execute("""
        SELECT
            id,
            numero_solicitud,
            objeto_compra,
            fecha_publicacion,
            fecha_limite,
            encargado,
            correo,
            tipo_publicacion,
            unidad_requirente,
            codigo_publicacion,
            numero_publicacion,
            notificado,
            oc_subida,
            estado,
            observaciones,
            codigo_proceso
        FROM publicaciones_necesidad
        WHERE id = %s
    """, (publicacion_id,))

    publicacion = cur.fetchone()

    if not publicacion:
        cur.close()
        conn.close()

        flash("❌ La publicación no existe.", "danger")
        return redirect(url_for("main.publicaciones_necesidad"))

    # Items actuales
    cur.execute("""
        SELECT
            id,
            cpc,
            descripcion_producto,
            cantidad,
            unidad,
            forma_pago
        FROM publicacion_items
        WHERE publicacion_id = %s
        ORDER BY id
    """, (publicacion_id,))

    items = cur.fetchall()

    # Unidades
    cur.execute("""
        SELECT
            id,
            nombre_unidad,
            departamento_principal,
            bloque,
            CASE
                WHEN UPPER(TRIM(nombre_unidad)) IN ('DECANATO', 'SUBDECANATO')
                    AND bloque IS NOT NULL
                    AND TRIM(bloque) <> ''
                THEN nombre_unidad || ' - ' || bloque
                ELSE nombre_unidad
            END AS nombre_visible
        FROM unidades
        ORDER BY nombre_visible
    """)

    unidades = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "publicaciones/publicacion_form.html",
        publicacion=publicacion,
        items=items,
        unidades=unidades
    )
# ==========================================
# ELIMINAR PUBLICACIÓN
# ==========================================
@main.route(
    "/publicaciones_necesidad/<int:publicacion_id>/eliminar",
    methods=["POST"]
)
@login_required()
def eliminar_publicacion_necesidad(publicacion_id):

    conn = get_connection()
    cur = conn.cursor()

    try:

        # ==========================================
        # ELIMINAR PROFORMAS
        # ==========================================
        cur.execute("""
            DELETE
            FROM proformas_publicacion
            WHERE publicacion_id = %s
        """, (
            publicacion_id,
        ))

        # ==========================================
        # ELIMINAR ITEMS
        # ==========================================
        cur.execute("""
            DELETE
            FROM publicacion_items
            WHERE publicacion_id = %s
        """, (
            publicacion_id,
        ))

        # ==========================================
        # ELIMINAR PUBLICACIÓN
        # ==========================================
        cur.execute("""
            DELETE
            FROM publicaciones_necesidad
            WHERE id = %s
        """, (
            publicacion_id,
        ))

        conn.commit()

        flash(
            "✅ Publicación eliminada correctamente.",
            "success"
        )

    except Exception as e:

        conn.rollback()

        print(
            "ERROR ELIMINANDO PUBLICACIÓN:",
            e
        )

        flash(
            f"❌ Error al eliminar la publicación: {e}",
            "danger"
        )

    finally:

        cur.close()
        conn.close()

    return redirect(
        url_for(
            "main.publicaciones_necesidad"
        )
    )

# ==========================================
# VISTA PREVIA DE MATRIZ DE PUBLICACIONES
# NO MODIFICA LA BASE DE DATOS
# ==========================================
@main.route(
    "/publicaciones_necesidad/vista_previa_excel",
    methods=["GET", "POST"]
)
@login_required()
def vista_previa_publicaciones_excel():

    resumen = None
    publicaciones_preview = []
    advertencias = []
    token_importacion = None
    ruta_temporal = None

    if request.method == "POST":

        archivo = request.files.get("archivo")

        if not archivo or not archivo.filename:

            flash(
                "❌ Debe seleccionar un archivo Excel.",
                "danger"
            )

            return redirect(
                url_for(
                    "main.vista_previa_publicaciones_excel"
                )
            )

        # ==========================================
        # VALIDAR EXTENSIÓN DEL ARCHIVO
        # ==========================================
        nombre_seguro = secure_filename(
            archivo.filename
        )

        extension = os.path.splitext(
            nombre_seguro
        )[1].lower()

        if extension != ".xlsx":

            flash(
                "❌ Solo se permiten archivos Excel .xlsx.",
                "danger"
            )

            return redirect(
                url_for(
                    "main.vista_previa_publicaciones_excel"
                )
            )

        # ==========================================
        # CREAR CARPETA TEMPORAL
        # ==========================================
        carpeta_temporal = os.path.join(
            current_app.root_path,
            "static",
            "uploads",
            "importaciones_temporales"
        )

        os.makedirs(
            carpeta_temporal,
            exist_ok=True
        )

        # ==========================================
        # GENERAR UN ÚNICO TOKEN
        # ==========================================
        token_importacion = uuid.uuid4().hex

        nombre_temporal = (
            f"{token_importacion}.xlsx"
        )

        ruta_temporal = os.path.join(
            carpeta_temporal,
            nombre_temporal
        )

        try:

            # Guardar el archivo una sola vez.
            archivo.save(
                ruta_temporal
            )

            print("===================================")
            print("ARCHIVO TEMPORAL GUARDADO")
            print("TOKEN:", token_importacion)
            print("RUTA:", ruta_temporal)
            print(
                "TAMAÑO:",
                os.path.getsize(ruta_temporal),
                "bytes"
            )
            print("===================================")

            # Utilizar la función auxiliar creada antes.
            resultado = leer_excel_publicaciones(
                ruta_temporal
            )

            resumen = resultado["resumen"]

            publicaciones_preview = resultado[
                "publicaciones_preview"
            ]

            advertencias = resultado[
                "advertencias"
            ]

        except Exception as e:

            import traceback
            traceback.print_exc()

            # Eliminar el archivo si falló la lectura.
            try:

                if (
                    ruta_temporal
                    and os.path.exists(ruta_temporal)
                ):

                    os.remove(
                        ruta_temporal
                    )

            except OSError as error_archivo:

                print(
                    "No se pudo eliminar el archivo "
                    "temporal:",
                    error_archivo
                )

            token_importacion = None

            flash(
                f"❌ No fue posible analizar el Excel: {e}",
                "danger"
            )

    return render_template(
        "publicaciones/vista_previa_excel.html",
        resumen=resumen,
        publicaciones=publicaciones_preview,
        advertencias=advertencias,
        token_importacion=token_importacion
    )
# ==========================================
# IMPORTAR PUBLICACIONES A SICOP
# ==========================================
@main.route(
    "/publicaciones_necesidad/importar_confirmado",
    methods=["POST"]
)
@login_required()
def importar_publicaciones_confirmado():

    token_importacion = str(
        request.form.get("token_importacion") or ""
    ).strip()

    if not token_importacion:

        flash(
            "❌ No se encontró el archivo previamente analizado.",
            "danger"
        )

        return redirect(
            url_for(
                "main.vista_previa_publicaciones_excel"
            )
        )

    # ==========================================
    # VALIDAR TOKEN
    # ==========================================
    token_valido = (
        len(token_importacion) == 32
        and all(
            caracter in "0123456789abcdef"
            for caracter in token_importacion.lower()
        )
    )

    if not token_valido:

        flash(
            "❌ El código de importación no es válido.",
            "danger"
        )

        return redirect(
            url_for(
                "main.vista_previa_publicaciones_excel"
            )
        )

    carpeta_temporal = os.path.join(
        current_app.root_path,
        "static",
        "uploads",
        "importaciones_temporales"
    )

    ruta_temporal = os.path.join(
        carpeta_temporal,
        f"{token_importacion}.xlsx"
    )

    if not os.path.exists(ruta_temporal):

        flash(
            "❌ El archivo temporal ya no está disponible. "
            "Vuelva a cargar la matriz.",
            "danger"
        )

        return redirect(
            url_for(
                "main.vista_previa_publicaciones_excel"
            )
        )

    conn = None
    cur = None
    importacion_exitosa = False

    try:

        print("===================================")
        print("INICIANDO IMPORTACIÓN")
        print("TOKEN:", token_importacion)
        print("RUTA:", ruta_temporal)
        print(
            "TAMAÑO:",
            os.path.getsize(ruta_temporal),
            "bytes"
        )
        print("===================================")

        # Leer nuevamente el mismo archivo temporal.
        resultado = leer_excel_publicaciones(
            ruta_temporal
        )

        publicaciones = resultado[
            "publicaciones"
        ]

        advertencias = resultado[
            "advertencias"
        ]

        if not publicaciones:

            raise ValueError(
                "No se encontraron publicaciones válidas "
                "para importar."
            )

        # Las advertencias no bloquean toda la importación.
        # Las filas inválidas ya fueron excluidas por
        # leer_excel_publicaciones().
        print(
            "ADVERTENCIAS DETECTADAS:",
            len(advertencias)
        )

        conn = get_connection()
        cur = conn.cursor()

        total_publicaciones = 0
        total_items = 0

        # ==========================================
        # INSERTAR PUBLICACIONES
        # ==========================================
        for datos in publicaciones.values():

            cur.execute("""
                INSERT INTO publicaciones_necesidad (
                    numero_solicitud,
                    objeto_compra,
                    fecha_publicacion,
                    fecha_limite,
                    encargado,
                    correo,
                    tipo_publicacion,
                    unidad_requirente,
                    codigo_publicacion,
                    numero_publicacion,
                    notificado,
                    oc_subida,
                    codigo_proceso,
                    estado,
                    observaciones,
                    proformas_historicas,
                    usuario_id
                )
                VALUES (
                    %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s,
                    %s, %s
                )
                RETURNING id
            """, (
                datos["numero_solicitud"],
                datos["objeto"],
                datos["fecha_publicacion"],
                datos["fecha_limite"],
                datos["encargado"],
                datos["correo"],
                datos["tipo_publicacion"],
                datos["unidad_requirente"],
                datos["codigo_publicacion"],
                datos["numero_publicacion"],
                datos["notificado"],
                datos["oc_subida"],
                datos["codigo_proceso"],
                "FINALIZADA",
                datos["observaciones"],
                datos["proformas_historicas"],
                session.get("user_id")
            ))

            resultado_insert = cur.fetchone()

            if not resultado_insert:

                raise ValueError(
                    "No fue posible obtener el ID "
                    "de la publicación insertada."
                )

            publicacion_id = resultado_insert[0]

            total_publicaciones += 1

            # ==========================================
            # INSERTAR ÍTEMS
            # ==========================================
            for item in datos["items"]:

                cur.execute("""
                    INSERT INTO publicacion_items (
                        publicacion_id,
                        cpc,
                        descripcion_producto,
                        cantidad,
                        unidad,
                        forma_pago
                    )
                    VALUES (
                        %s, %s, %s, %s, %s, %s
                    )
                """, (
                    publicacion_id,
                    item["cpc"],
                    item["descripcion"],
                    item["cantidad"],
                    item["unidad"],
                    item["forma_pago"]
                ))

                total_items += 1

        conn.commit()

        importacion_exitosa = True

        print("===================================")
        print("IMPORTACIÓN COMPLETADA")
        print(
            "PUBLICACIONES:",
            total_publicaciones
        )
        print(
            "ÍTEMS:",
            total_items
        )
        print("===================================")

        flash(
            f"✅ Importación completada: "
            f"{total_publicaciones} publicaciones y "
            f"{total_items} ítems guardados.",
            "success"
        )

        return redirect(
            url_for(
                "main.publicaciones_necesidad"
            )
        )

    except Exception as e:

        if conn:

            conn.rollback()

        import traceback
        traceback.print_exc()

        flash(
            f"❌ No se realizó la importación: {e}",
            "danger"
        )

        return redirect(
            url_for(
                "main.vista_previa_publicaciones_excel"
            )
        )

    finally:

        if cur:

            cur.close()

        if conn:

            conn.close()

        # El archivo solo se elimina cuando la
        # importación terminó correctamente.
        if importacion_exitosa:

            try:

                if os.path.exists(ruta_temporal):

                    os.remove(
                        ruta_temporal
                    )

                    print(
                        "ARCHIVO TEMPORAL ELIMINADO:",
                        ruta_temporal
                    )

            except OSError as error_archivo:

                print(
                    "No se pudo eliminar el archivo temporal:",
                    error_archivo
                )

# ==========================================
# ACTUALIZAR PUBLICACIÓN DE NECESIDAD
# ==========================================
@main.route(
    "/publicaciones_necesidad/<int:publicacion_id>/actualizar",
    methods=["POST"]
)
@login_required()
def actualizar_publicacion_necesidad(publicacion_id):

    conn = get_connection()
    cur = conn.cursor()

    try:

        numero_solicitud = request.form.get("numero_solicitud")
        objeto_compra = request.form.get("objeto_compra")
        fecha_publicacion = request.form.get("fecha_publicacion")
        fecha_limite = request.form.get("fecha_limite") or None
        encargado = request.form.get("encargado")
        correo = request.form.get("correo")
        tipo_publicacion = request.form.get("tipo_publicacion")
        unidad_requirente = request.form.get("unidad_requirente")
        codigo_publicacion = request.form.get("codigo_publicacion")
        numero_publicacion = (
            request.form.get("numero_publicacion") or 1
        )
        estado = request.form.get("estado")
        observaciones = request.form.get("observaciones")

        notificado = (
            request.form.get("notificado") == "on"
        )

        oc_subida = (
            request.form.get("oc_subida") == "on"
        )

        # ==========================================
        # ACTUALIZAR CABECERA
        # ==========================================
        cur.execute("""
            UPDATE publicaciones_necesidad
            SET
                numero_solicitud = %s,
                objeto_compra = %s,
                fecha_publicacion = %s,
                fecha_limite = %s,
                encargado = %s,
                correo = %s,
                tipo_publicacion = %s,
                unidad_requirente = %s,
                codigo_publicacion = %s,
                numero_publicacion = %s,
                notificado = %s,
                oc_subida = %s,
                estado = %s,
                observaciones = %s
            WHERE id = %s
        """, (
            numero_solicitud,
            objeto_compra,
            fecha_publicacion,
            fecha_limite,
            encargado,
            correo,
            tipo_publicacion,
            unidad_requirente,
            codigo_publicacion,
            numero_publicacion,
            notificado,
            oc_subida,
            estado,
            observaciones,
            publicacion_id
        ))

        # ==========================================
        # BORRAR ÍTEMS ANTERIORES
        # ==========================================
        cur.execute("""
            DELETE FROM publicacion_items
            WHERE publicacion_id = %s
        """, (publicacion_id,))

        # ==========================================
        # RECIBIR ÍTEMS DEL FORMULARIO
        # ==========================================
        cpcs = request.form.getlist("item_cpc[]")

        descripciones = request.form.getlist(
            "item_descripcion[]"
        )

        cantidades = request.form.getlist(
            "item_cantidad[]"
        )

        unidades_items = request.form.getlist(
            "item_unidad[]"
        )

        formas_pago = request.form.getlist(
            "item_forma_pago[]"
        )

        # ==========================================
        # VOLVER A GUARDAR LOS ÍTEMS
        # ==========================================
        for i, descripcion in enumerate(descripciones):

            if not descripcion.strip():
                continue

            cpc = (
                cpcs[i]
                if i < len(cpcs)
                else None
            )

            cantidad = (
                cantidades[i]
                if i < len(cantidades)
                else None
            )

            unidad_item = (
                unidades_items[i]
                if i < len(unidades_items)
                else None
            )

            forma_pago = (
                formas_pago[i]
                if i < len(formas_pago)
                else None
            )

            cantidad = cantidad or None

            cur.execute("""
                INSERT INTO publicacion_items (
                    publicacion_id,
                    cpc,
                    descripcion_producto,
                    cantidad,
                    unidad,
                    forma_pago
                )
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (
                publicacion_id,
                cpc,
                descripcion,
                cantidad,
                unidad_item,
                forma_pago
            ))

        conn.commit()

        flash(
            "✅ Publicación actualizada correctamente.",
            "success"
        )

    except Exception as e:

        conn.rollback()

        print(
            "ERROR ACTUALIZANDO PUBLICACIÓN:",
            e
        )

        flash(
            f"❌ Error al actualizar la publicación: {e}",
            "danger"
        )

    finally:

        cur.close()
        conn.close()

    return redirect(
        url_for(
            "main.publicacion_necesidad_detalle",
            publicacion_id=publicacion_id
        )
    )

# ==========================================
# CONSULTA RÁPIDA DE PUBLICACIONES
# ==========================================
@main.route("/publicaciones_necesidad/consulta")
@login_required()
def consulta_publicaciones_necesidad():

    busqueda = request.args.get("busqueda", "").strip()

    resultados = []

    if busqueda:

        conn = get_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT
                p.id,
                p.codigo_publicacion,
                p.objeto_compra,
                p.unidad_requirente,
                p.fecha_publicacion,
                p.fecha_limite,
                p.numero_publicacion,
                p.estado,
                COUNT(pr.id) AS total_proformas
            FROM publicaciones_necesidad p

            LEFT JOIN publicacion_items i
                ON i.publicacion_id = p.id

            LEFT JOIN proformas_publicacion pr
                ON pr.publicacion_id = p.id

            WHERE
                p.objeto_compra ILIKE %s
                OR i.descripcion_producto ILIKE %s
                OR i.cpc ILIKE %s
                OR p.numero_solicitud ILIKE %s

            GROUP BY
                p.id,
                p.codigo_publicacion,
                p.objeto_compra,
                p.unidad_requirente,
                p.fecha_publicacion,
                p.fecha_limite,
                p.numero_publicacion,
                p.estado

            ORDER BY
                p.fecha_publicacion DESC,
                p.id DESC
        """, (
            f"%{busqueda}%",
            f"%{busqueda}%",
            f"%{busqueda}%",
            f"%{busqueda}%"
        ))

        resultados = cur.fetchall()

        cur.close()
        conn.close()

    return render_template(
        "publicaciones/consulta_publicaciones.html",
        busqueda=busqueda,
        resultados=resultados
    )
# ==========================================
# LEER MATRIZ EXCEL DE PUBLICACIONES
# ==========================================
def leer_excel_publicaciones(ruta_excel):

    libro = None

    publicaciones = {}
    advertencias = []

    filas_leidas = 0
    filas_validas = 0
    filas_invalidas = 0
    publicaciones_sin_codigo = 0
    items_detectados = 0
    filas_vacias_consecutivas = 0

    try:

        # ==========================================
        # VALIDAR ARCHIVO
        # ==========================================
        if not os.path.exists(ruta_excel):

            raise ValueError(
                "El archivo Excel no existe."
            )

        tamanio_archivo = os.path.getsize(
            ruta_excel
        )

        if tamanio_archivo <= 0:

            raise ValueError(
                "El archivo Excel está vacío."
            )

        # Los archivos XLSX son contenedores ZIP.
        with open(ruta_excel, "rb") as archivo:

            firma = archivo.read(4)

        if firma != b"PK\x03\x04":

            raise ValueError(
                "El archivo seleccionado no es un "
                "Excel .xlsx válido."
            )

        # ==========================================
        # ABRIR LIBRO
        # ==========================================
        libro = load_workbook(
            ruta_excel,
            read_only=True,
            data_only=True
        )

        if "Hoja1" not in libro.sheetnames:

            raise ValueError(
                "El archivo no contiene la hoja 'Hoja1'."
            )

        hoja = libro["Hoja1"]

        print("===================================")
        print("LEYENDO MATRIZ DE PUBLICACIONES")
        print("ARCHIVO:", ruta_excel)
        print("TAMAÑO:", tamanio_archivo, "bytes")
        print("HOJA:", hoja.title)
        print(
            "DIMENSIÓN REPORTADA:",
            hoja.max_row,
            "filas x",
            hoja.max_column,
            "columnas"
        )
        print("===================================")

        # El Excel utiliza 20 columnas: A hasta T.
        ultima_fila_analizar = min(
            hoja.max_row,
            3000
        )

        # ==========================================
        # RECORRER FILAS DEL EXCEL
        # ==========================================
        for numero_fila, fila in enumerate(
            hoja.iter_rows(
                min_row=2,
                max_row=ultima_fila_analizar,
                min_col=1,
                max_col=20,
                values_only=True
            ),
            start=2
        ):

            filas_leidas += 1

            fila_esta_vacia = all(
                valor is None
                or str(valor).strip() == ""
                for valor in fila
            )

            if fila_esta_vacia:

                filas_vacias_consecutivas += 1

                # Se detiene cuando encuentra 20 filas
                # consecutivas completamente vacías.
                if filas_vacias_consecutivas >= 20:

                    print(
                        "LECTURA FINALIZADA EN LA FILA:",
                        numero_fila
                    )

                    break

                continue

            filas_vacias_consecutivas = 0

            (
                numero_solicitud,
                objeto,
                fecha_publicacion,
                fecha_limite,
                tiempo_publicacion,
                encargado,
                correo,
                tipo_publicacion,
                cpc,
                descripcion_producto,
                cantidad,
                unidad,
                forma_pago,
                unidad_requirente,
                codigo_publicacion,
                notificado,
                manifestacion,
                oc_subida,
                codigo_proceso,
                observaciones
            ) = fila

            objeto_texto = str(
                objeto or ""
            ).strip()

            codigo_texto = str(
                codigo_publicacion or ""
            ).strip()

            observacion_texto = str(
                observaciones or ""
            ).strip()

            observacion_mayuscula = (
                observacion_texto.upper()
            )

            # ==========================================
            # DETECTAR PUBLICACIONES INVÁLIDAS
            # ==========================================
            es_invalida = (
                "NO VALE" in observacion_mayuscula
                or "PUBLICACIÓN CON ERROR"
                in observacion_mayuscula
                or "PUBLICACION CON ERROR"
                in observacion_mayuscula
            )

            if es_invalida:

                filas_invalidas += 1

                advertencias.append({
                    "fila": numero_fila,
                    "codigo":
                        codigo_texto or "SIN CÓDIGO",
                    "mensaje": (
                        "Fila marcada en la matriz como "
                        "publicación inválida o con error."
                    )
                })

                continue

            # Una publicación válida debe tener objeto.
            if not objeto_texto:

                advertencias.append({
                    "fila": numero_fila,
                    "codigo":
                        codigo_texto or "SIN CÓDIGO",
                    "mensaje":
                        "Fila sin objeto de compra."
                })

                continue

            filas_validas += 1

            # ==========================================
            # CLAVE PARA AGRUPAR LOS ÍTEMS
            # ==========================================
            if codigo_texto:

                clave = (
                    "CODIGO",
                    codigo_texto
                )

            else:

                clave = (
                    "SIN_CODIGO",
                    str(
                        numero_solicitud or ""
                    ).strip(),
                    str(
                        fecha_publicacion or ""
                    ),
                    objeto_texto
                )

            # ==========================================
            # NÚMERO DE PUBLICACIÓN
            # ==========================================
            numero_publicacion = 1

            if (
                "2DA" in observacion_mayuscula
                or "SEGUNDA" in observacion_mayuscula
            ):

                numero_publicacion = 2

            elif (
                "3RA" in observacion_mayuscula
                or "TERCERA" in observacion_mayuscula
            ):

                numero_publicacion = 3

            elif (
                "4TA" in observacion_mayuscula
                or "CUARTA" in observacion_mayuscula
            ):

                numero_publicacion = 4

            # ==========================================
            # CONVERTIR CAMPOS BOOLEANOS
            # ==========================================
            texto_notificado = str(
                notificado or ""
            ).strip().upper()

            valor_notificado = (
                texto_notificado
                in (
                    "SI",
                    "SÍ",
                    "X",
                    "1",
                    "TRUE"
                )
            )

            texto_oc_subida = str(
                oc_subida or ""
            ).strip().upper()

            valor_oc_subida = (
                texto_oc_subida
                in (
                    "SI",
                    "SÍ",
                    "X",
                    "1",
                    "TRUE"
                )
            )

            # ==========================================
            # MANIFESTACIONES HISTÓRICAS
            # ==========================================
            try:

                proformas_historicas = int(
                    float(
                        manifestacion or 0
                    )
                )

            except (TypeError, ValueError):

                proformas_historicas = 0

                advertencias.append({
                    "fila": numero_fila,
                    "codigo":
                        codigo_texto or "SIN CÓDIGO",
                    "mensaje": (
                        "El número de manifestaciones "
                        "no pudo convertirse a un número."
                    )
                })

            # ==========================================
            # CREAR CABECERA DE LA PUBLICACIÓN
            # ==========================================
            if clave not in publicaciones:

                if not codigo_texto:

                    publicaciones_sin_codigo += 1

                publicaciones[clave] = {
                    "numero_solicitud":
                        numero_solicitud,

                    "objeto":
                        objeto_texto,

                    "fecha_publicacion":
                        fecha_publicacion,

                    "fecha_limite":
                        fecha_limite,

                    "tiempo_publicacion":
                        tiempo_publicacion,

                    "encargado":
                        encargado,

                    "correo":
                        correo,

                    "tipo_publicacion":
                        tipo_publicacion,

                    "unidad_requirente":
                        unidad_requirente,

                    "codigo_publicacion":
                        codigo_texto or None,

                    "numero_publicacion":
                        numero_publicacion,

                    "notificado":
                        valor_notificado,

                    "oc_subida":
                        valor_oc_subida,

                    "codigo_proceso":
                        codigo_proceso,

                    "observaciones":
                        observacion_texto,

                    "proformas_historicas":
                        proformas_historicas,

                    "items": []
                }

            else:

                # Evita sumar las manifestaciones
                # repetidas en cada fila de producto.
                publicaciones[clave][
                    "proformas_historicas"
                ] = max(
                    publicaciones[clave][
                        "proformas_historicas"
                    ],
                    proformas_historicas
                )

                # Conserva el mayor número de publicación.
                publicaciones[clave][
                    "numero_publicacion"
                ] = max(
                    publicaciones[clave][
                        "numero_publicacion"
                    ],
                    numero_publicacion
                )

                # Si alguna fila está marcada como sí,
                # conserva el valor verdadero.
                publicaciones[clave][
                    "notificado"
                ] = (
                    publicaciones[clave]["notificado"]
                    or valor_notificado
                )

                publicaciones[clave][
                    "oc_subida"
                ] = (
                    publicaciones[clave]["oc_subida"]
                    or valor_oc_subida
                )

            # ==========================================
            # AGREGAR ÍTEM
            # ==========================================
            if descripcion_producto:

                descripcion_texto = str(
                    descripcion_producto
                ).strip()

                publicaciones[clave][
                    "items"
                ].append({
                    "cpc":
                        cpc,

                    "descripcion":
                        descripcion_texto,

                    "cantidad":
                        cantidad,

                    "unidad":
                        unidad,

                    "forma_pago":
                        forma_pago
                })

                items_detectados += 1

        # ==========================================
        # CONSTRUIR DATOS PARA VISTA PREVIA
        # ==========================================
        publicaciones_preview = []

        for datos in publicaciones.values():

            publicaciones_preview.append({
                "numero_solicitud":
                    datos["numero_solicitud"],

                "codigo_publicacion":
                    datos["codigo_publicacion"],

                "objeto":
                    datos["objeto"],

                "fecha_publicacion":
                    datos["fecha_publicacion"],

                "unidad_requirente":
                    datos["unidad_requirente"],

                "numero_publicacion":
                    datos["numero_publicacion"],

                "proformas_historicas":
                    datos["proformas_historicas"],

                "total_items":
                    len(datos["items"])
            })

        # ==========================================
        # ORDENAR POR FECHA DE PUBLICACIÓN
        # ==========================================
        def convertir_fecha_orden(valor):

            if isinstance(valor, datetime):
                return valor.date()

            if isinstance(valor, date):
                return valor

            return date.min

        publicaciones_preview.sort(
            key=lambda registro: convertir_fecha_orden(
                registro["fecha_publicacion"]
            ),
            reverse=True
        )

        # ==========================================
        # RESUMEN DE LA LECTURA
        # ==========================================
        resumen = {
            "filas_leidas":
                filas_leidas,

            "filas_validas":
                filas_validas,

            "filas_invalidas":
                filas_invalidas,

            "publicaciones_detectadas":
                len(publicaciones),

            "items_detectados":
                items_detectados,

            "publicaciones_sin_codigo":
                publicaciones_sin_codigo,

            "advertencias":
                len(advertencias)
        }

        print("===================================")
        print("LECTURA COMPLETADA")
        print(
            "PUBLICACIONES:",
            len(publicaciones)
        )
        print(
            "ÍTEMS:",
            items_detectados
        )
        print(
            "ADVERTENCIAS:",
            len(advertencias)
        )
        print("===================================")

        return {
            "publicaciones":
                publicaciones,

            "publicaciones_preview":
                publicaciones_preview,

            "advertencias":
                advertencias,

            "resumen":
                resumen
        }

    finally:

        if libro:

            libro.close()
@main.route("/api/tareas/<int:tarea_id>/certificacion")
@login_required()
def api_tarea_certificacion(tarea_id):

    conn = get_connection()
    cur = conn.cursor()

    try:

        cur.execute("""
            SELECT
                numero_certificacion,
                fecha_certificacion,
                valor_certificacion,
                certificacion_plurianual
            FROM tareas
            WHERE id = %s
        """, (tarea_id,))

        tarea = cur.fetchone()

        if not tarea:
            return {
                "ok": False,
                "mensaje": "No se encontró la tarea."
            }, 404


        cur.execute("""
            SELECT
                anio,
                monto
            FROM certificaciones_plurianuales
            WHERE tarea_id = %s
            ORDER BY anio
        """, (tarea_id,))

        plurianuales = cur.fetchall()


        return {
            "ok": True,

            "certificacion": {
                "numero": tarea[0],
                "fecha": (
                    tarea[1].isoformat()
                    if tarea[1] else None
                ),
                "valor": float(tarea[2] or 0),
                "plurianual": bool(tarea[3])
            },

            "anios": [
                {
                    "anio": fila[0],
                    "monto": float(fila[1] or 0)
                }
                for fila in plurianuales
            ]
        }

    finally:
        cur.close()
        conn.close()
# ============================================================
# MEMORANDO - OFERTAS RECIBIDAS EN PUBLICACIÓN
# ============================================================
@main.route(
    "/publicaciones_necesidad/<int:publicacion_id>/memorando-ofertas"
)
@login_required()
def memorando_ofertas_publicacion(publicacion_id):

    conn = get_connection()
    cur = conn.cursor()

    try:

        # ========================================================
        # 1. DATOS DE LA PUBLICACIÓN
        # ========================================================
        cur.execute("""
            SELECT
                id,
                numero_solicitud,
                objeto_compra,
                fecha_publicacion,
                fecha_limite,
                encargado,
                correo,
                tipo_publicacion,
                unidad_requirente,
                codigo_publicacion,
                numero_publicacion,
                estado,
                observaciones,
                codigo_proceso
            FROM publicaciones_necesidad
            WHERE id = %s
        """, (publicacion_id,))

        publicacion = cur.fetchone()

        if not publicacion:
            return "La publicación no existe.", 404

        # ========================================================
        # HISTORIAL DE PUBLICACIONES DEL MISMO REQUERIMIENTO
        # ========================================================

        cur.execute("""
            SELECT
                p.id,
                p.tipo_publicacion,
                p.codigo_publicacion,
                p.numero_publicacion,
                p.fecha_publicacion,
                p.fecha_limite,
                COUNT(pr.id) AS total_proformas
            FROM publicaciones_necesidad p

            LEFT JOIN proformas_publicacion pr
                ON pr.publicacion_id = p.id

            WHERE
                UPPER(TRIM(p.objeto_compra)) =
                UPPER(TRIM(%s))

                AND
                UPPER(TRIM(p.unidad_requirente)) =
                UPPER(TRIM(%s))

            GROUP BY
                p.id,
                p.tipo_publicacion,
                p.codigo_publicacion,
                p.numero_publicacion,
                p.fecha_publicacion,
                p.fecha_limite

            ORDER BY
                p.fecha_publicacion ASC,
                p.id ASC
        """, (
            publicacion[2],
            publicacion[8]
        ))

        historial_publicaciones = cur.fetchall()

        # ========================================================
        # 2. PROFORMAS RECIBIDAS
        # ========================================================
        cur.execute("""
            SELECT
                proveedor,
                ruc,
                fecha_recepcion,
                monto_proforma,
                observaciones
            FROM proformas_publicacion
            WHERE publicacion_id = %s
            ORDER BY fecha_recepcion ASC NULLS LAST, id ASC
        """, (publicacion_id,))

        proformas = cur.fetchall()


        # ========================================================
        # 3. MOSTRAR MEMORANDO
        # ========================================================
        return render_template(
            "publicaciones/memorando_ofertas.html",
            publicacion=publicacion,
            proformas=proformas,
            total_ofertas=len(proformas),
            historial_publicaciones=historial_publicaciones
        )

    finally:

        cur.close()
        conn.close()